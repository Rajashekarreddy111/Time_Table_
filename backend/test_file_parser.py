from io import BytesIO

import openpyxl

from services.file_parser import parse_tabular_upload


def _make_grouped_block(ws, start_row: int, year_value: int, subject_rows: list[tuple[int, int, int, int, int, int]]) -> None:
    ws.cell(start_row, 1, "YEAR")
    ws.cell(start_row, 2, "SUBJECT")
    ws.cell(start_row, 3, "C1")
    ws.cell(start_row, 6, "C2")

    ws.cell(start_row + 1, 3, "NO OF HOURS")
    ws.cell(start_row + 1, 4, "FACULTY-ID")
    ws.cell(start_row + 1, 5, "CONTINUOUS HOURS")
    ws.cell(start_row + 1, 6, "NO OF HOURS")
    ws.cell(start_row + 1, 7, "FACULTY-ID")
    ws.cell(start_row + 1, 8, "CONTINUOUS HOURS")

    row_ptr = start_row + 2
    for subject_id, c1_hours, c1_faculty, c2_hours, c2_faculty, continuous_hours in subject_rows:
        ws.cell(row_ptr, 1, year_value)
        ws.cell(row_ptr, 2, subject_id)
        ws.cell(row_ptr, 3, c1_hours)
        ws.cell(row_ptr, 4, c1_faculty)
        ws.cell(row_ptr, 5, continuous_hours)
        ws.cell(row_ptr, 6, c2_hours)
        ws.cell(row_ptr, 7, c2_faculty)
        ws.cell(row_ptr, 8, continuous_hours)
        row_ptr += 1


def test_parse_grouped_main_timetable_reads_multiple_year_blocks() -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    _make_grouped_block(
        worksheet,
        1,
        2,
        [
            (2, 7, 5, 7, 5, 2),
            (3, 6, 40, 6, 40, 2),
        ],
    )
    _make_grouped_block(
        worksheet,
        8,
        3,
        [
            (12, 5, 26, 5, 36, 2),
        ],
    )

    stream = BytesIO()
    workbook.save(stream)

    frame = parse_tabular_upload("main.xlsx", stream.getvalue())

    assert len(frame) == 3
    assert {int(v) for v in frame["YEAR"].tolist()} == {2, 3}
    assert {int(v) for v in frame["SUBJECT_ID"].tolist()} == {2, 3, 12}


def test_parse_grouped_main_timetable_accepts_leading_blank_rows() -> None:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    _make_grouped_block(
        worksheet,
        4,
        2,
        [
            (2, 7, 5, 7, 5, 2),
        ],
    )

    stream = BytesIO()
    workbook.save(stream)

    frame = parse_tabular_upload("main.xlsx", stream.getvalue())

    assert len(frame) == 1
    assert int(frame.iloc[0]["YEAR"]) == 2
    assert int(frame.iloc[0]["SUBJECT_ID"]) == 2

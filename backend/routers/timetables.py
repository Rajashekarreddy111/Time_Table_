from typing import Any
import copy

from fastapi import APIRouter, HTTPException
from openpyxl import Workbook

from models.schemas import GenerateTimetableRequest, GenerateTimetableResponse
from services.timetable_generator import (
    DAYS,
    PERIODS,
    _build_faculty_schedule_details_from_section_grids,
    _build_faculty_workload_workbook_from_details,
    _build_section_timetables_workbook_from_schedule_map,
    _build_section_timetables_workbook,
    _build_room_timetables_workbook_from_schedule_map,
    _encode_workbook,
    generate_timetable,
)
from storage.memory_store import store

router = APIRouter(tags=["timetables"])


def _normalize_subject_token(value: Any) -> str:
    token = str(value or "").strip()
    if token.endswith(".0"):
        token = token[:-2]
    return token


def _get_global_period_config() -> list[dict] | None:
    payload = store.get_scoped_mapping("period_configuration", "global")
    if payload and payload.get("rows"):
        return payload["rows"]
    return None


def _build_subject_name_map() -> dict[str, str]:
    mapping: dict[str, str] = {}
    payload = store.get_scoped_mapping("subject_id_mapping", "global")
    if not payload:
        return mapping
    for row in payload.get("rows", []):
        sid = _normalize_subject_token(
            row.get("subject_id", "") or row.get("subject id", "") or row.get("id", "")
        )
        sname = str(
            row.get("subject_name", "") or row.get("subject name", "") or row.get("subject", "")
        ).strip()
        if sid:
            mapping[sid] = sname or sid
    return mapping


def _enrich_subject_names(record: dict[str, Any]) -> dict[str, Any]:
    subject_map = _build_subject_name_map()
    if not subject_map:
        return record

    def enrich_grid(grid: dict[str, list[Any]] | None) -> None:
        if not isinstance(grid, dict):
            return
        for _day, slots in grid.items():
            if not isinstance(slots, list):
                continue
            for slot in slots:
                if not isinstance(slot, dict):
                    continue
                sid = _normalize_subject_token(slot.get("subjectId") or slot.get("subject"))
                if not sid:
                    continue
                subject_name = subject_map.get(sid, sid)
                slot["subjectId"] = sid
                slot["subjectName"] = subject_name
                slot["subject"] = subject_name

    if isinstance(record.get("allGrids"), dict):
        for grid in record["allGrids"].values():
            enrich_grid(grid)
    else:
        enrich_grid(record.get("grid"))

    faculty_workloads = record.get("facultyWorkloads")
    if isinstance(faculty_workloads, dict):
        for _faculty, day_map in faculty_workloads.items():
            if not isinstance(day_map, dict):
                continue
            for _day, slots in day_map.items():
                if not isinstance(slots, list):
                    continue
                for idx, slot in enumerate(slots):
                    if not isinstance(slot, str) or not slot.strip():
                        continue
                    head, sep, tail = slot.partition("(")
                    sid = _normalize_subject_token(head)
                    if not sid:
                        continue
                    subject_name = subject_map.get(sid)
                    if not subject_name:
                        continue
                    slots[idx] = f"{subject_name} ({tail}" if sep else subject_name

    for key in ("sharedClasses", "constraintViolations", "unscheduledSubjects"):
        items = record.get(key)
        if not isinstance(items, list):
            continue
        for item in items:
            if not isinstance(item, dict):
                continue
            sid = _normalize_subject_token(item.get("subject_id", ""))
            if not sid:
                continue
            item["subject_id"] = sid
            item["subject_name"] = subject_map.get(sid, sid)
    try:
        _refresh_generated_workbooks(record)
    except Exception:
        # Keep timetable listing available even if workbook regeneration fails
        # for a particular record. Dedicated download endpoints can still rebuild
        # files later after layout issues are fixed.
        pass
    return record


def _refresh_generated_workbooks(record: dict[str, Any]) -> None:
    year = str(record.get("year", "")).strip()
    section = str(record.get("section", "")).strip()
    all_grids = record.get("allGrids")
    single_grid = record.get("grid")
    timetable_metadata = record.get("timetableMetadata")
    if not year:
        return

    generated_files = dict(record.get("generatedFiles") or {})

    schedules: dict[tuple[str, str], dict[str, dict[int, dict[str, Any] | None]]] = {}
    if isinstance(all_grids, dict) and all_grids:
        for sec, grid in all_grids.items():
            if not isinstance(grid, dict):
                continue
            schedules[(year, sec)] = {
                day: {
                    period: (grid.get(day, [None] * len(PERIODS))[period - 1] if len(grid.get(day, [])) >= period else None)
                    for period in PERIODS
                }
                for day in DAYS
            }
    elif isinstance(single_grid, dict) and section:
        schedules[(year, section)] = {
            day: {
                period: (single_grid.get(day, [None] * len(PERIODS))[period - 1] if len(single_grid.get(day, [])) >= period else None)
                for period in PERIODS
            }
            for day in DAYS
        }

    if schedules:
        sections = [sec for _, sec in schedules.keys()]
        generated_files["sectionTimetables"] = _encode_workbook(
            "section_timetables.xlsx",
            _build_section_timetables_workbook(year, sections, schedules, timetable_metadata, _get_global_period_config()),
        )

    if schedules:
        section_grids = {
            (record_year, record_section): {
                day: [
                    schedules[(record_year, record_section)][day][period]
                    for period in PERIODS
                ]
                for day in DAYS
            }
            for record_year, record_section in schedules
        }
        faculty_schedules = _build_faculty_schedule_details_from_section_grids(section_grids)
        if faculty_schedules:
            generated_files["facultyWorkload"] = _encode_workbook(
                "faculty_workload.xlsx",
                _build_faculty_workload_workbook_from_details(faculty_schedules, timetable_metadata, _get_global_period_config()),
            )

    if generated_files:
        record["generatedFiles"] = generated_files


def _grid_has_data(grid: dict[str, list[Any]]) -> bool:
    """Check if a grid has any non-null data."""
    if not isinstance(grid, dict):
        return False
    for day_slots in grid.values():
        if isinstance(day_slots, list) and any(slot is not None for slot in day_slots):
            return True
    return False


def _extract_section_grids_from_record(
    record: dict[str, Any],
    require_data: bool = True,
) -> dict[tuple[str, str], dict[str, list[Any]]]:
    year = str(record.get("year", "")).strip()
    section = str(record.get("section", "")).strip()
    result: dict[tuple[str, str], dict[str, list[Any]]] = {}
    all_grids = record.get("allGrids")

    if isinstance(all_grids, dict) and all_grids:
        for sec, grid in all_grids.items():
            if isinstance(grid, dict) and (not require_data or _grid_has_data(grid)):
                result[(year, str(sec).strip())] = grid
        return result
    single_grid = record.get("grid")
    if year and section and isinstance(single_grid, dict) and (not require_data or _grid_has_data(single_grid)):
        result[(year, section)] = single_grid
    return result


def _latest_section_grids(
    records: list[dict[str, Any]],
    include_invalid: bool = False,
    require_data: bool = True,
) -> dict[tuple[str, str], dict[str, list[Any]]]:
    latest: dict[tuple[str, str], dict[str, list[Any]]] = {}
    for raw_record in records:
        record = _enrich_subject_names(raw_record)
        if not include_invalid and record.get("hasValidTimetable") is False:
            continue
        for key, grid in _extract_section_grids_from_record(record, require_data).items():
            if key not in latest:
                latest[key] = grid
    return latest


def _resolve_download_metadata(records: list[dict[str, Any]]) -> dict[str, Any] | None:
    for raw_record in records:
        metadata = raw_record.get("timetableMetadata")
        if isinstance(metadata, dict) and any(
            str(metadata.get(key) or "").strip()
            for key in ("academicYear", "semester", "withEffectFrom")
        ):
            return metadata
    return None


def _section_schedules_from_grids(
    section_grids: dict[tuple[str, str], dict[str, list[Any]]],
) -> dict[tuple[str, str], dict[str, dict[int, dict[str, Any] | None]]]:
    return {
        (year, section): {
            day: {
                period: (grid.get(day, [None] * len(PERIODS))[period - 1] if len(grid.get(day, [])) >= period else None)
                for period in PERIODS
            }
            for day in DAYS
        }
        for (year, section), grid in section_grids.items()
    }


def _latest_room_grids(
    records: list[dict[str, Any]],
    include_invalid: bool = False,
    require_data: bool = True,
) -> dict[str, dict[str, list[Any]]]:
    latest: dict[str, dict[str, list[Any]]] = {}
    for raw_record in records:
        record = _enrich_subject_names(raw_record)
        if not include_invalid and record.get("hasValidTimetable") is False:
            continue
        grids = record.get("roomGrids")
        if isinstance(grids, dict) and grids:
            for room, grid in grids.items():
                if isinstance(grid, dict) and (not require_data or _grid_has_data(grid)):
                    if room not in latest:
                        latest[room] = copy.deepcopy(grid)
                    else:
                        for day, slots in grid.items():
                            if day not in latest[room]:
                                latest[room][day] = copy.deepcopy(slots)
                                continue
                            
                            for idx, slot in enumerate(slots):
                                if slot is not None:
                                    if len(latest[room][day]) <= idx:
                                        latest[room][day].extend([None] * (idx - len(latest[room][day]) + 1))
                                    current = latest[room][day][idx]
                                    if current is None:
                                        latest[room][day][idx] = copy.deepcopy(slot)
                                    else:
                                        y1 = str(current.get("year") or "").replace(" Year", "").strip()
                                        y2 = str(slot.get("year") or "").replace(" Year", "").strip()
                                        years = {y.strip() for y in y1.split(",") if y.strip()}
                                        if y2:
                                            years.update({y.strip() for y in y2.split(",") if y.strip()})
                                            
                                        s1 = str(current.get("section") or "").strip()
                                        s2 = str(slot.get("section") or "").strip()
                                        sections = {s.strip() for s in s1.split(",") if s.strip()}
                                        if s2:
                                            sections.update({s.strip() for s in s2.split(",") if s.strip()})
                                            
                                        current["year"] = ", ".join(filter(None, sorted(list(years))))
                                        current["section"] = ", ".join(filter(None, sorted(list(sections))))
    return latest


def _room_schedules_from_grids(
    room_grids: dict[str, dict[str, list[Any]]],
) -> dict[str, dict[str, dict[int, dict[str, Any] | None]]]:
    return {
        room: {
            day: {
                period: (grid.get(day, [None] * len(PERIODS))[period - 1] if len(grid.get(day, [])) >= period else None)
                for period in PERIODS
            }
            for day in DAYS
        }
        for room, grid in room_grids.items()
    }


@router.post("/timetables/generate", response_model=GenerateTimetableResponse)
async def create_timetable(payload: GenerateTimetableRequest):
    result = generate_timetable(payload, store)
    return GenerateTimetableResponse(**result)


@router.post("/timetables/feasibility")
async def check_timetable_feasibility(payload: GenerateTimetableRequest):
    return generate_timetable(payload, store, precheck_only=True)


@router.get("/timetables")
async def list_generated_timetables():
    items = [_enrich_subject_names(item) for item in store.list_timetables()]
    return {"items": items}


@router.get("/timetables/all-sections-workbook")
async def get_all_sections_workbook():
    records = store.list_timetables()
    metadata = _resolve_download_metadata(records)
    section_grids = _latest_section_grids(records)

    if not section_grids:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "No Timetables"
        ws.append(["No Timetables Found"])
        ws.append([f"Total records in database: {len(records)}"])
        ws.append(["Please generate timetables first before downloading."])
    else:
        schedules = _section_schedules_from_grids(section_grids)
        workbook = _build_section_timetables_workbook_from_schedule_map(schedules, metadata, _get_global_period_config())

    return _encode_workbook("All_Class_Timetables_Format.xlsx", workbook)


@router.get("/timetables/all-rooms-workbook")
async def get_all_rooms_workbook(room: str | None = None):
    records = store.list_timetables()
    metadata = _resolve_download_metadata(records)
    room_grids = _latest_room_grids(records)

    if room:
        room_grids = {r: g for r, g in room_grids.items() if r == room}

    if not room_grids:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "No Timetables"
        ws.append(["No Timetables Found"])
        if room:
            ws.append([f"No data for room: {room}"])
        ws.append([f"Total records in database: {len(records)}"])
        ws.append(["Please generate timetables first before downloading."])
        file_name = f"Room_Timetable_{room.replace(' ', '_')}.xlsx" if room else "All_Room_Timetables_Format.xlsx"
        return _encode_workbook(file_name, workbook)
    else:
        schedules = _room_schedules_from_grids(room_grids)
        workbook = _build_room_timetables_workbook_from_schedule_map(schedules, metadata, _get_global_period_config())
        file_name = f"Room_Timetable_{room.replace(' ', '_')}.xlsx" if room else "All_Room_Timetables_Format.xlsx"
        return _encode_workbook(file_name, workbook)


@router.get("/timetables/{timetable_id}")
async def get_generated_timetable(timetable_id: str):
    payload = store.get_timetable(timetable_id)
    if not payload:
        raise HTTPException(
            status_code=404,
            detail={"error": "NotFound", "message": "Timetable not found", "details": []},
        )
    return _enrich_subject_names(payload)


@router.get("/timetables/{timetable_id}/section-workbook")
async def get_section_workbook(timetable_id: str, section: str):
    payload = store.get_timetable(timetable_id)
    if not payload:
        raise HTTPException(
            status_code=404,
            detail={"error": "NotFound", "message": "Timetable not found", "details": []},
        )
    record = _enrich_subject_names(payload)
    grids = _extract_section_grids_from_record(record)
    year = str(record.get("year", "")).strip()
    target_section = str(section).strip()
    target_grid = grids.get((year, target_section))
    if not target_grid:
        raise HTTPException(
            status_code=404,
            detail={"error": "NotFound", "message": "Section timetable not found", "details": []},
        )
    schedules = _section_schedules_from_grids({(year, target_section): target_grid})
    return _encode_workbook(
        f"Timetable_{year.replace(' ', '_')}_{target_section}_Format.xlsx",
        _build_section_timetables_workbook(year, [target_section], schedules, record.get("timetableMetadata"), _get_global_period_config()),
    )


@router.get("/timetables/all-sections-workbook-legacy-disabled", include_in_schema=False)
async def get_all_sections_workbook_legacy_disabled():
    records = store.list_timetables()
    metadata = _resolve_download_metadata(records)

    # ✅ Get valid grids only (remove include_invalid and require_data parameters)
    section_grids = _latest_section_grids(records)

    # ✅ Build workbook safely
    if not section_grids:
        workbook = Workbook()
        ws = workbook.active
        ws.title = "No Timetables"
        ws.append(["No Timetables Found"])
        ws.append([f"Total records in database: {len(records)}"])
        ws.append(["Please generate timetables first before downloading."])
    else:
        schedules = _section_schedules_from_grids(section_grids)
        workbook = _build_section_timetables_workbook_from_schedule_map(
            schedules, metadata, period_config=_get_global_period_config()
        )

    # ✅ CRITICAL: Write to stream properly
    return _encode_workbook("All_Class_Timetables_Format.xlsx", workbook)

    # ✅ Return as downloadable file


@router.get("/faculty-workloads/workbook")
async def get_faculty_workload_workbook(facultyName: str | None = None):
    records = store.list_timetables()
    metadata = _resolve_download_metadata(records)
    section_grids = _latest_section_grids(records)
    if not section_grids:
        raise HTTPException(
            status_code=404,
            detail={"error": "NotFound", "message": "No faculty workloads available", "details": []},
        )
    faculty_schedules = _build_faculty_schedule_details_from_section_grids(section_grids)
    selected_name = str(facultyName or "").strip()
    if selected_name:
        schedule = faculty_schedules.get(selected_name)
        if not schedule:
            raise HTTPException(
                status_code=404,
                detail={"error": "NotFound", "message": "Faculty workload not found", "details": []},
            )
        faculty_schedules = {selected_name: schedule}
        file_name = f"Workload_{selected_name.replace(' ', '_')}_Format.xlsx"
    else:
        file_name = "All_Faculty_Workloads_Format.xlsx"

    return _encode_workbook(
        file_name,
        _build_faculty_workload_workbook_from_details(
            faculty_schedules,
            metadata,
            period_config=_get_global_period_config()
        ),
    )


@router.delete("/timetables/{timetable_id}")
async def delete_timetable(timetable_id: str):
    success = store.delete_timetable(timetable_id)
    if not success:
        raise HTTPException(status_code=404, detail="Timetable not found")
    store.delete_occupancy_by_source(timetable_id)
    return {"message": "Timetable deleted successfully"}


@router.delete("/timetables")
async def reset_all_timetables():
    deleted_count = store.delete_all_timetables()
    store.delete_occupancy_by_source(None)
    return {"message": f"Deleted {deleted_count} timetables", "deletedCount": deleted_count}





from fastapi.responses import Response
import base64

@router.get("/timetables/{timetable_id}/room-workbook")
async def get_room_workbook(timetable_id: str):
    payload = store.get_timetable(timetable_id)
    if not payload or not payload.get("generatedFiles", {}).get("roomTimetables"):
        raise HTTPException(status_code=404, detail="Room workbook not found")
    
    return payload["generatedFiles"]["roomTimetables"]

@router.get("/timetables/{timetable_id}/constraint-report-workbook")
async def get_constraint_report_workbook(timetable_id: str):
    payload = store.get_timetable(timetable_id)
    if not payload or not payload.get("generatedFiles", {}).get("constraintReport"):
        raise HTTPException(status_code=404, detail="Constraint report not found")
        
    return payload["generatedFiles"]["constraintReport"]

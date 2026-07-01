import sys
import unittest
from pathlib import Path

# Disable MongoDB for tests to ensure full test isolation and socket stability
import pymongo
pymongo.MongoClient = lambda *args, **kwargs: exec("raise(pymongo.errors.ConnectionFailure('MongoDB disabled for tests'))") or None
# Or define a simple helper function
def _disable_mongo(*args, **kwargs):
    raise pymongo.errors.ConnectionFailure("MongoDB disabled for tests")
pymongo.MongoClient = _disable_mongo


ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from models.schemas import (  # noqa: E402
    FacultyWeeklyAvailabilityEntry,
    GenerateTimetableRequest,
    ManualEntryMode,
    ManualLabEntry,
    SharedClassEntry,
)
from services.timetable_generator import (  # noqa: E402
    DAYS,
    Requirement,
    _allocate_classrooms_to_schedule,
    _build_daily_subject_counts,
    _build_session_adjacency,
    _build_room_inventory,
    _build_section_strength_map,
    _choose_faculty_for_slot,
    _enumerate_slot_candidates,
    _sections_are_free,
    _subject_daily_limit_penalty,
    generate_timetable,
)
from storage.memory_store import MemoryStore  # noqa: E402


def _count_filled_slots(grid: dict[str, list[dict | None]]) -> int:
    return sum(1 for day in DAYS for entry in grid[day] if entry is not None)


def _collect_subject_slots(grid: dict[str, list[dict | None]], subject_id: str) -> list[tuple[str, int]]:
    slots: list[tuple[str, int]] = []
    for day in DAYS:
        for period_index, entry in enumerate(grid[day], start=1):
            if entry and entry.get("subject") == subject_id:
                slots.append((day, period_index))
    return slots


def _metadata() -> dict[str, object]:
    return {
        "academicYear": "2026-2027",
        "semester": 2,
        "withEffectFrom": "2026-06-01",
    }


class TimetableSolverTests(unittest.TestCase):
    def test_faculty_cannot_take_consecutive_different_section_within_session(self) -> None:
        requirement = Requirement(
            subject_id="SUB2",
            faculty_id="F1",
            faculty_options=("F1",),
            faculty_team=(),
            sections=("B",),
            hours=1,
            min_consecutive_hours=1,
            max_consecutive_hours=1,
            shared=False,
        )
        selected = _choose_faculty_for_slot(
            requirement,
            "Monday",
            2,
            1,
            {"F1": set()},
            {"F1": {day: {1, 2, 3, 4, 5, 6, 7} for day in DAYS}},
            [1, 2, 3, 4, 5, 6, 7],
            {"F1": {("Monday", 1): ("A",)}},
            _build_session_adjacency([(1, 2), (3, 4), (5, 6, 7)]),
        )

        self.assertEqual(selected, "F1")


    def test_faculty_can_take_same_section_consecutively_and_after_break(self) -> None:
        requirement_same = Requirement(
            subject_id="SUB2",
            faculty_id="F1",
            faculty_options=("F1",),
            faculty_team=(),
            sections=("A",),
            hours=1,
            min_consecutive_hours=1,
            max_consecutive_hours=1,
            shared=False,
        )
        availability = {"F1": {day: {1, 2, 3, 4, 5, 6, 7} for day in DAYS}}
        faculty_slots = {"F1": {("Monday", 1): ("A",)}}
        adjacency = _build_session_adjacency([(1, 2), (3, 4), (5, 6, 7)])

        same_section_selected = _choose_faculty_for_slot(
            requirement_same,
            "Monday",
            2,
            1,
            {"F1": set()},
            availability,
            [1, 2, 3, 4, 5, 6, 7],
            faculty_slots,
            adjacency,
        )
        after_break_selected = _choose_faculty_for_slot(
            Requirement(
                subject_id="SUB3",
                faculty_id="F1",
                faculty_options=("F1",),
                faculty_team=(),
                sections=("B",),
                hours=1,
                min_consecutive_hours=1,
                max_consecutive_hours=1,
                shared=False,
            ),
            "Monday",
            3,
            1,
            {"F1": set()},
            availability,
            [1, 2, 3, 4, 5, 6, 7],
            faculty_slots,
            adjacency,
        )

        self.assertEqual("F1", same_section_selected)
        self.assertEqual("F1", after_break_selected)

    def test_single_section_keeps_same_room_within_session(self) -> None:
        year = "2nd Year"
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "SUB1",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "A")]["Monday"][2] = {
            "subject": "SUB2",
            "isLab": False,
            "sharedSections": [],
        }
        session_log = [
            {"day": "Monday", "periods": [1], "sections": ["A"], "isLab": False},
            {"day": "Monday", "periods": [2], "sections": ["A"], "isLab": False},
        ]
        constraint_violations: list[dict] = []

        _allocate_classrooms_to_schedule(
            year,
            ["A"],
            schedules,
            session_log,
            ["C101", "C102"],
            constraint_violations,
            [1, 2, 3, 4, 5, 6, 7],
            [(1, 2), (3, 4), (5, 6, 7)],
            section_strength_map={"A": 30},
        )

        self.assertEqual([], constraint_violations)
        self.assertEqual(
            schedules[(year, "A")]["Monday"][1].get("classroom"),
            schedules[(year, "A")]["Monday"][2].get("classroom"),
        )

    def test_partial_session_shared_class_uses_same_room_only_for_shared_period(self) -> None:
        year = "2nd Year"
        sections = ["A", "B"]
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
            (year, "B"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "A1",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "B")]["Monday"][1] = {
            "subject": "B1",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "A")]["Monday"][2] = {
            "subject": "SHARED",
            "isLab": False,
            "sharedSections": ["A", "B"],
        }
        schedules[(year, "B")]["Monday"][2] = {
            "subject": "SHARED",
            "isLab": False,
            "sharedSections": ["A", "B"],
        }
        session_log = [
            {"day": "Monday", "periods": [1], "sections": ["A"], "isLab": False},
            {"day": "Monday", "periods": [1], "sections": ["B"], "isLab": False},
            {"day": "Monday", "periods": [2], "sections": ["A", "B"], "isLab": False},
        ]
        constraint_violations: list[dict] = []

        _allocate_classrooms_to_schedule(
            year,
            sections,
            schedules,
            session_log,
            ["C101", "C102", "C103"],
            constraint_violations,
            [1, 2, 3, 4, 5, 6, 7],
            [(1, 2), (3, 4), (5, 6, 7)],
            section_strength_map={"A": 30, "B": 30},
        )

        self.assertEqual([], constraint_violations)
        self.assertNotEqual(
            schedules[(year, "A")]["Monday"][1].get("classroom"),
            schedules[(year, "B")]["Monday"][1].get("classroom"),
        )
        self.assertEqual(
            schedules[(year, "A")]["Monday"][2].get("classroom"),
            schedules[(year, "B")]["Monday"][2].get("classroom"),
        )

    def test_strength_can_push_allocation_to_lab_room_when_needed(self) -> None:
        year = "2nd Year"
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "BIG",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "A")]["Monday"][2] = {
            "subject": "BIG2",
            "isLab": False,
            "sharedSections": [],
        }
        session_log = [
            {"day": "Monday", "periods": [1], "sections": ["A"], "isLab": False},
            {"day": "Monday", "periods": [2], "sections": ["A"], "isLab": False},
        ]
        constraint_violations: list[dict] = []

        _allocate_classrooms_to_schedule(
            year,
            ["A"],
            schedules,
            session_log,
            ["C101", "L201"],
            constraint_violations,
            [1, 2, 3, 4, 5, 6, 7],
            [(1, 2), (3, 4), (5, 6, 7)],
            lab_room_names={"L201"},
            room_capacity_map={"C101": 60, "L201": 80},
            section_strength_map={"2nd Year|A": 75},
        )

        self.assertEqual([], constraint_violations)
        self.assertEqual("L201", schedules[(year, "A")]["Monday"][1].get("classroom"))
        self.assertEqual("L201", schedules[(year, "A")]["Monday"][2].get("classroom"))

    def test_fixed_classroom_blocks_bypass_auto_allocation_and_reserve_named_room(self) -> None:
        year = "2nd Year"
        sections = ["A", "B"]
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
            (year, "B"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "A_FIXED",
            "subjectId": "A_FIXED",
            "facultyId": "FA",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "A")]["Monday"][2] = {
            "subject": "A_BLANK",
            "subjectId": "A_BLANK",
            "facultyId": "FA",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "B")]["Monday"][1] = {
            "subject": "B_AUTO",
            "subjectId": "B_AUTO",
            "facultyId": "FB",
            "isLab": False,
            "sharedSections": [],
        }
        session_log = [
            {"day": "Monday", "periods": [1], "sections": ["A"], "isLab": False},
            {"day": "Monday", "periods": [2], "sections": ["A"], "isLab": False},
            {"day": "Monday", "periods": [1], "sections": ["B"], "isLab": False},
        ]
        constraint_violations: list[dict] = []

        _allocate_classrooms_to_schedule(
            year,
            sections,
            schedules,
            session_log,
            ["Seminar Hall", "C101"],
            constraint_violations,
            [1, 2, 3, 4, 5, 6, 7],
            [(1, 2), (3, 4), (5, 6, 7)],
            fixed_classroom_blocks={
                ("A", "Monday", 1): "Seminar Hall",
                ("A", "Monday", 2): "",
            },
        )

        self.assertEqual([], constraint_violations)
        self.assertEqual("Seminar Hall", schedules[(year, "A")]["Monday"][1].get("classroom"))
        self.assertEqual("", schedules[(year, "A")]["Monday"][2].get("classroom"))
        self.assertEqual("C101", schedules[(year, "B")]["Monday"][1].get("classroom"))

    def test_mixed_lab_session_uses_lab_for_entire_session_and_frees_classroom(self) -> None:
        year = "2nd Year"
        sections = ["A", "B"]
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
            (year, "B"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "A_THEORY",
            "subjectId": "A_THEORY",
            "facultyId": "FA",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "A")]["Monday"][2] = {
            "subject": "A_LAB",
            "subjectId": "A_LAB",
            "facultyId": "FA",
            "isLab": True,
            "venue": "L201",
            "labRoom": "L201",
            "sharedSections": [],
        }
        schedules[(year, "B")]["Monday"][1] = {
            "subject": "B_THEORY",
            "subjectId": "B_THEORY",
            "facultyId": "FB",
            "isLab": False,
            "sharedSections": [],
        }
        session_log = [
            {"day": "Monday", "periods": [1], "sections": ["A"], "isLab": False},
            {"day": "Monday", "periods": [2], "sections": ["A"], "isLab": True, "venue": "L201"},
            {"day": "Monday", "periods": [1], "sections": ["B"], "isLab": False},
        ]
        constraint_violations: list[dict] = []

        _allocate_classrooms_to_schedule(
            year,
            sections,
            schedules,
            session_log,
            ["C101", "L201"],
            constraint_violations,
            [1, 2, 3, 4, 5, 6, 7],
            [(1, 2), (3, 4), (5, 6, 7)],
            lab_room_names={"L201"},
            section_strength_map={"A": 30, "B": 30},
        )

        self.assertEqual([], constraint_violations)
        self.assertEqual("L201", schedules[(year, "A")]["Monday"][1].get("fallbackLab"))
        self.assertEqual("", schedules[(year, "A")]["Monday"][1].get("classroom"))
        self.assertEqual("C101", schedules[(year, "B")]["Monday"][1].get("classroom"))

    def test_single_section_uses_available_free_classrooms_period_by_period(self) -> None:
        year = "2nd Year"
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "SUB1",
            "subjectId": "SUB1",
            "facultyId": "F1",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "A")]["Monday"][2] = {
            "subject": "SUB2",
            "subjectId": "SUB2",
            "facultyId": "F2",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "A")]["Monday"][3] = {
            "subject": "SUB3",
            "subjectId": "SUB3",
            "facultyId": "F3",
            "isLab": False,
            "sharedSections": [],
        }
        session_log = [
            {"day": "Monday", "periods": [1], "sections": ["A"], "isLab": False},
            {"day": "Monday", "periods": [2], "sections": ["A"], "isLab": False},
            {"day": "Monday", "periods": [3], "sections": ["A"], "isLab": False},
        ]
        constraint_violations: list[dict] = []

        _allocate_classrooms_to_schedule(
            year,
            ["A"],
            schedules,
            session_log,
            ["C101", "C102"],
            constraint_violations,
            [1, 2, 3, 4, 5, 6, 7],
            [(1, 2, 3), (4, 5), (6, 7)],
            prior_room_busy={"C101": {("Monday", 2)}, "C102": {("Monday", 1)}},
            section_strength_map={"A": 30},
        )

        self.assertEqual([], constraint_violations)
        self.assertEqual("C101", schedules[(year, "A")]["Monday"][1].get("classroom"))
        self.assertEqual("C102", schedules[(year, "A")]["Monday"][2].get("classroom"))
        self.assertIn(schedules[(year, "A")]["Monday"][3].get("classroom"), {"C101", "C102"})

    def test_final_room_repair_assigns_free_room_when_session_log_misses_slot(self) -> None:
        year = "2nd Year"
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "SUB1",
            "subjectId": "SUB1",
            "facultyId": "F1",
            "isLab": False,
            "sharedSections": [],
        }
        constraint_violations: list[dict] = []

        _allocate_classrooms_to_schedule(
            year,
            ["A"],
            schedules,
            [],
            ["C101"],
            constraint_violations,
            [1, 2, 3, 4, 5, 6, 7],
            [(1, 2), (3, 4), (5, 6, 7)],
            section_strength_map={"A": 30},
        )

        self.assertEqual([], constraint_violations)
        self.assertEqual("C101", schedules[(year, "A")]["Monday"][1].get("classroom"))

    def test_classroom_template_2c1_section_maps_strength_from_same_file(self) -> None:
        store = MemoryStore()
        store.save_scoped_mapping(
            "classrooms",
            "global",
            {
                "rows": [
                    {"class_number": "1101", "room_type": "classroom", "capacity": 75, "section": "2C1", "strength": 70},
                    {"class_number": "1108", "room_type": "classroom", "capacity": 170, "section": "2G1", "strength": 70},
                    {"class_number": "2301", "room_type": "lab", "capacity": "", "section": "3C4", "strength": 70},
                ]
            },
            allow_overwrite=True,
        )
        request = GenerateTimetableRequest(
            year="2nd Year",
            section="C1",
            timetableMetadata=_metadata(),
        )

        room_inventory = _build_room_inventory(request, store)
        section_strength_map = _build_section_strength_map(request, store)

        self.assertEqual(70, section_strength_map.get("2nd Year|C1"))
        self.assertEqual("1101", room_inventory[0].get("name"))
        self.assertEqual(75, room_inventory[0].get("capacity"))

    def test_post_allocation_assigns_classroom_and_keeps_continuous_block_room(self) -> None:
        store = MemoryStore()
        store.save_scoped_mapping(
            "classrooms",
            "global",
            {"rows": [{"class_number": "C101"}, {"class_number": "C102"}]},
            allow_overwrite=True,
        )
        request = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="BLOCK",
                    facultyId="FBLOCK",
                    noOfHours=2,
                    continuousHours=2,
                    compulsoryContinuousHours=2,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="FILL",
                    facultyId="FFILL",
                    noOfHours=40,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])

        self.assertTrue(payload["hasValidTimetable"])
        self.assertEqual([], [v for v in payload["constraintViolations"] if v.get("constraint") == "classroom allocation constraint"])

        grid = payload["allGrids"]["A"]
        block_slots = _collect_subject_slots(grid, "BLOCK")
        self.assertEqual(2, len(block_slots))
        block_rooms = {
            grid[day][period - 1].get("classroom")
            for day, period in block_slots
            if grid[day][period - 1]
        }
        self.assertEqual({"C101"}, block_rooms)

        for day in DAYS:
            for cell in grid[day]:
                if not cell or cell.get("isLab"):
                    continue
                self.assertTrue(cell.get("classroom"))

    def test_post_allocation_reports_insufficient_classrooms(self) -> None:
        store = MemoryStore()
        store.save_scoped_mapping(
            "classrooms",
            "global",
            {"rows": [{"class_number": "C101"}]},
            allow_overwrite=True,
        )
        request = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="A_MAIN",
                    facultyId="FA",
                    noOfHours=42,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="B",
                    subjectId="B_MAIN",
                    facultyId="FB",
                    noOfHours=42,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])

        classroom_violations = [
            v for v in payload["constraintViolations"]
            if v.get("constraint") == "classroom allocation constraint"
        ]
        self.assertGreater(len(classroom_violations), 0)

    def test_post_allocation_reports_missing_classrooms_when_none_exist(self) -> None:
        year = "2nd Year"
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "SUB1",
            "subjectId": "SUB1",
            "facultyId": "F1",
            "isLab": False,
            "sharedSections": [],
        }
        schedules[(year, "A")]["Monday"][2] = {
            "subject": "SUB2",
            "subjectId": "SUB2",
            "facultyId": "F2",
            "isLab": False,
            "sharedSections": [],
        }
        constraint_violations: list[dict] = []

        _allocate_classrooms_to_schedule(
            year,
            ["A"],
            schedules,
            [{"day": "Monday", "periods": [1, 2], "sections": ["A"], "isLab": False}],
            [],
            constraint_violations,
            [1, 2, 3, 4, 5, 6, 7],
            [(1, 2), (3, 4), (5, 6, 7)],
            section_strength_map={"A": 30},
        )

        self.assertTrue(
            any(v.get("constraint") == "classroom allocation constraint" for v in constraint_violations),
            constraint_violations,
        )

    def test_compulsory_continuous_hours_enforced_as_contiguous_blocks(self) -> None:
        store = MemoryStore()
        store.save_scoped_mapping(
            "classrooms",
            "global",
            {"rows": [{"class_number": "C101"}, {"class_number": "C102"}]},
            allow_overwrite=True,
        )
        request = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="PAIR",
                    facultyId="FPAIR",
                    noOfHours=4,
                    continuousHours=2,
                    compulsoryContinuousHours=2,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="FILL",
                    facultyId="FFILL",
                    noOfHours=38,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])
        pair_slots = _collect_subject_slots(payload["allGrids"]["A"], "PAIR")

        self.assertEqual(4, len(pair_slots))
        grouped_by_day: dict[str, list[int]] = {}
        for day, period in pair_slots:
            grouped_by_day.setdefault(day, []).append(period)

        self.assertEqual(2, len(grouped_by_day), grouped_by_day)
        for periods in grouped_by_day.values():
            periods.sort()
            self.assertEqual(2, len(periods), periods)
            self.assertEqual(1, periods[1] - periods[0], periods)

    def test_compulsory_continuous_hours_allow_final_remainder_hour(self) -> None:
        store = MemoryStore()
        store.save_scoped_mapping(
            "classrooms",
            "global",
            {"rows": [{"class_number": "C101"}, {"class_number": "C102"}]},
            allow_overwrite=True,
        )
        request = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="FIVEBLOCK",
                    facultyId="FFIVE",
                    noOfHours=5,
                    continuousHours=2,
                    compulsoryContinuousHours=2,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="FILL",
                    facultyId="FFILL",
                    noOfHours=37,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])
        slots = _collect_subject_slots(payload["allGrids"]["A"], "FIVEBLOCK")

        self.assertEqual(5, len(slots))
        grouped_by_day: dict[str, list[int]] = {}
        for day, period in slots:
            grouped_by_day.setdefault(day, []).append(period)

        two_hour_blocks = 0
        single_hour_blocks = 0
        for periods in grouped_by_day.values():
            periods.sort()
            run_lengths: list[int] = []
            run = 1
            for idx in range(1, len(periods)):
                if periods[idx] == periods[idx - 1] + 1:
                    run += 1
                else:
                    run_lengths.append(run)
                    run = 1
            run_lengths.append(run)
            two_hour_blocks += sum(1 for length in run_lengths if length == 2)
            single_hour_blocks += sum(1 for length in run_lengths if length == 1)

        self.assertEqual(2, two_hour_blocks, grouped_by_day)
        self.assertEqual(1, single_hour_blocks, grouped_by_day)

    def test_subject_daily_preference_penalizes_a_third_period_when_another_day_is_available(self) -> None:
        year = "2nd Year"
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {"subject": "MATH", "subjectId": "MATH"}
        schedules[(year, "A")]["Monday"][2] = {"subject": "MATH", "subjectId": "MATH"}
        daily_subject_counts = _build_daily_subject_counts(schedules, year, ["A"], [1, 2, 3, 4, 5, 6, 7])
        requirement = Requirement(
            subject_id="MATH",
            faculty_id="F1",
            faculty_options=("F1",),
            faculty_team=(),
            sections=("A",),
            hours=1,
            min_consecutive_hours=1,
            max_consecutive_hours=1,
            shared=False,
        )

        monday_penalty = _subject_daily_limit_penalty(requirement, "Monday", 1, daily_subject_counts)
        tuesday_penalty = _subject_daily_limit_penalty(requirement, "Tuesday", 1, daily_subject_counts)
        candidates = _enumerate_slot_candidates(
            requirement=requirement,
            remaining_hours=1,
            schedules=schedules,
            daily_subject_counts=daily_subject_counts,
            faculty_busy={"F1": set()},
            faculty_availability={"F1": {day: {1, 2, 3, 4, 5, 6, 7} for day in DAYS}},
            faculty_section_slots={},
            year=year,
            days_order=["Monday", "Tuesday"],
            periods_order=[3],
            instructional_periods=[1, 2, 3, 4, 5, 6, 7],
            sessions=[(1, 2), (3, 4), (5, 6, 7)],
            session_adjacency=_build_session_adjacency([(1, 2), (3, 4), (5, 6, 7)]),
            candidate_limit=4,
            free_slots_tracker={"A": 40},
        )

        self.assertLess(monday_penalty, tuesday_penalty)
        self.assertGreater(len(candidates), 0)
        self.assertEqual("Tuesday", candidates[0].day)

    def test_subject_daily_hard_limit_blocks_fourth_period_for_non_continuous_subject(self) -> None:
        year = "2nd Year"
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        for period in (1, 2, 3):
            schedules[(year, "A")]["Monday"][period] = {"subject": "MATH", "subjectId": "MATH"}
        daily_subject_counts = _build_daily_subject_counts(schedules, year, ["A"], [1, 2, 3, 4, 5, 6, 7])
        requirement = Requirement(
            subject_id="MATH",
            faculty_id="F1",
            faculty_options=("F1",),
            faculty_team=(),
            sections=("A",),
            hours=1,
            min_consecutive_hours=1,
            max_consecutive_hours=1,
            shared=False,
        )

        allowed = _sections_are_free(
            sections=("A",),
            requirement=requirement,
            day="Monday",
            start_period=4,
            block_size=1,
            schedules=schedules,
            year=year,
            instructional_periods=[1, 2, 3, 4, 5, 6, 7],
            sessions=[(1, 2), (3, 4), (5, 6, 7)],
            daily_subject_counts=daily_subject_counts,
        )

        self.assertTrue(allowed)


    def test_subject_daily_hard_limit_exempts_explicit_continuous_blocks(self) -> None:
        year = "2nd Year"
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        for period in (3, 4):
            schedules[(year, "A")]["Monday"][period] = {"subject": "LAB", "subjectId": "LAB"}
        daily_subject_counts = _build_daily_subject_counts(schedules, year, ["A"], [1, 2, 3, 4, 5, 6, 7])
        requirement = Requirement(
            subject_id="LAB",
            faculty_id="F1",
            faculty_options=("F1",),
            faculty_team=(),
            sections=("A",),
            hours=4,
            min_consecutive_hours=2,
            max_consecutive_hours=2,
            shared=False,
            daily_limit_exempt=True,
        )

        allowed = _sections_are_free(
            sections=("A",),
            requirement=requirement,
            day="Monday",
            start_period=1,
            block_size=2,
            schedules=schedules,
            year=year,
            instructional_periods=[1, 2, 3, 4, 5, 6, 7],
            sessions=[(1, 2), (3, 4), (5, 6, 7)],
            daily_subject_counts=daily_subject_counts,
        )

        self.assertTrue(allowed)

    def test_solver_does_not_auto_group_matching_subjects_across_sections(self) -> None:
        store = MemoryStore()
        request = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="GLOBAL",
                    facultyId="VERBAL_POOL",
                    noOfHours=2,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="B",
                    subjectId="GLOBAL",
                    facultyId="VERBAL_POOL",
                    noOfHours=2,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="A_FILL",
                    facultyId="FA",
                    noOfHours=40,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="B",
                    subjectId="B_FILL",
                    facultyId="FB",
                    noOfHours=40,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
            facultyIdNameMapping=[
                {"facultyId": "FVA", "facultyName": "VERBAL_POOL Alpha"},
                {"facultyId": "FVB", "facultyName": "VERBAL_POOL Beta"},
            ],
            facultyAvailability=[
                FacultyWeeklyAvailabilityEntry(
                    facultyId="FVA",
                    availablePeriodsByDay={
                        "Monday": [1, 2],
                        "Tuesday": [],
                        "Wednesday": [],
                        "Thursday": [],
                        "Friday": [],
                        "Saturday": [],
                    },
                ),
                FacultyWeeklyAvailabilityEntry(
                    facultyId="FVB",
                    availablePeriodsByDay={
                        "Monday": [],
                        "Tuesday": [1, 2],
                        "Wednesday": [],
                        "Thursday": [],
                        "Friday": [],
                        "Saturday": [],
                    },
                )
            ],
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])

        self.assertTrue(payload["hasValidTimetable"])
        self.assertEqual([], payload["sharedClasses"])
        self.assertEqual(42, _count_filled_slots(payload["allGrids"]["A"]))
        self.assertEqual(42, _count_filled_slots(payload["allGrids"]["B"]))
        self.assertEqual(2, len(_collect_subject_slots(payload["allGrids"]["A"], "GLOBAL")))
        self.assertEqual(2, len(_collect_subject_slots(payload["allGrids"]["B"], "GLOBAL")))

    def test_solver_prioritizes_scarce_slots_over_high_hour_subject(self) -> None:
        store = MemoryStore()
        request = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="FLEX",
                    facultyId="F3",
                    noOfHours=40,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="LIMITED_MON",
                    facultyId="F1",
                    noOfHours=1,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="LIMITED_TUE",
                    facultyId="F2",
                    noOfHours=1,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
            facultyAvailability=[
                FacultyWeeklyAvailabilityEntry(
                    facultyId="F1",
                    availablePeriodsByDay={
                        "Monday": [1],
                        "Tuesday": [],
                        "Wednesday": [],
                        "Thursday": [],
                        "Friday": [],
                        "Saturday": [],
                    },
                ),
                FacultyWeeklyAvailabilityEntry(
                    facultyId="F2",
                    availablePeriodsByDay={
                        "Monday": [],
                        "Tuesday": [1],
                        "Wednesday": [],
                        "Thursday": [],
                        "Friday": [],
                        "Saturday": [],
                    },
                ),
            ],
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])

        self.assertTrue(payload["hasValidTimetable"])
        self.assertEqual([], payload["constraintViolations"])
        self.assertEqual([], payload["unscheduledSubjects"])
        self.assertEqual(30, payload["generationMeta"]["timeoutSeconds"])  # §18: small config ≤5 sections = 30s
        self.assertEqual(2, payload["generationMeta"]["retryStrategies"])
        self.assertIn("shared-first", payload["generationMeta"]["attemptStrategies"])
        self.assertEqual(42, _count_filled_slots(payload["allGrids"]["A"]))
        self.assertEqual("LIMITED_MON", payload["allGrids"]["A"]["Monday"][0]["subject"])
        self.assertEqual("LIMITED_TUE", payload["allGrids"]["A"]["Tuesday"][0]["subject"])

    def test_solver_generates_full_timetable_with_locked_and_shared_sessions(self) -> None:
        store = MemoryStore()
        request = GenerateTimetableRequest(
            year="3rd Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="3rd Year",
                    section="A",
                    subjectId="LABSH",
                    facultyId="FLAB",
                    noOfHours=3,
                    continuousHours=3,
                    compulsoryContinuousHours=3,
                ),
                ManualEntryMode(
                    year="3rd Year",
                    section="B",
                    subjectId="LABSH",
                    facultyId="FLAB",
                    noOfHours=3,
                    continuousHours=3,
                    compulsoryContinuousHours=3,
                ),
                ManualEntryMode(
                    year="3rd Year",
                    section="A",
                    subjectId="SHARED",
                    facultyId="FSH",
                    noOfHours=3,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="3rd Year",
                    section="B",
                    subjectId="SHARED",
                    facultyId="FSH",
                    noOfHours=3,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="3rd Year",
                    section="A",
                    subjectId="A_ONLY",
                    facultyId="FA",
                    noOfHours=36,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="3rd Year",
                    section="B",
                    subjectId="B_ONLY",
                    facultyId="FB",
                    noOfHours=36,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
            manualLabEntries=[
                ManualLabEntry(year="3rd Year", section="A", subjectId="LABSH", day=1, hours=[1, 2, 3], venue="Lab 1"),
                ManualLabEntry(year="3rd Year", section="B", subjectId="LABSH", day=1, hours=[1, 2, 3], venue="Lab 1"),
            ],
            sharedClasses=[SharedClassEntry(year="3rd Year", sections=["A", "B"], subject="SHARED")],
            facultyAvailability=[
                FacultyWeeklyAvailabilityEntry(
                    facultyId="FLAB",
                    availablePeriodsByDay={
                        "Monday": [1, 2, 3],
                        "Tuesday": [],
                        "Wednesday": [],
                        "Thursday": [],
                        "Friday": [],
                        "Saturday": [],
                    },
                ),
                FacultyWeeklyAvailabilityEntry(
                    facultyId="FSH",
                    availablePeriodsByDay={
                        "Monday": [],
                        "Tuesday": [1, 2, 3],
                        "Wednesday": [],
                        "Thursday": [],
                        "Friday": [],
                        "Saturday": [],
                    },
                ),
            ],
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])

        self.assertTrue(payload["hasValidTimetable"])
        self.assertEqual(42, _count_filled_slots(payload["allGrids"]["A"]))
        self.assertEqual(42, _count_filled_slots(payload["allGrids"]["B"]))
        self.assertEqual("LABSH", payload["allGrids"]["A"]["Monday"][0]["subject"])
        self.assertEqual("LABSH", payload["allGrids"]["B"]["Monday"][0]["subject"])
        shared_subjects = {item["subject_id"] for item in payload["sharedClasses"]}
        self.assertIn("LABSH", shared_subjects)
        self.assertIn("SHARED", shared_subjects)
        self.assertEqual([], payload["constraintViolations"])
        self.assertIn("sectionTimetables", payload["generatedFiles"])
        self.assertIn("facultyWorkload", payload["generatedFiles"])
        self.assertIn("sharedClassesReport", payload["generatedFiles"])

    def test_solver_resets_state_between_year_runs(self) -> None:
        store = MemoryStore()
        year_one = GenerateTimetableRequest(
            year="1st Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="1st Year",
                    section="A",
                    subjectId="Y1_MAIN",
                    facultyId="F_SHARED",
                    noOfHours=42,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                )
            ],
        )
        year_two = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="Y2_MAIN",
                    facultyId="F_SHARED",
                    noOfHours=42,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                )
            ],
        )

        first_result = generate_timetable(year_one, store)
        second_result = generate_timetable(year_two, store)
        first_payload = store.get_timetable(first_result["timetableId"])
        second_payload = store.get_timetable(second_result["timetableId"])

        self.assertTrue(first_payload["hasValidTimetable"])
        self.assertTrue(second_payload["hasValidTimetable"])
        self.assertEqual(42, _count_filled_slots(second_payload["allGrids"]["A"]))


    def test_lab_shared_session_no_false_faculty_conflict(self) -> None:
        """
        Requirement §10: Faculty teaching multiple sections the SAME lab at the
        same time is VALID. This must not produce a faculty conflict violation.
        C4 and C5 both have Lab FL1 with faculty FLAB, Monday P3+P4.
        Expected: no constraint violations, one shared lab entry in sharedClasses.
        """
        store = MemoryStore()
        request = GenerateTimetableRequest(
            year="2nd Year",
            section="C4",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="C4",
                    subjectId="FL1",
                    facultyId="FLAB",
                    noOfHours=2,
                    continuousHours=2,
                    compulsoryContinuousHours=2,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="C4",
                    subjectId="FILL_C4",
                    facultyId="FA",
                    noOfHours=40,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="C5",
                    subjectId="FL1",
                    facultyId="FLAB",
                    noOfHours=2,
                    continuousHours=2,
                    compulsoryContinuousHours=2,
                ),
                ManualEntryMode(
                    year="2nd Year",
                    section="C5",
                    subjectId="FILL_C5",
                    facultyId="FB",
                    noOfHours=40,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
            manualLabEntries=[
                ManualLabEntry(year="2nd Year", section="C4", subjectId="FL1", day=1, hours=[3, 4]),
                ManualLabEntry(year="2nd Year", section="C5", subjectId="FL1", day=1, hours=[3, 4]),
            ],
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])

        # No faculty conflicts — same faculty same session across two sections is valid
        faculty_conflict_violations = [
            v for v in payload["constraintViolations"]
            if "faculty" in v.get("constraint", "").lower()
        ]
        self.assertEqual([], faculty_conflict_violations, f"Unexpected faculty conflicts: {faculty_conflict_violations}")

        # The lab session should appear in sharedClasses (source = "lab")
        lab_shared = [s for s in payload["sharedClasses"] if s.get("subject_id") == "FL1"]
        self.assertGreater(len(lab_shared), 0, "FL1 lab should appear in shared class report")

        # Both sections should be in the shared entry
        for entry in lab_shared:
            self.assertIn("C4", entry.get("sections", []))
            self.assertIn("C5", entry.get("sections", []))

        # Both sections fully scheduled
        self.assertTrue(payload["hasValidTimetable"])

    def test_shared_report_only_contains_file_driven_sessions(self) -> None:
        """
        Requirement §21: The shared class report must NEVER include auto-detected
        sessions. Two sections with the same subject + same faculty should NOT
        appear in sharedClasses unless explicitly declared in the shared class file.
        """
        store = MemoryStore()
        # Two sections X and Y both have COMMON taught by FCOMMON — but NO shared class declaration
        request = GenerateTimetableRequest(
            year="3rd Year",
            section="X",
            timetableMetadata=_metadata(),
            manualEntries=[
                ManualEntryMode(
                    year="3rd Year",
                    section="X",
                    subjectId="COMMON",
                    facultyId="FCOMMON",
                    noOfHours=2,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="3rd Year",
                    section="X",
                    subjectId="FILL_X",
                    facultyId="FX",
                    noOfHours=40,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="3rd Year",
                    section="Y",
                    subjectId="COMMON",
                    facultyId="FCOMMON",
                    noOfHours=2,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
                ManualEntryMode(
                    year="3rd Year",
                    section="Y",
                    subjectId="FILL_Y",
                    facultyId="FY",
                    noOfHours=40,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                ),
            ],
            # NO sharedClasses entry for COMMON — it must NOT appear in shared report
        )

        result = generate_timetable(request, store)
        payload = store.get_timetable(result["timetableId"])

        # COMMON must NOT appear in the shared class report (it's solver-placed, not file-driven)
        common_in_shared = [s for s in payload["sharedClasses"] if s.get("subject_id") == "COMMON"]
        self.assertEqual(
            [],
            common_in_shared,
            f"COMMON was auto-detected as shared — this is wrong! Entries: {common_in_shared}",
        )
        self.assertTrue(payload["hasValidTimetable"])

    def test_team_teaching_requires_simultaneous_availability(self) -> None:
        # F1 is busy on Mon P1, F2 is free. Team [F1, F2] cannot schedule Mon P1.
        # F1 and F2 are free on Mon P2. Team [F1, F2] can schedule Mon P2.
        requirement = Requirement(
            subject_id="TEAM_SUBJ",
            faculty_id="F1, F2",
            faculty_options=(),
            faculty_team=("F1", "F2"),
            sections=("A",),
            hours=1,
            min_consecutive_hours=1,
            max_consecutive_hours=1,
            shared=False,
        )
        schedules = {
            ("2nd Year", "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        faculty_availability = {
            "F1": {day: {1, 2, 3, 4, 5, 6, 7} for day in DAYS},
            "F2": {day: {1, 2, 3, 4, 5, 6, 7} for day in DAYS},
        }

        # Case 1: F1 is busy on Monday period 1
        faculty_busy_1 = {"F1": {("Monday", 1)}, "F2": set()}
        cand_busy_1 = _choose_faculty_for_slot(
            requirement, "Monday", 1, 1,
            faculty_busy_1, faculty_availability, [1, 2, 3, 4, 5, 6, 7]
        )
        self.assertIsNone(cand_busy_1)

        # Case 2: F2 is busy on Monday period 1
        faculty_busy_2 = {"F1": set(), "F2": {("Monday", 1)}}
        cand_busy_2 = _choose_faculty_for_slot(
            requirement, "Monday", 1, 1,
            faculty_busy_2, faculty_availability, [1, 2, 3, 4, 5, 6, 7]
        )
        self.assertIsNone(cand_busy_2)

        # Case 3: Both are free on Monday period 2
        faculty_busy_3 = {"F1": set(), "F2": set()}
        cand_free = _choose_faculty_for_slot(
            requirement, "Monday", 2, 1,
            faculty_busy_3, faculty_availability, [1, 2, 3, 4, 5, 6, 7]
        )
        self.assertEqual("F1,F2", cand_free)

    def test_shared_classroom_allocation_dynamic_preference(self) -> None:
        year = "2nd Year"
        sections = ["C4", "C5"]
        schedules = {
            (year, "C4"): {day: {period: None for period in range(1, 8)} for day in DAYS},
            (year, "C5"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        # Mon P1: Independent class for C4 & C5
        schedules[(year, "C4")]["Monday"][1] = {"subject": "MATH", "subjectId": "MATH", "isLab": False}
        schedules[(year, "C5")]["Monday"][1] = {"subject": "PHY", "subjectId": "PHY", "isLab": False}

        # Mon P2: Shared class for C4 & C5
        schedules[(year, "C4")]["Monday"][2] = {"subject": "CHEM", "subjectId": "CHEM", "isLab": False, "sharedSections": ["C4", "C5"]}
        schedules[(year, "C5")]["Monday"][2] = {"subject": "CHEM", "subjectId": "CHEM", "isLab": False, "sharedSections": ["C4", "C5"]}

        # Home classrooms and capacities:
        # C4 home: Room 2304 (capacity 75, strength 35)
        # C5 home: Room 2305 (capacity 60, strength 30)
        classrooms = ["Room 2304", "Room 2305", "Auditorium"]
        room_capacity_map = {"Room 2304": 75, "Room 2305": 60, "Auditorium": 100}
        section_strength_map = {"2nd Year|C4": 35, "2nd Year|C5": 30}
        section_home_room_map = {"2nd Year|C4": "Room 2304", "2nd Year|C5": "Room 2305"}

        session_log = [
            {"day": "Monday", "periods": [1], "sections": ["C4"], "subject_id": "MATH", "isLab": False},
            {"day": "Monday", "periods": [1], "sections": ["C5"], "subject_id": "PHY", "isLab": False},
            {"day": "Monday", "periods": [2], "sections": ["C4", "C5"], "subject_id": "CHEM", "isLab": False},
        ]

        constraint_violations = []
        _allocate_classrooms_to_schedule(
            year=year,
            sections=sections,
            schedules=schedules,
            session_log=session_log,
            classrooms=classrooms,
            constraint_violations=constraint_violations,
            instructional_periods=[1, 2, 3, 4, 5, 6, 7],
            academic_sessions=[(1, 2), (3, 4), (5, 6, 7)],
            prior_room_busy=None,
            lab_room_names=None,
            room_capacity_map=room_capacity_map,
            section_strength_map=section_strength_map,
            fixed_classroom_blocks=None,
            section_home_room_map=section_home_room_map,
        )

        self.assertEqual([], constraint_violations)
        # Non-shared period uses home room:
        self.assertEqual("Room 2304", schedules[(year, "C4")]["Monday"][1].get("classroom"))
        self.assertEqual("Room 2305", schedules[(year, "C5")]["Monday"][1].get("classroom"))

        # Shared period uses the home room with larger capacity (Room 2304 has capacity 75 > Room 2305 capacity 60)
        # Combined strength = 35 + 30 = 65, Room 2304 fits (75 >= 65)
        self.assertEqual("Room 2304", schedules[(year, "C4")]["Monday"][2].get("classroom"))
        self.assertEqual("Room 2304", schedules[(year, "C5")]["Monday"][2].get("classroom"))

    def test_existing_faculty_workload_blocks_allocation(self) -> None:
        from storage.memory_store import store
        store.save_scoped_mapping(
            "existing_faculty_workloads",
            "global",
            {
                "rows": [
                    {
                        "faculty_id": "F101",
                        "faculty_name": "Dr. Rao",
                        "day": "Monday",
                        "period": 3,
                        "cell_value": "MATH",
                    }
                ]
            },
            allow_overwrite=True
        )

        request_data = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata={
                "academicYear": "2025-2026",
                "semester": 1,
                "withEffectFrom": "2025-06-16"
            },
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="MATH",
                    facultyId="F101",
                    noOfHours=1,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                )
            ]
        )

        from services.timetable_generator import generate_timetable
        store.save_scoped_mapping("faculty_id_map", "global", {"rows": [{"faculty_id": "F101", "faculty_name": "Dr. Rao"}]}, allow_overwrite=True)
        store.save_scoped_mapping("main_timetable_config", "global", {"rows": []}, allow_overwrite=True)
        store.save_scoped_mapping("lab_timetable_config", "global", {"rows": []}, allow_overwrite=True)

        res = generate_timetable(request_data, store)
        timetable_id = res["timetableId"]
        timetable = store.get_timetable(timetable_id)

        monday_grid = timetable["allGrids"]["A"]["Monday"]
        self.assertIsNone(monday_grid[2])

    def test_existing_classroom_timetable_blocks_room_allocation(self) -> None:
        from storage.memory_store import store
        store.save_scoped_mapping(
            "existing_classroom_timetables",
            "global",
            {
                "rows": [
                    {
                        "classroom": "C101",
                        "day": "Monday",
                        "period": 1,
                        "cell_value": "BUSY_CLASS",
                    }
                ]
            },
            allow_overwrite=True
        )

        year = "2nd Year"
        sections = ["A"]
        schedules = {
            (year, "A"): {day: {period: None for period in range(1, 8)} for day in DAYS},
        }
        schedules[(year, "A")]["Monday"][1] = {
            "subject": "MATH",
            "subjectId": "MATH",
            "facultyId": "F1",
            "isLab": False,
            "sharedSections": [],
        }
        session_log = [
            {"day": "Monday", "periods": [1], "sections": ["A"], "isLab": False},
        ]
        constraint_violations = []

        _allocate_classrooms_to_schedule(
            year=year,
            sections=sections,
            schedules=schedules,
            session_log=session_log,
            classrooms=["C101", "C102"],
            constraint_violations=constraint_violations,
            instructional_periods=[1, 2, 3, 4, 5, 6, 7],
            academic_sessions=[(1, 2), (3, 4), (5, 6, 7)],
            prior_room_busy={"C101": {("Monday", 1)}},
            lab_room_names=None,
            room_capacity_map={"C101": 50, "C102": 50},
            section_strength_map={"A": 30},
        )

        self.assertEqual([], constraint_violations)
        self.assertEqual("C102", schedules[(year, "A")]["Monday"][1].get("classroom"))

    def test_master_workbook_template_endpoint(self) -> None:
        from fastapi.testclient import TestClient
        from main import app
        import openpyxl
        from io import BytesIO

        client = TestClient(app)
        
        # Test example workbook
        response = client.get("/api/templates/master-workbook?type=example")
        self.assertEqual(200, response.status_code)
        self.assertEqual("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", response.headers["content-type"])
        self.assertIn("attachment; filename=\"master-workbook-template.xlsx\"", response.headers["content-disposition"])
        
        # Parse the output workbook to verify sheets and content
        wb = openpyxl.load_workbook(BytesIO(response.content), data_only=True)
        expected_sheets = [
            "Subjects",
            "Faculty Mapping",
            "Constraints",
            "Labs",
            "Shared Classes",
            "Sessions",
            "Classrooms",
            "Faculty Availability",
            "Continuous Rules",
            "Fixed Classroom Blocks",
        ]
        self.assertEqual(expected_sheets, wb.sheetnames)
        
        # Verify Subjects sheet data exists
        ws_subjects = wb["Subjects"]
        self.assertEqual("SUBJECT_ID", ws_subjects.cell(row=1, column=1).value)
        self.assertEqual("SUBJECT_NAME", ws_subjects.cell(row=1, column=2).value)
        self.assertEqual("7", str(ws_subjects.cell(row=2, column=1).value))
        self.assertEqual("Data Structures", ws_subjects.cell(row=2, column=2).value)

        # Test empty workbook
        response_empty = client.get("/api/templates/master-workbook?type=empty")
        self.assertEqual(200, response_empty.status_code)
        wb_empty = openpyxl.load_workbook(BytesIO(response_empty.content), data_only=True)
        self.assertEqual(expected_sheets, wb_empty.sheetnames)
        
        # Verify empty sheet contains only header and no example rows
        ws_subjects_empty = wb_empty["Subjects"]
        self.assertEqual("SUBJECT_ID", ws_subjects_empty.cell(row=1, column=1).value)
        self.assertIsNone(ws_subjects_empty.cell(row=2, column=1).value)

    def test_existing_timetables_merged_into_generated_outputs(self) -> None:
        from storage.memory_store import store
        from services.timetable_generator import generate_timetable
        import base64
        import openpyxl
        from io import BytesIO

        store.save_scoped_mapping(
            "existing_faculty_workloads",
            "global",
            {
                "rows": [
                    {
                        "faculty_id": "F101",
                        "faculty_name": "Dr. Rao",
                        "day": "Monday",
                        "period": 3,
                        "cell_value": "EXISTING_FAC_WORKLOAD",
                    }
                ]
            },
            allow_overwrite=True
        )
        store.save_scoped_mapping(
            "existing_classroom_timetables",
            "global",
            {
                "rows": [
                    {
                        "classroom": "C101",
                        "day": "Monday",
                        "period": 1,
                        "cell_value": "EXISTING_ROOM_TIMETABLE",
                    }
                ]
            },
            allow_overwrite=True
        )

        request_data = GenerateTimetableRequest(
            year="2nd Year",
            section="A",
            timetableMetadata={
                "academicYear": "2025-2026",
                "semester": 1,
                "withEffectFrom": "2025-06-16"
            },
            manualEntries=[
                ManualEntryMode(
                    year="2nd Year",
                    section="A",
                    subjectId="MATH",
                    facultyId="F101",
                    noOfHours=1,
                    continuousHours=1,
                    compulsoryContinuousHours=1,
                )
            ]
        )
        store.save_scoped_mapping("faculty_id_map", "global", {"rows": [{"faculty_id": "F101", "faculty_name": "Dr. Rao"}]}, allow_overwrite=True)
        store.save_scoped_mapping("main_timetable_config", "global", {"rows": []}, allow_overwrite=True)
        store.save_scoped_mapping("lab_timetable_config", "global", {"rows": []}, allow_overwrite=True)
        store.save_scoped_mapping("classroom_inventory", "global", {"rows": [{"classroom": "C101", "capacity": 60}]}, allow_overwrite=True)

        res = generate_timetable(request_data, store)
        timetable_id = res["timetableId"]
        timetable = store.get_timetable(timetable_id)

        # 1. Verify JSON output
        # Verify faculty workloads
        fac_workload = timetable["facultyWorkloads"].get("Dr. Rao") or timetable["facultyWorkloads"].get("F101")
        self.assertIsNotNone(fac_workload)
        # Period 3 is index 2
        self.assertEqual("EXISTING_FAC_WORKLOAD", fac_workload["Monday"][2])

        # Verify room grids
        room_grid = timetable["roomGrids"].get("C101")
        self.assertIsNotNone(room_grid)
        # Period 1 is index 0
        self.assertEqual("EXISTING_ROOM_TIMETABLE", room_grid["Monday"][0]["subject"])

        # 2. Verify downloadables
        # Verify roomTimetables Excel
        room_tt_b64 = timetable["generatedFiles"]["roomTimetables"]["contentBase64"]
        room_wb = openpyxl.load_workbook(BytesIO(base64.b64decode(room_tt_b64)), data_only=True)
        self.assertIn("C101", room_wb.sheetnames)
        # In room sheet, row 9 is Monday. Period 1 is column 2
        self.assertEqual("EXISTING_ROOM_TIMETABLE", room_wb["C101"].cell(row=9, column=2).value)

        # Verify facultyWorkload Excel
        fac_wl_b64 = timetable["generatedFiles"]["facultyWorkload"]["contentBase64"]
        fac_wb = openpyxl.load_workbook(BytesIO(base64.b64decode(fac_wl_b64)), data_only=True)
        sheet_name = [name for name in fac_wb.sheetnames if "Rao" in name or "F101" in name][0]
        # In faculty sheet, row 9 is Monday. Period 3 is column 5 (since column 2 is period 1, 3 is period 2, 4 is break, 5 is period 3)
        self.assertEqual("EXISTING_FAC_WORKLOAD", fac_wb[sheet_name].cell(row=9, column=5).value)


if __name__ == "__main__":
    unittest.main()

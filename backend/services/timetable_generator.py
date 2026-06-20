from __future__ import annotations

import base64
import time
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, Side

from models.schemas import GenerateTimetableRequest
from services.utils import normalize_year
from storage.memory_store import MemoryStore

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
PERIODS = [1, 2, 3, 4, 5, 6, 7]
DAY_INDEX = {index + 1: day for index, day in enumerate(DAYS)}
DAY_SHORT_LABELS = {
    "Monday": "M\nO\nN",
    "Tuesday": "T\nU\nE",
    "Wednesday": "W\nE\nD",
    "Thursday": "T\nH\nU",
    "Friday": "F\nR\nI",
    "Saturday": "S\nA\nT",
}
ACADEMIC_METADATA = {
    "college": "NARASARAOPETA ENGINEERING COLLEGE :: NARASARAOPET",
    "department": "Department of Computer Science & Engineering",
    "semester": "II SEMESTER",
}
THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
MEDIUM_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGNMENT = Alignment(horizontal="left", vertical="center", wrap_text=True)
BOLD_FONT = Font(bold=True)
PREFERRED_SUBJECT_PERIODS_PER_DAY = 2
HARD_SUBJECT_PERIODS_PER_DAY = 3


def _academic_year_label() -> str:
    now = datetime.now()
    start_year = now.year if now.month >= 6 else now.year - 1
    return f"{start_year}-{start_year + 1}"


def _semester_label(value: int | str | None) -> str:
    semester = str(value or "").strip()
    if semester == "1":
        return "I SEMESTER"
    if semester == "2":
        return "II SEMESTER"
    return ACADEMIC_METADATA["semester"]


def _display_effective_date(value: str | None) -> str:
    raw = str(value or "").strip()
    if not raw:
        return ""
    try:
        return datetime.strptime(raw, "%Y-%m-%d").strftime("%d-%m-%Y")
    except ValueError:
        return raw


def _resolve_timetable_metadata(raw_meta: dict[str, Any] | None) -> dict[str, str]:
    meta = raw_meta or {}
    academic_year = str(meta.get("academicYear") or "").strip() or _academic_year_label()
    semester = _semester_label(meta.get("semester"))
    with_effect_from = str(meta.get("withEffectFrom") or "").strip()
    return {
        "academicYear": academic_year,
        "semester": semester,
        "withEffectFrom": with_effect_from,
        "withEffectFromDisplay": _display_effective_date(with_effect_from),
    }


def _compact_academic_year_label(value: str) -> str:
    academic_year = str(value or "").strip()
    parts = academic_year.split("-")
    if len(parts) == 2 and len(parts[0]) == 4 and len(parts[1]) == 4:
        return f"{parts[0]} - {parts[1][-2:]}"
    return academic_year


def _faculty_workload_academic_line(metadata: dict[str, str]) -> str:
    semester = str(metadata.get("semester", "")).strip().upper()
    return (
        f"ACADEMIC YEAR : {_compact_academic_year_label(metadata.get('academicYear', ''))} "
        f"({semester})  Time Table"
    )


def _class_title(year: str, section: str, semester_label: str | None = None) -> str:
    year_map = {
        "2nd Year": "II B.Tech",
        "3rd Year": "III B.Tech",
        "4th Year": "IV B.Tech",
    }
    normalized = year_map.get(year, year)
    return f"{normalized} [CSE - {section}] {semester_label or ACADEMIC_METADATA['semester']} TIME TABLE"


@dataclass
class Requirement:
    subject_id: str
    faculty_id: str
    faculty_options: tuple[str, ...]
    faculty_team: tuple[str, ...]
    sections: tuple[str, ...]
    hours: int
    min_consecutive_hours: int
    max_consecutive_hours: int
    shared: bool
    preferred_daily_limit: int = PREFERRED_SUBJECT_PERIODS_PER_DAY
    hard_daily_limit: int = HARD_SUBJECT_PERIODS_PER_DAY
    daily_limit_exempt: bool = False
    common_faculty: bool = False
    phase: int = 4


@dataclass(frozen=True)
class SlotCandidate:
    day: str
    start_period: int
    block_size: int
    faculty_id: str
    faculty_ids: tuple[str, ...]
    score: int


def _validation_error(message: str, details: list | None = None) -> HTTPException:
    return HTTPException(
        status_code=400,
        detail={"error": "ValidationError", "message": message, "details": details or []},
    )


def _normalize_day(value: str | int) -> str | None:
    if isinstance(value, int):
        return DAY_INDEX.get(value)
    text = str(value).strip()
    if not text:
        return None
    if text.isdigit():
        return DAY_INDEX.get(int(text))
    lowered = text.lower()
    for day in DAYS:
        if day.lower().startswith(lowered[:3]):
            return day
    return None


def _resolve_faculty_output(faculty_id: str, faculty_id_to_name: dict[str, str]) -> tuple[str, str]:
    token = str(faculty_id).strip()
    faculty_name = faculty_id_to_name.get(token, token)
    return token, faculty_name


def _resolve_subject_output(subject_id: str, subject_id_to_name: dict[str, str]) -> tuple[str, str]:
    token = str(subject_id).strip()
    if token.endswith(".0"):
        token = token[:-2]
    subject_name = subject_id_to_name.get(token, token)
    return token, subject_name


def _normalize_id_token(value: str | int | float | None) -> str:
    token = str(value or "").strip()
    if token.lower() == "nan":
        return ""
    if token.endswith(".0"):
        token = token[:-2]
    return token


def _parse_year_and_section_token(value: str | int | float | None) -> tuple[str, str]:
    import re

    text = _normalize_id_token(value)
    if not text:
        return "", ""

    normalized_text = re.sub(r"\s+", " ", text).strip()
    match = re.search(
        r"(?P<year>(?:\d+(?:ST|ND|RD|TH)?|I|II|III|IV))\s*(?:YEAR)?\s*[-/]?\s*(?P<section>[A-Z]\d+)",
        normalized_text,
        re.IGNORECASE,
    )
    if not match:
        compact = re.sub(r"[^A-Z0-9]", "", normalized_text.upper())
        match = re.search(r"(?P<year>(?:1|2|3|4|I|II|III|IV))(?P<section>[A-Z]\d+)", compact, re.IGNORECASE)
    if not match:
        return "", ""
    return normalize_year(match.group("year") or ""), _normalize_id_token(match.group("section") or "").upper()


def _split_faculty_tokens(raw_faculty_id: str) -> tuple[str, ...]:
    tokens = [str(raw_faculty_id or "").strip()]
    for delimiter in [",", "/", "|", "+", "&"]:
        next_tokens: list[str] = []
        for token in tokens:
            next_tokens.extend(part.strip() for part in token.split(delimiter))
        tokens = next_tokens
    cleaned: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        normalized = _normalize_id_token(token)
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        cleaned.append(normalized)
    return tuple(cleaned)


def _normalize_faculty_field(raw_faculty_id: str | int | float | None) -> str:
    return ",".join(_split_faculty_tokens(str(raw_faculty_id or "")))


def _faculty_alias_to_id_map(faculty_id_to_name: dict[str, str]) -> dict[str, str]:
    alias_map: dict[str, str] = {}
    for faculty_id, faculty_name in faculty_id_to_name.items():
        normalized_id = _normalize_id_token(faculty_id)
        if normalized_id:
            alias_map[normalized_id] = normalized_id
            alias_map[normalized_id.lower()] = normalized_id
        normalized_name = str(faculty_name).strip()
        if normalized_name:
            alias_map[normalized_name] = normalized_id or normalized_name
            alias_map[normalized_name.lower()] = normalized_id or normalized_name
    return alias_map


def _canonicalize_faculty_token(
    value: str | int | float | None,
    faculty_id_to_name: dict[str, str],
) -> str:
    token = _normalize_id_token(value)
    if not token:
        return ""
    alias_map = _faculty_alias_to_id_map(faculty_id_to_name)
    return alias_map.get(token, alias_map.get(token.lower(), token))


def _resolve_faculty_display(
    faculty_ids: tuple[str, ...],
    faculty_id_to_name: dict[str, str],
) -> tuple[str, str]:
    ids = tuple(
        _canonicalize_faculty_token(fid, faculty_id_to_name)
        for fid in faculty_ids
        if _canonicalize_faculty_token(fid, faculty_id_to_name)
    )
    names = tuple(faculty_id_to_name.get(fid, fid) for fid in ids)
    return ",".join(ids), ", ".join(names)


def _build_faculty_maps(request_data: GenerateTimetableRequest, store: MemoryStore) -> dict[str, str]:
    faculty_id_to_name: dict[str, str] = {}
    fac_map_payload = store.get_scoped_mapping("faculty_id_map", "global")
    # Fallback: load from file-map if caller supplied mappingFileIds but
    # scoped mappings are not present (or empty) (common when backend restarts).
    if (not fac_map_payload or not fac_map_payload.get("rows")) and request_data.mappingFileIds and request_data.mappingFileIds.facultyIdMap:
        fac_map_payload = store.get_file_map(request_data.mappingFileIds.facultyIdMap)
    if fac_map_payload:
        for row in fac_map_payload.get("rows", []):
            faculty_id = _normalize_faculty_field(row.get("faculty_id", ""))
            faculty_name = str(row.get("faculty_name", "")).strip()
            if faculty_id:
                faculty_id_to_name[faculty_id] = faculty_name or faculty_id
    for row in request_data.facultyIdNameMapping:
        faculty_id = _normalize_faculty_field(row.facultyId)
        faculty_name = str(row.facultyName).strip()
        if faculty_id:
            faculty_id_to_name[faculty_id] = faculty_name or faculty_id
    return faculty_id_to_name


def _build_subject_maps(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
) -> tuple[dict[str, str], dict[str, int]]:
    subject_id_to_name: dict[str, str] = {}
    compulsory_continuous: dict[str, int] = {}

    subject_map_payload = store.get_scoped_mapping("subject_id_mapping", "global")
    if (not subject_map_payload or not subject_map_payload.get("rows")) and request_data.mappingFileIds and request_data.mappingFileIds.subjectIdMapping:
        subject_map_payload = store.get_file_map(request_data.mappingFileIds.subjectIdMapping)
    if subject_map_payload:
        for row in subject_map_payload.get("rows", []):
            subject_id = _normalize_id_token(row.get("subject_id", ""))
            subject_name = str(row.get("subject_name", "")).strip()
            if subject_id:
                subject_id_to_name[subject_id] = subject_name or subject_id

    rule_payload = store.get_scoped_mapping("subject_continuous_rules", "global")
    if (not rule_payload or not rule_payload.get("rows")) and request_data.mappingFileIds and request_data.mappingFileIds.subjectContinuousRules:
        rule_payload = store.get_file_map(request_data.mappingFileIds.subjectContinuousRules)
    if rule_payload:
        for row in rule_payload.get("rows", []):
            subject_id = _normalize_id_token(row.get("subject_id", ""))
            if subject_id:
                compulsory_continuous[subject_id] = max(
                    1, int(row.get("compulsory_continuous_hours", 1) or 1)
                )

    for row in request_data.subjectIdNameMapping:
        subject_id = _normalize_id_token(row.subjectId)
        subject_name = str(row.subjectName).strip()
        if subject_id:
            subject_id_to_name[subject_id] = subject_name or subject_id

    for row in request_data.subjectContinuousRules:
        subject_id = _normalize_id_token(row.subjectId)
        if subject_id:
            compulsory_continuous[subject_id] = max(1, int(row.compulsoryContinuousHours or 1))

    return subject_id_to_name, compulsory_continuous


DEFAULT_PERIOD_CONFIG = [
    {"period": "1", "time": "9.10-10.00"},
    {"period": "2", "time": "10.00-10.50"},
    {"period": "Break", "time": "10.50-11.00"},
    {"period": "3", "time": "11.00-11.50"},
    {"period": "4", "time": "11.50-12.40"},
    {"period": "Lunch", "time": "12.40-1.30"},
    {"period": "5", "time": "1.30-2.20"},
    {"period": "6", "time": "2.20-3.10"},
    {"period": "7", "time": "3.10-4.00"},
]


def _build_period_config(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
) -> list[dict[str, str]]:
    payload = store.get_scoped_mapping("period_configuration", "global")
    if (not payload or not payload.get("rows")) and request_data.mappingFileIds and request_data.mappingFileIds.periodConfiguration:
        payload = store.get_file_map(request_data.mappingFileIds.periodConfiguration)
    
    rows = []
    if payload:
        rows = payload.get("rows", [])
    
    # Merge with direct entries
    for entry in request_data.periodConfiguration:
        rows.append({"period": entry.period, "time": entry.time})
    
    if not rows:
        return DEFAULT_PERIOD_CONFIG
    return rows



def _derive_sessions(period_config: list[dict[str, str]]) -> tuple[list[int], list[tuple[int, ...]]]:
    instructional_periods = []
    sessions = []
    current_session = []
    
    for row in period_config:
        p = str(row.get("period", "")).strip()
        if p.isdigit():
            val = int(p)
            instructional_periods.append(val)
            current_session.append(val)
        else:
            if current_session:
                sessions.append(tuple(current_session))
                current_session = []
    if current_session:
        sessions.append(tuple(current_session))
    
    return sorted(list(set(instructional_periods))), sessions


def _build_classroom_list(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
) -> list[str]:
    return [item["name"] for item in _build_room_inventory(request_data, store)]


def _build_room_inventory(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
) -> list[dict[str, object]]:
    payload = store.get_scoped_mapping("classrooms", "global")
    if (not payload or not payload.get("rows")) and request_data.mappingFileIds and request_data.mappingFileIds.classrooms:
        payload = store.get_file_map(request_data.mappingFileIds.classrooms)

    rooms: list[dict[str, object]] = []
    seen: set[str] = set()
    for row in (payload or {}).get("rows", []):
        room_name = _normalize_id_token(
            row.get("class_number", "")
            or row.get("classroom", "")
            or row.get("room", "")
        )
        if not room_name or room_name in seen:
            continue
        seen.add(room_name)
        room_type = str(row.get("room_type", "") or row.get("type", "") or row.get("category", "")).strip().lower()
        is_lab_flag = str(row.get("is_lab", "")).strip().lower()
        is_lab = room_type == "lab" or is_lab_flag in {"1", "true", "yes", "y"}
        raw_capacity = row.get("capacity", row.get("room_capacity", row.get("class_capacity")))
        capacity = None
        if str(raw_capacity or "").strip():
            try:
                capacity = int(float(str(raw_capacity).strip()))
            except ValueError:
                capacity = None
        rooms.append(
            {
                "name": room_name,
                "is_lab": is_lab,
                "capacity": capacity,
            }
        )
    return rooms


def _build_section_strength_map(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
) -> dict[str, int]:
    payload = store.get_scoped_mapping("classrooms", "global")
    if (not payload or not payload.get("rows")) and request_data.mappingFileIds and request_data.mappingFileIds.classrooms:
        payload = store.get_file_map(request_data.mappingFileIds.classrooms)

    strengths: dict[str, int] = {}
    for row in (payload or {}).get("rows", []):
        section_name = str(
            _normalize_id_token(
                row.get("section_name", "")
                or row.get("section", "")
                or row.get("section_names", "")
            )
        ).strip()
        raw_strength = row.get("strength", row.get("section_strength", row.get("student_strength")))
        if not section_name or not str(raw_strength or "").strip():
            continue
        try:
            strength = int(float(str(raw_strength).strip()))
        except ValueError:
            continue
        candidate_keys = {section_name, section_name.upper()}
        parsed_year, parsed_section = _parse_year_and_section_token(section_name)
        if parsed_section:
            candidate_keys.add(parsed_section)
        if parsed_year and parsed_section:
            candidate_keys.add(f"{parsed_year}|{parsed_section}")
        for key in candidate_keys:
            if key and key not in strengths:
                strengths[key] = strength
    return strengths


def _build_section_home_room_map(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
) -> dict[str, str]:
    payload = store.get_scoped_mapping("classrooms", "global")
    if (not payload or not payload.get("rows")) and request_data.mappingFileIds and request_data.mappingFileIds.classrooms:
        payload = store.get_file_map(request_data.mappingFileIds.classrooms)

    home_rooms: dict[str, str] = {}
    for row in (payload or {}).get("rows", []):
        section_name = str(
            _normalize_id_token(
                row.get("section_name", "")
                or row.get("section", "")
                or row.get("section_names", "")
            )
        ).strip()
        room_name = _normalize_id_token(
            row.get("class_number", "")
            or row.get("classroom", "")
            or row.get("room", "")
        )
        if not section_name or not room_name:
            continue
        candidate_keys = {section_name, section_name.upper()}
        parsed_year, parsed_section = _parse_year_and_section_token(section_name)
        if parsed_section:
            candidate_keys.add(parsed_section)
        if parsed_year and parsed_section:
            candidate_keys.add(f"{parsed_year}|{parsed_section}")
        for key in candidate_keys:
            if key and key not in home_rooms:
                home_rooms[key] = room_name
    return home_rooms


def _build_fixed_classroom_blocks(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
) -> dict[tuple[str, str, str, int], str]:
    payload = store.get_scoped_mapping("fixed_classroom_blocks", "global")
    if (
        (not payload or not payload.get("rows"))
        and request_data.mappingFileIds
        and request_data.mappingFileIds.fixedClassroomBlocks
    ):
        payload = store.get_file_map(request_data.mappingFileIds.fixedClassroomBlocks)

    blocks: dict[tuple[str, str, str, int], str] = {}
    for row in (payload or {}).get("rows", []):
        year = normalize_year(str(row.get("year", "")))
        section = _normalize_id_token(row.get("section", "")).upper()
        day = _normalize_day(row.get("day", ""))
        periods = row.get("periods", [])
        classroom = _normalize_id_token(row.get("classroom", ""))
        if not year or not section or not day:
            continue
        for raw_period in periods:
            try:
                period = int(raw_period)
            except (TypeError, ValueError):
                continue
            if period < 1 or period > 7:
                continue
            blocks[(year, section, day, period)] = classroom
    return blocks


def _build_faculty_availability(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
    all_faculties: set[str],
    faculty_id_to_name: dict[str, str],
    instructional_periods: list[int],
) -> dict[str, dict[str, set[int]]]:
    availability: dict[str, dict[str, set[int]]] = {
        fid: {day: set(instructional_periods) for day in DAYS} for fid in all_faculties
    }

    reverse_faculty_name_map = {
        name.strip(): faculty_id for faculty_id, name in faculty_id_to_name.items() if name.strip()
    }

    def resolve_faculty_key(raw_id: str, raw_name: str = "") -> str:
        faculty_id = _canonicalize_faculty_token(_normalize_faculty_field(raw_id), faculty_id_to_name)
        if faculty_id:
            return faculty_id
        faculty_name = str(raw_name).strip()
        return _canonicalize_faculty_token(reverse_faculty_name_map.get(faculty_name, faculty_name), faculty_id_to_name)

    uploaded_payload = store.get_scoped_mapping("faculty_availability", "global")
    if uploaded_payload:
        uploaded_faculties: set[str] = set()
        for row in uploaded_payload.get("rows", []):
            faculty_key = resolve_faculty_key(row.get("faculty_id", ""), row.get("faculty_name", ""))
            faculty_keys = _split_faculty_tokens(faculty_key) or ((faculty_key,) if faculty_key else ())
            day = _normalize_day(str(row.get("day", "")))
            period = int(row.get("period", 0) or 0)
            if not faculty_keys or not day or period not in instructional_periods:
                continue
            for key in faculty_keys:
                availability.setdefault(key, {name: set(instructional_periods) for name in DAYS})
                if key not in uploaded_faculties:
                    availability[key] = {name: set() for name in DAYS}
                    uploaded_faculties.add(key)
                availability[key][day].add(period)

    for entry in request_data.facultyAvailability:
        faculty_keys = _split_faculty_tokens(str(entry.facultyId).strip())
        if not faculty_keys:
            continue
        for faculty_key in faculty_keys:
            availability.setdefault(faculty_key, {name: set(instructional_periods) for name in DAYS})
            for raw_day, periods in entry.availablePeriodsByDay.items():
                day = _normalize_day(raw_day)
                if not day:
                    continue
                availability[faculty_key][day] = {int(p) for p in periods if int(p) in instructional_periods}
    return availability


def _resolve_faculty_pool(
    raw_faculty_id: str,
    faculty_id_to_name: dict[str, str],
    faculty_availability: dict[str, dict[str, set[int]]],
) -> tuple[str, ...]:
    faculty_raw = str(raw_faculty_id or "").strip()
    if any(delimiter in faculty_raw for delimiter in [",", "/", "|", "+", "&"]):
        split_pool = tuple(
            _canonicalize_faculty_token(token, faculty_id_to_name)
            for token in _split_faculty_tokens(faculty_raw)
            if _canonicalize_faculty_token(token, faculty_id_to_name)
        )
        if split_pool:
            return tuple(sorted(set(split_pool)))
    faculty_token = _canonicalize_faculty_token(faculty_raw, faculty_id_to_name)
    if not faculty_token:
        return ()
    if faculty_token in faculty_id_to_name or faculty_token in faculty_availability:
        return (faculty_token,)

    normalized_matches = [
        faculty_id
        for faculty_id, faculty_name in faculty_id_to_name.items()
        if faculty_token.lower() in str(faculty_name).strip().lower()
    ]
    if normalized_matches:
        return tuple(sorted(set(normalized_matches)))

    cleaned_split = tuple(
        _canonicalize_faculty_token(token, faculty_id_to_name)
        for token in _split_faculty_tokens(faculty_token)
        if _canonicalize_faculty_token(token, faculty_id_to_name)
    )
    if len(cleaned_split) > 1:
        return tuple(sorted(set(cleaned_split)))
    return (faculty_token,)


def _pick_best_faculty_option_for_locked_session(
    faculty_options: tuple[str, ...],
    day: str,
    periods: tuple[int, ...],
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    instructional_periods: list[int],
    sections: tuple[str, ...] = (),
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]] | None = None,
    session_adjacency: dict[int, set[int]] | None = None,
    is_team: bool = False,
) -> str | None:
    if not faculty_options:
        return None

    if is_team or len(faculty_options) > 1:
        # Check that ALL co-teachers in the team are simultaneously free
        for faculty_id in faculty_options:
            allowed = faculty_availability.get(faculty_id, _default_day_availability(instructional_periods)).get(day, set(instructional_periods))
            if any(period not in allowed for period in periods):
                return None
            if any((day, period) in faculty_busy.setdefault(faculty_id, set()) for period in periods):
                return None
            if sections and faculty_section_slots is not None and session_adjacency is not None:
                if _faculty_has_consecutive_different_section_conflict(
                    faculty_id,
                    day,
                    periods,
                    _section_signature(sections),
                    faculty_section_slots,
                    session_adjacency,
                ):
                    return None
        return ",".join(faculty_options)

    # Existing single-faculty selection
    best: tuple[int, int, int, str] | None = None
    selected: str | None = None
    for faculty_id in faculty_options:
        allowed = faculty_availability.get(faculty_id, _default_day_availability(instructional_periods)).get(day, set(instructional_periods))
        unavailable = sum(1 for period in periods if period not in allowed)
        conflicts = sum(1 for period in periods if (day, period) in faculty_busy.setdefault(faculty_id, set()))
        section_conflict = 0
        if sections and faculty_section_slots is not None and session_adjacency is not None:
            section_conflict = 1 if _faculty_has_consecutive_different_section_conflict(
                faculty_id,
                day,
                periods,
                _section_signature(sections),
                faculty_section_slots,
                session_adjacency,
            ) else 0
        load = _faculty_daily_load(faculty_id, day, faculty_busy)
        score = (unavailable, conflicts, section_conflict, load, faculty_id)
        if best is None or score < best:
            best = score
            selected = faculty_id
    return selected


def _choose_faculty_for_slot(
    requirement: Requirement,
    day: str,
    start_period: int,
    block_size: int,
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    instructional_periods: list[int],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]] | None = None,
    session_adjacency: dict[int, set[int]] | None = None,
    enforce_section_transition_rule: bool = True,
    faculty_daily_loads: dict[str, dict[str, int]] | None = None,
) -> str | None:
    periods = range(start_period, start_period + block_size)
    faculty_section_slots = faculty_section_slots or {}
    session_adjacency = session_adjacency or {}
    section_signature = _section_signature(requirement.sections)
    if requirement.faculty_team:
        for faculty_id in requirement.faculty_team:
            allowed_periods = faculty_availability.get(faculty_id, _default_day_availability(instructional_periods)).get(day, set(instructional_periods))
            if any(period not in allowed_periods for period in periods):
                return None
            if any((day, period) in faculty_busy.setdefault(faculty_id, set()) for period in periods):
                return None
            if enforce_section_transition_rule and _faculty_has_consecutive_different_section_conflict(
                faculty_id,
                day,
                periods,
                section_signature,
                faculty_section_slots,
                session_adjacency,
            ):
                return None
        return ",".join(requirement.faculty_team)

    best_faculty: str | None = None
    best_key: tuple[int, int, str] | None = None
    faculty_options = requirement.faculty_options or ((requirement.faculty_id,) if requirement.faculty_id else ())
    for faculty_id in faculty_options:
        allowed_periods = faculty_availability.get(faculty_id, _default_day_availability(instructional_periods)).get(day, set(instructional_periods))
        if any(period not in allowed_periods for period in periods):
            continue
        if any((day, period) in faculty_busy.setdefault(faculty_id, set()) for period in periods):
            continue
        if enforce_section_transition_rule and _faculty_has_consecutive_different_section_conflict(
            faculty_id,
            day,
            periods,
            section_signature,
            faculty_section_slots,
            session_adjacency,
        ):
            continue
        key = (
            faculty_daily_loads.get(faculty_id, {}).get(day, 0)
            if faculty_daily_loads is not None
            else _faculty_daily_load(faculty_id, day, faculty_busy),
            _faculty_weekly_capacity(faculty_id, faculty_availability, instructional_periods),
            faculty_id,
        )
        if best_key is None or key < best_key:
            best_key = key
            best_faculty = faculty_id
    return best_faculty


def _candidate_block_sizes(
    remaining_hours: int,
    min_consecutive_hours: int,
    max_consecutive_hours: int,
) -> list[int]:
    if remaining_hours <= 0:
        return []
    minimum = max(1, min_consecutive_hours)
    maximum = max(minimum, max_consecutive_hours)
    if remaining_hours < minimum:
        return [remaining_hours]
    upper_bound = min(remaining_hours, maximum)
    sizes = list(range(upper_bound, minimum - 1, -1))
    remainder = remaining_hours % minimum
    if remainder > 0 and remainder not in sizes:
        sizes.append(remainder)
    return sizes


def _is_allowed_compulsory_block_placement(
    requirement: Requirement,
    start_period: int,
    block_size: int,
    sessions: list[tuple[int, ...]],
) -> bool:
    if block_size <= 1 or requirement.min_consecutive_hours <= 1:
        return True

    end_period = start_period + block_size - 1
    
    if block_size in (2, 3):
        for s in sessions:
            if start_period in s and end_period in s:
                return True
        return False
        
    if block_size == 4:
        start_s = -1
        end_s = -1
        for i, s in enumerate(sessions):
            if start_period in s: start_s = i
            if end_period in s: end_s = i
        if start_s == -1 or end_s == -1:
            return False
        # Can be in same session or span 2 consecutive sessions
        return end_s == start_s or end_s == start_s + 1

    return True


def _default_day_availability(periods: list[int]) -> dict[str, set[int]]:
    return {day: set(periods) for day in DAYS}


def _build_session_adjacency(sessions: list[tuple[int, ...]]) -> dict[int, set[int]]:
    adjacency: dict[int, set[int]] = {}
    for session in sessions:
        ordered = list(session)
        for idx, period in enumerate(ordered):
            if idx > 0:
                adjacency.setdefault(period, set()).add(ordered[idx - 1])
            if idx + 1 < len(ordered):
                adjacency.setdefault(period, set()).add(ordered[idx + 1])
    return adjacency


def _section_signature(sections: tuple[str, ...] | list[str]) -> tuple[str, ...]:
    return tuple(sorted(str(section).strip() for section in sections if str(section).strip()))


def _faculty_has_consecutive_different_section_conflict(
    faculty_id: str,
    day: str,
    periods: range | list[int] | tuple[int, ...],
    section_signature: tuple[str, ...],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
    session_adjacency: dict[int, set[int]],
) -> bool:
    assigned_slots = faculty_section_slots.get(faculty_id, {})
    period_set = {int(period) for period in periods}
    for period in period_set:
        for adjacent_period in session_adjacency.get(period, set()):
            if adjacent_period in period_set:
                continue
            adjacent_sections = assigned_slots.get((day, adjacent_period))
            if adjacent_sections and adjacent_sections != section_signature:
                return True
    return False


def _mark_faculty_section_assignment(
    faculty_ids: tuple[str, ...],
    day: str,
    periods: list[int] | tuple[int, ...] | range,
    sections: tuple[str, ...] | list[str],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
) -> None:
    signature = _section_signature(sections)
    if not signature:
        return
    for faculty_id in faculty_ids:
        slot_map = faculty_section_slots.setdefault(faculty_id, {})
        for period in periods:
            slot_map[(day, int(period))] = signature


def _unmark_faculty_section_assignment(
    faculty_ids: tuple[str, ...],
    day: str,
    periods: list[int] | tuple[int, ...] | range,
    sections: tuple[str, ...] | list[str],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
) -> None:
    signature = _section_signature(sections)
    for faculty_id in faculty_ids:
        slot_map = faculty_section_slots.setdefault(faculty_id, {})
        for period in periods:
            key = (day, int(period))
            if slot_map.get(key) == signature:
                slot_map.pop(key, None)


def _sections_can_host_block(
    sections: tuple[str, ...],
    requirement: Requirement,
    day: str,
    start_period: int,
    block_size: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
    sessions: list[tuple[int, ...]],
) -> bool:
    if not instructional_periods:
        return False
    if start_period + block_size - 1 > instructional_periods[-1]:
        return False
    if not _is_allowed_compulsory_block_placement(requirement, start_period, block_size, sessions):
        return False
    periods = range(start_period, start_period + block_size)
    for period in periods:
        if period not in instructional_periods:
            return False
        for section in sections:
            if schedules[(year, section)][day].get(period) is not None:
                return False
    return True


def _sections_are_free(
    sections: tuple[str, ...],
    requirement: Requirement,
    day: str,
    start_period: int,
    block_size: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
    sessions: list[tuple[int, ...]],
    daily_subject_counts: dict[str, dict[str, dict[str, int]]] | None = None,
) -> bool:
    if not _sections_can_host_block(
        sections,
        requirement,
        day,
        start_period,
        block_size,
        schedules,
        year,
        instructional_periods,
        sessions,
    ):
        return False
    if daily_subject_counts is not None and _exceeds_subject_daily_hard_limit(
        requirement,
        day,
        block_size,
        daily_subject_counts,
    ):
        return False
    return True


def _slot_is_free(
    sections: tuple[str, ...],
    requirement: Requirement,
    day: str,
    start_period: int,
    block_size: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    year: str,
    instructional_periods: list[int],
    sessions: list[tuple[int, ...]],
    daily_subject_counts: dict[str, dict[str, dict[str, int]]] | None = None,
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]] | None = None,
    session_adjacency: dict[int, set[int]] | None = None,
) -> bool:
    if not _sections_are_free(
        sections,
        requirement,
        day,
        start_period,
        block_size,
        schedules,
        year,
        instructional_periods,
        sessions,
        daily_subject_counts,
    ):
        return False
    selected_faculty = _choose_faculty_for_slot(
        requirement,
        day,
        start_period,
        block_size,
        faculty_busy,
        faculty_availability,
        instructional_periods,
        faculty_section_slots,
        session_adjacency,
    )
    return selected_faculty is not None


def _count_free_periods_for_sections(
    sections: tuple[str, ...],
    day: str,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
) -> int:
    free_count = 0
    for period in instructional_periods:
        if all(schedules[(year, section)][day].get(period) is None for section in sections):
            free_count += 1
    return free_count


def _count_available_periods_for_faculty(
    faculty_id: str,
    day: str,
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    instructional_periods: list[int],
) -> int:
    allowed_periods = faculty_availability.get(faculty_id, _default_day_availability(instructional_periods)).get(day, set(instructional_periods))
    return sum(1 for period in allowed_periods if (day, period) not in faculty_busy.setdefault(faculty_id, set()))


def _faculty_weekly_capacity(
    faculty_id: str,
    faculty_availability: dict[str, dict[str, set[int]]],
    instructional_periods: list[int],
) -> int:
    day_map = faculty_availability.get(faculty_id, _default_day_availability(instructional_periods))
    return sum(len(periods) for periods in day_map.values())


def _requirement_weekly_capacity(
    requirement: Requirement,
    faculty_availability: dict[str, dict[str, set[int]]],
    instructional_periods: list[int],
) -> int:
    if requirement.faculty_team:
        capacities = [_faculty_weekly_capacity(fid, faculty_availability, instructional_periods) for fid in requirement.faculty_team]
        return min(capacities) if capacities else 0
    faculty_options = requirement.faculty_options or ((requirement.faculty_id,) if requirement.faculty_id else ())
    if not faculty_options:
        return len(DAYS) * len(instructional_periods)
    return max(_faculty_weekly_capacity(faculty_id, faculty_availability, instructional_periods) for faculty_id in faculty_options)


def _faculty_daily_load(
    faculty_id: str,
    day: str,
    faculty_busy: dict[str, set[tuple[str, int]]],
) -> int:
    return sum(1 for busy_day, _ in faculty_busy.setdefault(faculty_id, set()) if busy_day == day)


def _subject_daily_load(
    requirement: Requirement,
    day: str,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
) -> int:
    count = 0
    for section in requirement.sections:
        for period in instructional_periods:
            entry = schedules[(year, section)][day].get(period)
            if not entry:
                continue
            slot_subject_id = str(entry.get("subjectId", "")).strip()
            slot_subject = str(entry.get("subject", "")).strip()
            if slot_subject_id == requirement.subject_id or slot_subject == requirement.subject_id:
                count += 1
    return count


def _subject_daily_load_by_section(
    requirement: Requirement,
    day: str,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
) -> dict[str, int]:
    result: dict[str, int] = {}
    for section in requirement.sections:
        count = 0
        for period in instructional_periods:
            entry = schedules[(year, section)][day].get(period)
            if not entry:
                continue
            slot_subject_id = str(entry.get("subjectId", "")).strip()
            slot_subject = str(entry.get("subject", "")).strip()
            if slot_subject_id == requirement.subject_id or slot_subject == requirement.subject_id:
                count += 1
        result[section] = count
    return result


def _daily_subject_count(
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
    section: str,
    day: str,
    subject_id: str,
) -> int:
    return int(daily_subject_counts.get(section, {}).get(day, {}).get(subject_id, 0) or 0)


def _daily_subject_counts_for_requirement(
    requirement: Requirement,
    day: str,
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
) -> dict[str, int]:
    return {
        section: _daily_subject_count(daily_subject_counts, section, day, requirement.subject_id)
        for section in requirement.sections
    }


def _exceeds_subject_daily_hard_limit(
    requirement: Requirement,
    day: str,
    block_size: int,
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
) -> bool:
    if requirement.daily_limit_exempt:
        return False
    return any(
        count + block_size > requirement.hard_daily_limit
        for count in _daily_subject_counts_for_requirement(requirement, day, daily_subject_counts).values()
    )


def _subject_daily_limit_penalty(
    requirement: Requirement,
    day: str,
    block_size: int,
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
) -> int:
    if requirement.daily_limit_exempt:
        return 0
    overflow = 0
    impacted_sections = 0
    for load in _daily_subject_counts_for_requirement(requirement, day, daily_subject_counts).values():
        next_total = load + block_size
        if next_total > requirement.preferred_daily_limit:
            overflow += next_total - requirement.preferred_daily_limit
            impacted_sections += 1
    if overflow <= 0:
        return 0
    return -(40 * overflow + 12 * impacted_sections)


def _subject_has_edge_period_for_section(
    section: str,
    subject_id: str,
    edge_period: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
) -> bool:
    for day in DAYS:
        entry = schedules[(year, section)][day].get(edge_period)
        if not entry:
            continue
        slot_subject_id = str(entry.get("subjectId", "")).strip()
        slot_subject = str(entry.get("subject", "")).strip()
        if slot_subject_id == subject_id or slot_subject == subject_id:
            return True
    return False


def _subject_edge_distribution_adjustment(
    requirement: Requirement,
    start_period: int,
    block_size: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
    subject_edge_counts: dict[str, dict[str, dict[int, int]]] | None = None,
) -> int:
    if not instructional_periods:
        return 0
    end_period = start_period + block_size - 1
    first_period = instructional_periods[0]
    last_period = instructional_periods[-1]
    touches_first = start_period == first_period
    touches_last = end_period == last_period
    if not touches_first and not touches_last:
        return 0

    score = 0
    for section in requirement.sections:
        if touches_first:
            has_first = (
                bool(subject_edge_counts.get(section, {}).get(requirement.subject_id, {}).get(first_period, 0))
                if subject_edge_counts is not None
                else _subject_has_edge_period_for_section(section, requirement.subject_id, first_period, schedules, year)
            )
            if has_first:
                score -= 2
            else:
                score += 8
        if touches_last:
            has_last = (
                bool(subject_edge_counts.get(section, {}).get(requirement.subject_id, {}).get(last_period, 0))
                if subject_edge_counts is not None
                else _subject_has_edge_period_for_section(section, requirement.subject_id, last_period, schedules, year)
            )
            if has_last:
                score -= 2
            else:
                score += 8
    return score


def _remaining_hours_per_section(
    requirements: list[Requirement],
    remaining_hours: dict[int, int],
) -> dict[str, int]:
    section_hours: dict[str, int] = {}
    for req_idx, requirement in enumerate(requirements):
        remaining = remaining_hours.get(req_idx, 0)
        if remaining <= 0:
            continue
        for section in requirement.sections:
            section_hours[section] = section_hours.get(section, 0) + remaining
    return section_hours


def _section_capacity_is_feasible(
    requirements: list[Requirement],
    remaining_hours: dict[int, int],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
) -> bool:
    remaining_by_section = _remaining_hours_per_section(requirements, remaining_hours)
    for section, remaining in remaining_by_section.items():
        free_slots = sum(
            1
            for day in DAYS
            for period in instructional_periods
            if schedules[(year, section)][day].get(period) is None
        )
        if remaining > free_slots:
            return False
    return True


def _compute_timeout_seconds(section_count: int, requirement_count: int) -> int:
    # §18: adaptive timeout – 30s small, 90s medium, 180s large
    if section_count <= 5:
        base = 30
    elif section_count <= 10:
        base = 90
    else:
        base = 180
    if requirement_count >= 20:
        base += 30
    if requirement_count >= 35:
        base += 30
    if requirement_count >= 50:
        base += 30
    if requirement_count >= 70:
        base += 30
    return min(300, base)


def _requirement_priority(
    requirement: Requirement,
    faculty_availability: dict[str, dict[str, set[int]]],
    instructional_periods: list[int],
) -> tuple[int, int, int, int, int, str, str]:
    weekly_capacity = _requirement_weekly_capacity(requirement, faculty_availability, instructional_periods)
    strict_availability = len(DAYS) * len(instructional_periods) - weekly_capacity
    return (
        requirement.phase,
        0 if requirement.common_faculty else 1,
        -requirement.hours,
        -requirement.min_consecutive_hours,
        -strict_availability,
        requirement.subject_id,
        ",".join(requirement.sections),
    )


def _determine_requirement_phase(
    shared: bool,
    min_consecutive_hours: int,
    hours: int,
    faculty_options: tuple[str, ...],
    faculty_availability: dict[str, dict[str, set[int]]],
) -> int:
    if shared:
        return 1
    if hours >= 4:
        return 2
    if min_consecutive_hours > 1:
        return 3
    return 4


def _requirement_faculty_tokens(requirement: Requirement) -> tuple[str, ...]:
    if requirement.faculty_team:
        return tuple(
            _normalize_id_token(faculty_id)
            for faculty_id in requirement.faculty_team
            if _normalize_id_token(faculty_id)
        )
    if requirement.faculty_options:
        return tuple(
            _normalize_id_token(faculty_id)
            for faculty_id in requirement.faculty_options
            if _normalize_id_token(faculty_id)
        )
    return _split_faculty_tokens(requirement.faculty_id)


def _extract_faculty_occupancy_from_timetable(record: dict | None) -> dict[str, set[tuple[str, int]]]:
    return _extract_faculty_occupancy_from_timetable_with_map(record, {})


def _extract_faculty_occupancy_from_timetable_with_map(
    record: dict | None,
    faculty_id_to_name: dict[str, str],
) -> dict[str, set[tuple[str, int]]]:
    occupancy: dict[str, set[tuple[str, int]]] = {}
    if not record:
        return occupancy

    all_grids = record.get("allGrids")
    if not isinstance(all_grids, dict):
        section = str(record.get("section", "")).strip()
        grid = record.get("grid")
        all_grids = {section: grid} if section and isinstance(grid, dict) else {}

    for grid in all_grids.values():
        if not isinstance(grid, dict):
            continue
        for raw_day, slots in grid.items():
            day = _normalize_day(str(raw_day))
            if not day or not isinstance(slots, list):
                continue
            for period_index, slot in enumerate(slots, start=1):
                if not isinstance(slot, dict):
                    continue
                raw_faculty_values = [
                    slot.get("facultyId") or "",
                    slot.get("facultyName") or "",
                    slot.get("faculty") or "",
                ]
                canonical_tokens: set[str] = set()
                for raw_value in raw_faculty_values:
                    for token in _split_faculty_tokens(str(raw_value)):
                        canonical = _canonicalize_faculty_token(token, faculty_id_to_name)
                        if canonical:
                            canonical_tokens.add(canonical)
                for faculty_id in canonical_tokens:
                    occupancy.setdefault(faculty_id, set()).add((day, period_index))
    return occupancy


def _extract_faculty_section_assignments_from_timetable_with_map(
    record: dict | None,
    faculty_id_to_name: dict[str, str],
) -> dict[str, dict[tuple[str, int], tuple[str, ...]]]:
    assignments: dict[str, dict[tuple[str, int], tuple[str, ...]]] = {}
    if not record:
        return assignments

    all_grids = record.get("allGrids")
    if not isinstance(all_grids, dict):
        section = str(record.get("section", "")).strip()
        grid = record.get("grid")
        all_grids = {section: grid} if section and isinstance(grid, dict) else {}

    for section_name, grid in all_grids.items():
        if not isinstance(grid, dict):
            continue
        normalized_section = str(section_name).strip()
        for raw_day, slots in grid.items():
            day = _normalize_day(str(raw_day))
            if not day or not isinstance(slots, list):
                continue
            for period_index, slot in enumerate(slots, start=1):
                if not isinstance(slot, dict):
                    continue
                raw_faculty_values = [
                    slot.get("facultyId") or "",
                    slot.get("facultyName") or "",
                    slot.get("faculty") or "",
                ]
                shared_sections = [
                    str(shared_section).strip()
                    for shared_section in (slot.get("sharedSections") or [])
                    if str(shared_section).strip()
                ]
                section_signature = _section_signature(shared_sections or [normalized_section])
                for raw_value in raw_faculty_values:
                    for token in _split_faculty_tokens(str(raw_value)):
                        canonical = _canonicalize_faculty_token(token, faculty_id_to_name)
                        if not canonical:
                            continue
                        assignments.setdefault(canonical, {})[(day, period_index)] = section_signature
    return assignments


def _build_prior_faculty_occupancy(
    timetable_ids: list[str],
    store: MemoryStore,
    faculty_id_to_name: dict[str, str],
) -> dict[str, set[tuple[str, int]]]:
    occupancy: dict[str, set[tuple[str, int]]] = {}
    for timetable_id in timetable_ids:
        record = store.get_timetable(str(timetable_id).strip())
        extracted = _extract_faculty_occupancy_from_timetable_with_map(record, faculty_id_to_name)
        for faculty_id, slots in extracted.items():
            occupancy.setdefault(faculty_id, set()).update(slots)
    return occupancy


def _build_prior_faculty_section_assignments(
    timetable_ids: list[str],
    store: MemoryStore,
    faculty_id_to_name: dict[str, str],
) -> dict[str, dict[tuple[str, int], tuple[str, ...]]]:
    assignments: dict[str, dict[tuple[str, int], tuple[str, ...]]] = {}
    for timetable_id in timetable_ids:
        record = store.get_timetable(str(timetable_id).strip())
        extracted = _extract_faculty_section_assignments_from_timetable_with_map(record, faculty_id_to_name)
        for faculty_id, slot_map in extracted.items():
            assignments.setdefault(faculty_id, {}).update(slot_map)
    return assignments


def _extract_room_occupancy_from_timetable(record: dict | None) -> dict[str, set[tuple[str, int]]]:
    occupancy: dict[str, set[tuple[str, int]]] = {}
    if not record:
        return occupancy

    all_grids = record.get("allGrids")
    if not isinstance(all_grids, dict):
        section = str(record.get("section", "")).strip()
        grid = record.get("grid")
        all_grids = {section: grid} if section and isinstance(grid, dict) else {}

    for grid in all_grids.values():
        if not isinstance(grid, dict):
            continue
        for raw_day, slots in grid.items():
            day = _normalize_day(str(raw_day))
            if not day or not isinstance(slots, list):
                continue
            for period_index, slot in enumerate(slots, start=1):
                if not isinstance(slot, dict):
                    continue
                # For cross-year room reuse, a mixed lab-session should reserve the lab first.
                # The optional classroom assignment for that same slot must not block another year.
                room = _normalize_id_token(
                    slot.get("fallbackLab")
                    or slot.get("labRoom")
                    or slot.get("venue")
                    or slot.get("classroom")
                    or ""
                )
                if room:
                    occupancy.setdefault(room, set()).add((day, period_index))
    return occupancy


def _build_prior_room_occupancy(
    timetable_ids: list[str],
    store: MemoryStore,
) -> dict[str, set[tuple[str, int]]]:
    occupancy: dict[str, set[tuple[str, int]]] = {}
    for timetable_id in timetable_ids:
        record = store.get_timetable(str(timetable_id).strip())
        extracted = _extract_room_occupancy_from_timetable(record)
        for room, slots in extracted.items():
            occupancy.setdefault(room, set()).update(slots)
    return occupancy


def _latest_timetable_ids_by_year(
    store: MemoryStore,
    exclude_year: str = "",
) -> list[str]:
    selected: list[str] = []
    seen_years: set[str] = set()
    for record in store.list_timetables():
        year = normalize_year(str(record.get("year", "")))
        timetable_id = str(record.get("id", "")).strip()
        if not timetable_id or not year or year == exclude_year or year in seen_years:
            continue
        seen_years.add(year)
        selected.append(timetable_id)
    return selected


def _is_lab_seed_record(record: dict | None) -> bool:
    if not isinstance(record, dict):
        return False
    if bool(record.get("labSeed")):
        return True
    generation_meta = record.get("generationMeta")
    return isinstance(generation_meta, dict) and bool(generation_meta.get("labSeed"))


def _build_global_section_subject_faculty_index(
    rows: list[dict],
    faculty_id_to_name: dict[str, str],
) -> dict[tuple[str, str, str], str]:
    index: dict[tuple[str, str, str], str] = {}
    for row in rows:
        year = normalize_year(str(row.get("year", "")))
        section = str(row.get("section", "")).strip()
        subject_id = _normalize_id_token(row.get("subject_id", ""))
        faculty_id = ",".join(
            _canonicalize_faculty_token(token, faculty_id_to_name)
            for token in _split_faculty_tokens(str(row.get("faculty_id", "")))
            if _canonicalize_faculty_token(token, faculty_id_to_name)
        )
        if year and section and subject_id and faculty_id:
            index[(year, section, subject_id)] = faculty_id
    return index


def _build_global_lab_occupancy(
    all_main_rows: list[dict],
    all_lab_rows: list[dict],
    current_year: str,
    faculty_id_to_name: dict[str, str],
    faculty_availability: dict[str, dict[str, set[int]]],
    instructional_periods: list[int],
) -> dict[str, set[tuple[str, int]]]:
    occupancy: dict[str, set[tuple[str, int]]] = {}
    faculty_index = _build_global_section_subject_faculty_index(all_main_rows, faculty_id_to_name)
    for row in all_lab_rows:
        lab_year = normalize_year(str(row.get("year", "")))
        if not lab_year or lab_year == current_year:
            continue
        section = str(row.get("section", "")).strip()
        subject_id = _normalize_id_token(row.get("subject_id", ""))
        day = _normalize_day(int(row.get("day", 0) or 0))
        periods = [int(period) for period in row.get("hours", []) if int(period) in instructional_periods]
        if not section or not subject_id or not day or not periods:
            continue
        raw_faculty_id = faculty_index.get((lab_year, section, subject_id), "")
        faculty_ids = _resolve_faculty_pool(raw_faculty_id, faculty_id_to_name, faculty_availability)
        for faculty_id in faculty_ids:
            occupancy.setdefault(faculty_id, set()).update((day, period) for period in periods)
    return occupancy


def _build_global_lab_room_occupancy(
    all_lab_rows: list[dict],
    current_year: str,
    instructional_periods: list[int],
) -> dict[str, set[tuple[str, int]]]:
    occupancy: dict[str, set[tuple[str, int]]] = {}
    for row in all_lab_rows:
        lab_year = normalize_year(str(row.get("year", "")))
        if not lab_year or lab_year == current_year:
            continue
        day = _normalize_day(int(row.get("day", 0) or 0))
        periods = [int(period) for period in row.get("hours", []) if int(period) in instructional_periods]
        venue = _normalize_id_token(row.get("venue") or "")
        if not day or not periods or not venue:
            continue
        occupancy.setdefault(venue, set()).update((day, period) for period in periods)
    return occupancy


def _persist_faculty_occupancy(
    store: MemoryStore,
    timetable_id: str,
    session_log: list[dict],
    instructional_periods: list[int],
) -> None:
    store.delete_occupancy_by_source(timetable_id)
    persisted: set[tuple[str, str, int]] = set()
    for session in session_log:
        day = _normalize_day(str(session.get("day", "")))
        if not day:
            continue
        periods = [int(period) for period in session.get("periods", []) if int(period) in instructional_periods]
        if not periods:
            continue
        faculty_ids = [
            _normalize_id_token(faculty_id)
            for faculty_id in session.get("faculty_ids", [])
            if _normalize_id_token(faculty_id)
        ]
        if not faculty_ids:
            faculty_ids = list(_split_faculty_tokens(str(session.get("faculty_id", ""))))
        if not faculty_ids:
            continue
        section_label = ",".join(
            sorted(str(section).strip() for section in session.get("sections", []) if str(section).strip())
        )
        year_label = str(session.get("year", "")).strip() or None
        for faculty_id in faculty_ids:
            for period in periods:
                marker = (faculty_id, day, period)
                if marker in persisted:
                    continue
                persisted.add(marker)
                store.mark_faculty_busy(
                    faculty=faculty_id,
                    day=day,
                    period=period,
                    source_id=timetable_id,
                    year=year_label,
                    section=section_label or None,
                )


def _remaining_empty_slots_for_sections(
    sections: tuple[str, ...],
    free_slots_tracker: dict[str, int],
) -> int:
    return sum(free_slots_tracker.get(section, 0) for section in sections)


def _build_daily_subject_counts(
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    sections: list[str],
    instructional_periods: list[int],
) -> dict[str, dict[str, dict[str, int]]]:
    counts: dict[str, dict[str, dict[str, int]]] = {
        section: {day: {} for day in DAYS}
        for section in sections
    }
    for section in sections:
        for day in DAYS:
            for period in instructional_periods:
                entry = schedules[(year, section)][day].get(period)
                if not isinstance(entry, dict):
                    continue
                subject_id = str(entry.get("subjectId", "")).strip() or str(entry.get("subject", "")).strip()
                if not subject_id:
                    continue
                day_counts = counts[section][day]
                day_counts[subject_id] = day_counts.get(subject_id, 0) + 1
    return counts


def _build_faculty_daily_loads(
    faculty_busy: dict[str, set[tuple[str, int]]],
) -> dict[str, dict[str, int]]:
    loads: dict[str, dict[str, int]] = {}
    for faculty_id, busy_slots in faculty_busy.items():
        day_counts: dict[str, int] = {}
        for day, _ in busy_slots:
            day_counts[day] = day_counts.get(day, 0) + 1
        loads[faculty_id] = day_counts
    return loads


def _build_subject_edge_counts(
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    sections: list[str],
    instructional_periods: list[int],
) -> dict[str, dict[str, dict[int, int]]]:
    if not instructional_periods:
        return {}
    edge_periods = {instructional_periods[0], instructional_periods[-1]}
    counts: dict[str, dict[str, dict[int, int]]] = {}
    for section in sections:
        section_counts: dict[str, dict[int, int]] = {}
        for day in DAYS:
            for period in edge_periods:
                entry = schedules[(year, section)][day].get(period)
                if not isinstance(entry, dict):
                    continue
                subject_id = str(entry.get("subjectId", "")).strip() or str(entry.get("subject", "")).strip()
                if not subject_id:
                    continue
                period_counts = section_counts.setdefault(subject_id, {})
                period_counts[period] = period_counts.get(period, 0) + 1
        if section_counts:
            counts[section] = section_counts
    return counts


def _adjust_daily_subject_counts(
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
    requirement: Requirement,
    day: str,
    block_size: int,
) -> None:
    for section in requirement.sections:
        day_counts = daily_subject_counts.setdefault(section, {}).setdefault(day, {})
        next_count = day_counts.get(requirement.subject_id, 0) + block_size
        if next_count > 0:
            day_counts[requirement.subject_id] = next_count
        else:
            day_counts.pop(requirement.subject_id, None)


def _adjust_subject_edge_counts(
    subject_edge_counts: dict[str, dict[str, dict[int, int]]],
    requirement: Requirement,
    start_period: int,
    block_size: int,
    instructional_periods: list[int],
    delta: int,
) -> None:
    if not instructional_periods:
        return
    end_period = start_period + block_size - 1
    touched_edge_periods: list[int] = []
    first_period = instructional_periods[0]
    last_period = instructional_periods[-1]
    if start_period == first_period:
        touched_edge_periods.append(first_period)
    if end_period == last_period and last_period not in touched_edge_periods:
        touched_edge_periods.append(last_period)
    if not touched_edge_periods:
        return

    for section in requirement.sections:
        section_counts = subject_edge_counts.setdefault(section, {})
        subject_counts = section_counts.setdefault(requirement.subject_id, {})
        for period in touched_edge_periods:
            next_count = subject_counts.get(period, 0) + delta
            if next_count > 0:
                subject_counts[period] = next_count
            else:
                subject_counts.pop(period, None)
        if not subject_counts:
            section_counts.pop(requirement.subject_id, None)
        if not section_counts:
            subject_edge_counts.pop(section, None)


def _score_slot_candidate(
    requirement: Requirement,
    faculty_id: str,
    day: str,
    start_period: int,
    block_size: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    year: str,
    instructional_periods: list[int],
    empty_slots: int,
    section_flex: int,
    subject_load: int,
) -> int:
    # §14: slot scoring weights – +5 section free, +5 faculty available, +3 continuous, -5 conflict
    faculty_flex = _count_available_periods_for_faculty(faculty_id, day, faculty_busy, faculty_availability, instructional_periods)
    faculty_load = _faculty_daily_load(faculty_id, day, faculty_busy)
    center_bonus = 2 if start_period not in (instructional_periods[0], instructional_periods[-1]) else 0
    # Primary scoring per spec
    section_free_score = 5 if section_flex > 0 else -5
    faculty_available_score = 5 if faculty_flex > 0 else -5
    continuous_score = 3 if block_size >= requirement.min_consecutive_hours else -10
    conflict_penalty = -5 if faculty_load + block_size > 5 else 0  # penalise overloaded days
    # Secondary heuristics
    shared_bonus = 2 if requirement.shared else 0
    longer_block_bonus = block_size * 2
    faculty_load_penalty = -3 * max(0, faculty_load + block_size - 4)
    clustering_penalty = -2 * subject_load
    scarcity_bonus = max(0, 8 - faculty_flex)
    tail_fill_bonus = 3 if empty_slots <= len(requirement.sections) * 8 else 0
    return (
        section_free_score
        + faculty_available_score
        + continuous_score
        + conflict_penalty
        + shared_bonus
        + longer_block_bonus
        + center_bonus
        + scarcity_bonus
        + tail_fill_bonus
        + section_flex
        + faculty_flex
        + faculty_load_penalty
        + clustering_penalty
    )


def _score_slot_candidate_fast(
    requirement: Requirement,
    faculty_id: str,
    day: str,
    start_period: int,
    block_size: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    year: str,
    instructional_periods: list[int],
) -> int:
    """
    Performance-focused scoring:
    - avoids expensive per-candidate metrics (subject_daily_load, empty slot scans, etc.)
    - keeps the spec's spirit: prefer section+faculty freedom and longer blocks
    """
    # Slots are already filtered to be free by the caller, so we treat "free" as boolean.
    section_free = 1
    faculty_free = 1
    # Mirror original heuristic: if block_size is too small for the requirement's
    # minimum consecutive hours, heavily penalize it.
    continuous_score = 3 if block_size >= requirement.min_consecutive_hours else -10

    # Light flex metric to differentiate candidates.
    section_flex = 0
    for p in instructional_periods:
        if all(schedules[(year, sec)][day][p] is None for sec in requirement.sections):
            section_flex += 1

    faculty_flex = 0
    allowed_periods = faculty_availability.get(faculty_id, _default_day_availability(instructional_periods)).get(day, set(instructional_periods))
    busy_slots = faculty_busy.get(faculty_id, set())
    for p in allowed_periods:
        if (day, p) not in busy_slots:
            faculty_flex += 1

    center_bonus = 2 if start_period not in (instructional_periods[0], instructional_periods[-1]) else 0
    shared_bonus = 2 if requirement.shared else 0

    # Core weights from spec.
    score = 0
    score += 5 * section_free
    score += 5 * faculty_free
    score += continuous_score

    # Cheap differentiators.
    score += section_flex
    score += faculty_flex
    # Prefer scarce faculty availability (approximation of the original solver).
    score += max(0, 8 - faculty_flex)
    score += (block_size * 2)  # prefer longer blocks
    score += center_bonus
    score += shared_bonus
    return int(score)


def _enumerate_slot_candidates(
    requirement: Requirement,
    remaining_hours: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
    year: str,
    days_order: list[str],
    periods_order: list[int],
    instructional_periods: list[int],
    sessions: list[tuple[int, ...]],
    session_adjacency: dict[int, set[int]],
    candidate_limit: int,
    free_slots_tracker: dict[str, int],
    faculty_daily_loads: dict[str, dict[str, int]] | None = None,
    subject_edge_counts: dict[str, dict[str, dict[int, int]]] | None = None,
) -> list[SlotCandidate]:
    """
    Optimized candidate enumeration:
    - EARLY EXIT: stops all loops once candidate_limit is reached
    - INLINE SCORING: lightweight score avoids function-call overhead
      and skips expensive _subject_daily_load entirely
    - SECTIONS FIRST: cheap grid check before expensive faculty lookup
    - DETERMINISTIC: sort includes faculty_id for stable tie-breaking
    """
    candidates: list[SlotCandidate] = []
    # Pre-compute once per call (O(1) via tracker)
    empty_slots = _remaining_empty_slots_for_sections(requirement.sections, free_slots_tracker)
    tail_fill = 3 if empty_slots <= len(requirement.sections) * 8 else 0
    shared_bonus = 2 if requirement.shared else 0
    first_period = instructional_periods[0] if instructional_periods else 1
    last_period = instructional_periods[-1] if instructional_periods else 7
    hit_limit = False

    for block_size in _candidate_block_sizes(
        remaining_hours,
        requirement.min_consecutive_hours,
        requirement.max_consecutive_hours,
    ):
        if hit_limit:
            break
        continuous_score = 3 if block_size >= requirement.min_consecutive_hours else -10
        longer_block_bonus = block_size * 2

        for day in days_order:
            if hit_limit:
                break
            # Pre-compute section flex ONCE per day (avoids repeat in inner loop)
            section_flex = _count_free_periods_for_sections(
                requirement.sections, day, schedules, year, instructional_periods
            )
            if section_flex == 0:
                continue  # no free slots for any section on this day

            for start_period in periods_order:
                # 1. Cheap section-grid check first
                if not _sections_are_free(
                    requirement.sections, requirement, day,
                    start_period, block_size, schedules,
                    year, instructional_periods, sessions, daily_subject_counts,
                ):
                    continue

                # 2. Faculty selection (expensive — only if sections are free)
                faculty_id = _choose_faculty_for_slot(
                    requirement, day, start_period, block_size,
                    faculty_busy, faculty_availability, instructional_periods,
                    faculty_section_slots, session_adjacency,
                    faculty_daily_loads=faculty_daily_loads,
                )
                if not faculty_id:
                    continue

                assigned_faculty_ids = _split_faculty_tokens(faculty_id)
                scoring_fid = assigned_faculty_ids[0] if assigned_faculty_ids else faculty_id

                # 3. INLINE FAST SCORING — avoids function call overhead
                #    Skips _subject_daily_load (expensive O(sections×periods) scan)
                busy_set = faculty_busy.get(scoring_fid, set())
                f_avail = faculty_availability.get(
                    scoring_fid, _default_day_availability(instructional_periods)
                ).get(day, set(instructional_periods))
                faculty_flex = sum(1 for p in f_avail if (day, p) not in busy_set)
                faculty_load = (
                    faculty_daily_loads.get(scoring_fid, {}).get(day, 0)
                    if faculty_daily_loads is not None
                    else sum(1 for bd, _ in busy_set if bd == day)
                )

                score = (
                    (5 if section_flex > 0 else -5)
                    + (5 if faculty_flex > 0 else -5)
                    + continuous_score
                    + (-5 if faculty_load + block_size > 5 else 0)
                    + shared_bonus
                    + longer_block_bonus
                    + (2 if start_period not in (first_period, last_period) else 0)
                    + max(0, 8 - faculty_flex)
                    + tail_fill
                    + section_flex
                    + faculty_flex
                    + (-3 * max(0, faculty_load + block_size - 4))
                    + _subject_daily_limit_penalty(
                        requirement,
                        day,
                        block_size,
                        daily_subject_counts,
                    )
                    + _subject_edge_distribution_adjustment(
                        requirement,
                        start_period,
                        block_size,
                        schedules,
                        year,
                        instructional_periods,
                        subject_edge_counts,
                    )
                )

                candidates.append(SlotCandidate(
                    day=day,
                    start_period=start_period,
                    block_size=block_size,
                    faculty_id=scoring_fid,
                    faculty_ids=assigned_faculty_ids or ((scoring_fid,) if scoring_fid else ()),
                    score=score,
                ))

                # EARLY EXIT — stop searching once we have enough candidates
                if len(candidates) >= candidate_limit * 2:
                    hit_limit = True
                    break

    # Deterministic sort: includes faculty_id so ties are always broken the same way
    candidates.sort(key=lambda c: (-c.score, -c.block_size, c.day, c.start_period, c.faculty_id))
    return candidates[:candidate_limit]


def _select_next_requirement(
    requirements: list[Requirement],
    remaining_hours: dict[int, int],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
    year: str,
    days_order: list[str],
    periods_order: list[int],
    instructional_periods: list[int],
    sessions: list[tuple[int, ...]],
    session_adjacency: dict[int, set[int]],
    candidate_limit: int,
    free_slots_tracker: dict[str, int],
    faculty_daily_loads: dict[str, dict[str, int]] | None = None,
    subject_edge_counts: dict[str, dict[str, dict[int, int]]] | None = None,
) -> tuple[int | None, list[SlotCandidate]]:
    """
    MRV (Minimum Remaining Values) heuristic:
    - Pick the requirement with the FEWEST valid candidates (most constrained)
    - FAIL FAST: if any active requirement has 0 candidates → return it
      immediately so the backtracker can undo the last placement
    - BREAK EARLY: if a requirement has exactly 1 candidate, pick it
      immediately (forced move, no point scanning further)
    """
    best_idx: int | None = None
    best_candidates: list[SlotCandidate] = []
    best_key: tuple[int, int, int, int, str, str] | None = None

    for req_idx, requirement in enumerate(requirements):
        remaining = remaining_hours.get(req_idx, 0)
        if remaining <= 0 or not requirement.faculty_id:
            continue

        candidates = _enumerate_slot_candidates(
            requirement, remaining, schedules, daily_subject_counts, faculty_busy,
            faculty_availability, faculty_section_slots, year, days_order, periods_order,
            instructional_periods, sessions, session_adjacency, candidate_limit,
            free_slots_tracker,
            faculty_daily_loads,
            subject_edge_counts,
        )

        # FAIL FAST: 0 candidates means this branch is dead
        if len(candidates) == 0:
            return req_idx, []

        # Deterministic MRV key — avoids expensive _requirement_weekly_capacity
        key = (
            len(candidates),                       # fewer candidates = more constrained
            0 if requirement.common_faculty else 1, # common faculty first
            -requirement.min_consecutive_hours,     # harder blocks first
            -remaining,                             # more hours remaining first
            requirement.subject_id,                 # stable tie-break
            ",".join(requirement.sections),          # stable tie-break
        )
        if best_key is None or key < best_key:
            best_idx = req_idx
            best_key = key
            best_candidates = candidates

            # BREAK EARLY: exactly 1 candidate = forced move, no point scanning more
            if len(candidates) == 1:
                break

    return best_idx, best_candidates


def _place_block(
    requirement: Requirement,
    faculty_ids: tuple[str, ...],
    day: str,
    start_period: int,
    block_size: int,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    year: str,
    subject_id_to_name: dict[str, str],
    faculty_id_to_name: dict[str, str],
    session_log: list[dict],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
    free_slots_tracker: dict[str, int],
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
    faculty_daily_loads: dict[str, dict[str, int]],
    subject_edge_counts: dict[str, dict[str, dict[int, int]]],
    instructional_periods: list[int],
    source: str = "solver",
) -> list[tuple[str, int]]:
    subject_id, subject_name = _resolve_subject_output(requirement.subject_id, subject_id_to_name)
    if not faculty_ids:
        faculty_ids = _split_faculty_tokens(requirement.faculty_id)
    faculty_ids = tuple(_normalize_id_token(fid) for fid in faculty_ids if _normalize_id_token(fid))
    faculty_id, faculty_name = _resolve_faculty_display(faculty_ids, faculty_id_to_name)
    periods = list(range(start_period, start_period + block_size))
    _mark_faculty_section_assignment(
        faculty_ids,
        day,
        periods,
        requirement.sections,
        faculty_section_slots,
    )
    for period in periods:
        for faculty_token in faculty_ids:
            faculty_busy.setdefault(faculty_token, set()).add((day, period))
            faculty_daily_loads.setdefault(faculty_token, {})
            faculty_daily_loads[faculty_token][day] = faculty_daily_loads[faculty_token].get(day, 0) + 1
        for section in requirement.sections:
            schedules[(year, section)][day][period] = {
                "subject": subject_name,
                "subjectName": subject_name,
                "subjectId": subject_id,
                "faculty": faculty_name,
                "facultyName": faculty_name,
                "facultyId": faculty_id,
                "isLab": False,
                "locked": False,
                "venue": "",
                "sharedSections": list(requirement.sections) if len(requirement.sections) > 1 else [],
            }
            if section in free_slots_tracker:
                free_slots_tracker[section] -= 1
    _adjust_daily_subject_counts(daily_subject_counts, requirement, day, block_size)
    _adjust_subject_edge_counts(
        subject_edge_counts,
        requirement,
        start_period,
        block_size,
        instructional_periods,
        1,
    )
    # §21: tag source so only file-driven sessions appear in the shared class report
    session_log.append(
        {
            "year": year,
            "subject_id": subject_id,
            "subject_name": subject_name,
            "faculty_id": faculty_id,
            "faculty_name": faculty_name,
            "faculty_ids": list(faculty_ids),
            "faculty_names": [faculty_id_to_name.get(fid, fid) for fid in faculty_ids],
            "sections": list(requirement.sections),
            "day": day,
            "periods": periods,
            "venue": "",
            "isLab": False,
            "shared": len(requirement.sections) > 1,
            "source": source,
        }
    )
    return [(day, period) for period in periods]


def _undo_block(
    requirement: Requirement,
    faculty_ids: tuple[str, ...],
    placements: list[tuple[str, int]],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    year: str,
    session_log: list[dict],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
    free_slots_tracker: dict[str, int],
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
    faculty_daily_loads: dict[str, dict[str, int]],
    subject_edge_counts: dict[str, dict[str, dict[int, int]]],
    instructional_periods: list[int],
) -> None:
    _unmark_faculty_section_assignment(
        faculty_ids,
        placements[0][0] if placements else "",
        [period for _, period in placements],
        requirement.sections,
        faculty_section_slots,
    )
    for day, period in placements:
        for faculty_token in faculty_ids:
            faculty_busy.setdefault(faculty_token, set()).discard((day, period))
            if faculty_token in faculty_daily_loads:
                next_count = faculty_daily_loads[faculty_token].get(day, 0) - 1
                if next_count > 0:
                    faculty_daily_loads[faculty_token][day] = next_count
                else:
                    faculty_daily_loads[faculty_token].pop(day, None)
                if not faculty_daily_loads[faculty_token]:
                    faculty_daily_loads.pop(faculty_token, None)
        for section in requirement.sections:
            schedules[(year, section)][day][period] = None
            if section in free_slots_tracker:
                free_slots_tracker[section] += 1
    _adjust_daily_subject_counts(
        daily_subject_counts,
        requirement,
        placements[0][0] if placements else "",
        -len(placements),
    )
    if placements:
        periods = [period for _, period in placements]
        _adjust_subject_edge_counts(
            subject_edge_counts,
            requirement,
            min(periods),
            len(periods),
            instructional_periods,
            -1,
        )
    if session_log:
        session_log.pop()


def _can_greedily_finish_remaining_requirements(
    requirements: list[Requirement],
    remaining_hours: dict[int, int],
) -> bool:
    active_indices = [idx for idx, req in enumerate(requirements) if remaining_hours.get(idx, 0) > 0 and req.faculty_id]
    if not active_indices:
        return False
    seen_sections: set[str] = set()
    for idx in active_indices:
        requirement = requirements[idx]
        if requirement.shared or len(requirement.sections) != 1:
            return False
        if requirement.min_consecutive_hours != 1 or requirement.max_consecutive_hours != 1:
            return False
        section = requirement.sections[0]
        if section in seen_sections:
            return False
        seen_sections.add(section)
    return True


def _finish_simple_section_requirements(
    requirements: list[Requirement],
    remaining_hours: dict[int, int],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
    year: str,
    days_order: list[str],
    periods_order: list[int],
    instructional_periods: list[int],
    sessions: list[tuple[int, ...]],
    session_adjacency: dict[int, set[int]],
    free_slots_tracker: dict[str, int],
    faculty_daily_loads: dict[str, dict[str, int]],
    subject_edge_counts: dict[str, dict[str, dict[int, int]]],
    subject_id_to_name: dict[str, str],
    faculty_id_to_name: dict[str, str],
    session_log: list[dict],
) -> bool:
    placed_blocks: list[tuple[int, tuple[str, ...], list[tuple[str, int]]]] = []
    active_indices = [idx for idx, req in enumerate(requirements) if remaining_hours.get(idx, 0) > 0 and req.faculty_id]
    try:
        for req_idx in active_indices:
            requirement = requirements[req_idx]
            while remaining_hours.get(req_idx, 0) > 0:
                candidates = _enumerate_slot_candidates(
                    requirement,
                    remaining_hours[req_idx],
                    schedules,
                    daily_subject_counts,
                    faculty_busy,
                    faculty_availability,
                    faculty_section_slots,
                    year,
                    days_order,
                    periods_order,
                    instructional_periods,
                    sessions,
                    session_adjacency,
                    1,
                    free_slots_tracker,
                    faculty_daily_loads,
                    subject_edge_counts,
                )
                if not candidates:
                    return False
                cand = candidates[0]
                placements = _place_block(
                    requirement,
                    cand.faculty_ids,
                    cand.day,
                    cand.start_period,
                    cand.block_size,
                    schedules,
                    faculty_busy,
                    year,
                    subject_id_to_name,
                    faculty_id_to_name,
                    session_log,
                    faculty_section_slots,
                    free_slots_tracker,
                    daily_subject_counts,
                    faculty_daily_loads,
                    subject_edge_counts,
                    instructional_periods,
                    source="solver",
                )
                remaining_hours[req_idx] -= cand.block_size
                placed_blocks.append((req_idx, cand.faculty_ids, placements))
        return True
    finally:
        if any(remaining_hours.get(idx, 0) > 0 for idx in active_indices):
            for req_idx, faculty_ids, placements in reversed(placed_blocks):
                remaining_hours[req_idx] += len(placements)
                _undo_block(
                    requirements[req_idx],
                    faculty_ids,
                    placements,
                    schedules,
                    faculty_busy,
                    year,
                    session_log,
                    faculty_section_slots,
                    free_slots_tracker,
                    daily_subject_counts,
                    faculty_daily_loads,
                    subject_edge_counts,
                    instructional_periods,
                )


def _infer_failure_reason(
    requirement: Requirement,
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    daily_subject_counts: dict[str, dict[str, dict[str, int]]],
    faculty_busy: dict[str, set[tuple[str, int]]],
    faculty_availability: dict[str, dict[str, set[int]]],
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]],
    year: str,
    instructional_periods: list[int],
    sessions: list[tuple[int, ...]],
    session_adjacency: dict[int, set[int]],
) -> str:
    faculty_has_any_slot = False
    faculty_has_continuous_slot = False
    section_has_any_slot = False
    faculty_only_blocked_by_section_transition = False
    blocked_only_by_daily_subject_limit = False

    for day in DAYS:
        for start in PERIODS:
            base_section_free = _sections_can_host_block(
                requirement.sections,
                requirement,
                day,
                start,
                1,
                schedules,
                year,
                instructional_periods,
                sessions,
            )
            if not base_section_free:
                continue
            if _exceeds_subject_daily_hard_limit(requirement, day, 1, daily_subject_counts):
                blocked_only_by_daily_subject_limit = True
                continue

            if _choose_faculty_for_slot(
                requirement,
                day,
                start,
                1,
                faculty_busy,
                faculty_availability,
                instructional_periods,
                faculty_section_slots,
                session_adjacency,
                enforce_section_transition_rule=False,
            ):
                if not _choose_faculty_for_slot(
                    requirement,
                    day,
                    start,
                    1,
                    faculty_busy,
                    faculty_availability,
                    instructional_periods,
                    faculty_section_slots,
                    session_adjacency,
                    enforce_section_transition_rule=True,
                ):
                    faculty_only_blocked_by_section_transition = True

            if _slot_is_free(
                requirement.sections,
                requirement,
                day,
                start,
                1,
                schedules,
                faculty_busy,
                faculty_availability,
                year,
                instructional_periods,
                sessions,
                daily_subject_counts,
                faculty_section_slots,
                session_adjacency,
            ):
                faculty_has_any_slot = True
                if len(requirement.sections) > 1:
                    section_has_any_slot = True
                
                if _slot_is_free(
                    requirement.sections,
                    requirement,
                    day,
                    start,
                    min(requirement.hours, max(1, requirement.min_consecutive_hours)),
                    schedules,
                    faculty_busy,
                    faculty_availability,
                    year,
                    instructional_periods,
                    sessions,
                    daily_subject_counts,
                    faculty_section_slots,
                    session_adjacency,
                ):
                    faculty_has_continuous_slot = True
                    section_has_any_slot = True

    if faculty_only_blocked_by_section_transition:
        return "faculty consecutive section constraint"
    if blocked_only_by_daily_subject_limit and not faculty_has_any_slot:
        return "subject daily allocation hard limit"
    if not faculty_has_any_slot:
        return "faculty availability conflict"
    if len(requirement.sections) > 1 and not section_has_any_slot:
        return "shared class constraint"
    if requirement.min_consecutive_hours > 1 and not faculty_has_continuous_slot:
        return "continuous hours constraint"
    return "no free slot"


def _group_issue_records(records: list[dict]) -> list[dict]:
    grouped: dict[tuple[str, str, str, tuple[str, ...], str], dict] = {}
    for record in records:
        sections = tuple(sorted(str(section).strip() for section in record.get("sections", []) if str(section).strip()))
        key = (
            str(record.get("year", "")).strip(),
            str(record.get("subject_id", "")).strip(),
            str(record.get("faculty_id", "")).strip(),
            sections,
            str(record.get("constraint", record.get("detail", ""))).strip(),
        )
        current = grouped.get(key)
        if current is None:
            grouped[key] = {**record, "sections": list(sections), "occurrences": 1}
            continue
        current["occurrences"] = int(current.get("occurrences", 1)) + 1
        if not current.get("detail") and record.get("detail"):
            current["detail"] = record["detail"]

    result: list[dict] = []
    for item in grouped.values():
        occurrences = int(item.pop("occurrences", 1))
        if occurrences > 1:
            base_detail = str(item.get("detail", "")).strip()
            item["detail"] = f"{base_detail} (Grouped {occurrences} similar issue(s))".strip()
        result.append(item)
    result.sort(
        key=lambda row: (
            str(row.get("constraint", "")),
            str(row.get("subject_id", "")),
            ",".join(row.get("sections", [])),
        )
    )
    return result


def _group_sections_by_remaining_capacity(
    requirements: list[Requirement],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
) -> list[dict]:
    section_rows: list[dict] = []
    seen_sections: set[str] = set()
    for requirement in requirements:
        for section in requirement.sections:
            if section in seen_sections:
                continue
            seen_sections.add(section)
            remaining_slots = sum(
                1
                for day in DAYS
                for period in PERIODS
                if schedules[(year, section)][day][period] is None
            )
            remaining_hours = sum(req.hours for req in requirements if section in req.sections)
            if remaining_hours > remaining_slots:
                section_rows.append(
                    {
                        "year": year,
                        "sections": [section],
                        "subject_id": "",
                        "faculty_id": "",
                        "constraint": "section capacity constraint",
                        "detail": f"Section {section} requires {remaining_hours} hour(s) but has only {remaining_slots} free slot(s).",
                    }
                )
    return section_rows


def _merge_overlapping_section_groups(
    groups: list[tuple[str, ...]],
) -> list[tuple[str, ...]]:
    """
    Merge overlapping section groups so the same subject is not counted twice
    for a section when shared-class rows overlap (e.g., A,B and B,C).
    """
    components: list[set[str]] = []
    for group in groups:
        current = {sec for sec in group if sec}
        if not current:
            continue
        merged: list[set[str]] = []
        for comp in components:
            if current.intersection(comp):
                current.update(comp)
            else:
                merged.append(comp)
        merged.append(current)
        components = merged
    return [tuple(sorted(comp)) for comp in components if comp]


def _serialize_section_grids(
    year: str,
    sections: list[str],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    instructional_periods: list[int],
) -> dict[str, dict[str, list[dict | None]]]:
    return {
        section: {day: [schedules[(year, section)][day][period] for period in instructional_periods] for day in DAYS}
        for section in sections
    }


def _collect_subject_daily_hard_limit_violations(
    requirements: list[Requirement],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
) -> list[dict]:
    requirement_by_section_subject = {
        (section, requirement.subject_id): requirement
        for requirement in requirements
        for section in requirement.sections
    }
    violations: list[dict] = []
    for section in sorted({sec for requirement in requirements for sec in requirement.sections}):
        for day in DAYS:
            subject_counts: dict[str, int] = {}
            for period in instructional_periods:
                cell = schedules[(year, section)][day].get(period)
                if not isinstance(cell, dict) or cell.get("isLab"):
                    continue
                subject_id = str(cell.get("subjectId", "")).strip() or str(cell.get("subject", "")).strip()
                if not subject_id:
                    continue
                subject_counts[subject_id] = subject_counts.get(subject_id, 0) + 1
            for subject_id, count in subject_counts.items():
                requirement = requirement_by_section_subject.get((section, subject_id))
                if requirement is None or requirement.daily_limit_exempt or count <= requirement.hard_daily_limit:
                    continue
                violations.append({
                    "year": year,
                    "sections": [section],
                    "subject_id": subject_id,
                    "faculty_id": requirement.faculty_id,
                    "constraint": "subject daily allocation hard limit",
                    "detail": (
                        f"Section {section} assigns subject {subject_id} {count} times on {day}; "
                        f"the hard daily limit is {requirement.hard_daily_limit}."
                    ),
                })
    return violations


def _build_timetable_quality_metrics(
    requirements: list[Requirement],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    year: str,
    instructional_periods: list[int],
    sessions: list[tuple[int, ...]],
) -> tuple[dict[str, int], list[dict]]:
    requirement_by_section_subject = {
        (section, requirement.subject_id): requirement
        for requirement in requirements
        for section in requirement.sections
    }
    warnings: list[dict] = []
    subject_daily_overloads = 0
    first_period_missing = 0
    last_period_missing = 0
    faculty_section_conflicts = 0
    first_period = instructional_periods[0] if instructional_periods else 1
    last_period = instructional_periods[-1] if instructional_periods else 7
    session_adjacency = _build_session_adjacency(sessions)

    for section in sorted({sec for requirement in requirements for sec in requirement.sections}):
        seen_subjects: set[str] = set()
        for day in DAYS:
            subject_counts: dict[str, int] = {}
            for period in instructional_periods:
                cell = schedules[(year, section)][day].get(period)
                if not isinstance(cell, dict) or cell.get("isLab"):
                    continue
                subject_id = str(cell.get("subjectId", "")).strip() or str(cell.get("subject", "")).strip()
                if not subject_id:
                    continue
                seen_subjects.add(subject_id)
                subject_counts[subject_id] = subject_counts.get(subject_id, 0) + 1
            overloaded_subjects: list[str] = []
            for subject_id, count in subject_counts.items():
                requirement = requirement_by_section_subject.get((section, subject_id))
                if requirement and requirement.daily_limit_exempt:
                    continue
                preferred_limit = requirement.preferred_daily_limit if requirement else PREFERRED_SUBJECT_PERIODS_PER_DAY
                if count <= preferred_limit:
                    continue
                overloaded_subjects.append(f"{subject_id} ({count})")
                subject_daily_overloads += 1
            if overloaded_subjects:
                warnings.append({
                    "year": year,
                    "sections": [section],
                    "subject_id": ", ".join(overloaded_subjects),
                    "faculty_id": "",
                    "constraint": "subject daily distribution preference",
                    "detail": (
                        f"Section {section} exceeds the preferred {PREFERRED_SUBJECT_PERIODS_PER_DAY}-period "
                        f"daily subject limit on {day}: {', '.join(overloaded_subjects)}."
                    ),
                })

        for subject_id in sorted(seen_subjects):
            has_first = _subject_has_edge_period_for_section(section, subject_id, first_period, schedules, year)
            has_last = _subject_has_edge_period_for_section(section, subject_id, last_period, schedules, year)
            if not has_first:
                first_period_missing += 1
            if not has_last:
                last_period_missing += 1
            missing_edges = []
            if not has_first:
                missing_edges.append(f"period {first_period}")
            if not has_last:
                missing_edges.append(f"period {last_period}")
            if missing_edges:
                warnings.append({
                    "year": year,
                    "sections": [section],
                    "subject_id": subject_id,
                    "faculty_id": "",
                    "constraint": "edge period distribution preference",
                    "detail": f"Subject {subject_id} for section {section} is missing weekly exposure in {', '.join(missing_edges)}.",
                })

    faculty_assignments: dict[str, dict[tuple[str, int], tuple[str, ...]]] = {}
    for section in sorted({sec for requirement in requirements for sec in requirement.sections}):
        for day in DAYS:
            for period in instructional_periods:
                cell = schedules[(year, section)][day].get(period)
                if not isinstance(cell, dict):
                    continue
                faculty_ids = _split_faculty_tokens(str(cell.get("facultyId") or ""))
                if not faculty_ids:
                    continue
                signature = _section_signature(cell.get("sharedSections") or [section])
                for faculty_id in faculty_ids:
                    faculty_assignments.setdefault(faculty_id, {})[(day, period)] = signature

    for faculty_id, slot_map in faculty_assignments.items():
        for (day, period), signature in slot_map.items():
            for adjacent_period in session_adjacency.get(period, set()):
                if adjacent_period <= period:
                    continue
                adjacent_signature = slot_map.get((day, adjacent_period))
                if adjacent_signature and adjacent_signature != signature:
                    faculty_section_conflicts += 1
                    warnings.append({
                        "year": year,
                        "sections": list(signature),
                        "subject_id": "",
                        "faculty_id": faculty_id,
                        "constraint": "faculty consecutive section constraint",
                        "detail": (
                            f"Faculty {faculty_id} is assigned to different sections in consecutive periods on {day}: "
                            f"{'/'.join(signature)} then {'/'.join(adjacent_signature)}."
                        ),
                    })

    metrics = {
        "facultyConsecutiveSectionConflicts": faculty_section_conflicts,
        "subjectDailyOverloadCount": subject_daily_overloads,
        "missingFirstPeriodCoverageCount": first_period_missing,
        "missingLastPeriodCoverageCount": last_period_missing,
    }
    return metrics, warnings


def _build_faculty_workloads_from_sessions(
    sessions_list: list[dict],
    instructional_periods: list[int],
) -> dict[str, dict[str, list[str | None]]]:
    workloads: dict[str, dict[str, dict[int, str | None]]] = {}
    for session in sessions_list:
        day = session["day"]
        faculty_names = [str(name).strip() for name in session.get("faculty_names", []) if str(name).strip()]
        faculty_ids = [str(fid).strip() for fid in session.get("faculty_ids", []) if str(fid).strip()]
        if not faculty_names:
            fallback_name = str(session.get("faculty_name", "")).strip()
            if fallback_name:
                faculty_names = [part.strip() for part in fallback_name.split(",") if part.strip()]
        if not faculty_ids:
            fallback_id = str(session.get("faculty_id", "")).strip()
            if fallback_id:
                faculty_ids = [part.strip() for part in fallback_id.split(",") if part.strip()]

        faculty_keys = faculty_names or faculty_ids
        if not faculty_keys:
            continue

        for faculty_key in faculty_keys:
            faculty_workload = workloads.setdefault(faculty_key, {d: {p: None for p in instructional_periods} for d in DAYS})
            for period in session["periods"]:
                if period not in instructional_periods: continue
                room_line = ""
                if session.get("isLab"):
                    lab_room = str(session.get("lab_room") or session.get("venue") or "").strip().strip("()")
                    if lab_room:
                        room_line = f"\n({lab_room})"
                else:
                    room_label = _format_non_lab_room_label(session).strip().strip("()")
                    if room_label:
                        room_line = f"\n({room_label})"
                
                yr = str(session.get('year', '')).strip()
                sec_text = _format_sections_ordinal(yr, session.get("sections", []))
                
                cell_text = f"{session['subject_name']}\n{sec_text}"
                if room_line:
                    cell_text += room_line
                faculty_workload[day][period] = cell_text
    
    final_workloads: dict[str, dict[str, list[str | None]]] = {}
    for fid, days_data in workloads.items():
        final_workloads[fid] = {day: [days_data[day][period] for period in instructional_periods] for day in DAYS}
    return final_workloads


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _encode_workbook(file_name: str, workbook: Workbook) -> dict:
    return {
        "fileName": file_name,
        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "contentBase64": base64.b64encode(_workbook_bytes(workbook)).decode("ascii"),
    }


def _style_range(
    worksheet,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
    *,
    alignment: Alignment | None = None,
    border: Border | None = None,
    font: Font | None = None,
) -> None:
    for row in worksheet.iter_rows(
        min_row=start_row,
        max_row=end_row,
        min_col=start_column,
        max_col=end_column,
    ):
        for cell in row:
            if isinstance(cell, MergedCell):
                continue
            if alignment:
                cell.alignment = alignment
            if border:
                cell.border = border
            if font:
                cell.font = font


def _apply_border_sides(
    cell,
    *,
    left: Side | None = None,
    right: Side | None = None,
    top: Side | None = None,
    bottom: Side | None = None,
) -> None:
    current = cell.border
    cell.border = Border(
        left=left or current.left,
        right=right or current.right,
        top=top or current.top,
        bottom=bottom or current.bottom,
    )


def _clear_cell_border(cell) -> None:
    cell.border = Border()


def _apply_merged_range_outline(
    worksheet,
    start_row: int,
    end_row: int,
    start_column: int,
    end_column: int,
    *,
    border: Border,
) -> None:
    for row_idx in range(start_row, end_row + 1):
        for col_idx in range(start_column, end_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            _clear_cell_border(cell)
            _apply_border_sides(
                cell,
                left=border.left if col_idx == start_column else None,
                right=border.right if col_idx == end_column else None,
                top=border.top if row_idx == start_row else None,
                bottom=border.bottom if row_idx == end_row else None,
            )


def _build_subject_legend_for_section(
    section_schedule: dict[str, dict[int, dict | None]],
) -> list[tuple[str, str]]:
    seen: dict[str, str] = {}
    for day in DAYS:
        for period in PERIODS:
            entry = section_schedule[day][period]
            if not entry:
                continue
            subject = str(entry.get("subjectName") or entry.get("subject") or "").strip()
            faculty = str(entry.get("facultyName") or entry.get("faculty") or "").strip()
            if subject and subject not in seen:
                seen[subject] = faculty
    return sorted(seen.items(), key=lambda item: item[0])


def _same_section_entry(left: dict | None, right: dict | None) -> bool:
    if not left or not right:
        return False
    return (
        str(left.get("subjectName") or left.get("subject") or "").strip()
        == str(right.get("subjectName") or right.get("subject") or "").strip()
        and str(left.get("facultyName") or left.get("faculty") or "").strip()
        == str(right.get("facultyName") or right.get("faculty") or "").strip()
        and bool(left.get("isLab")) == bool(right.get("isLab"))
        and ",".join(left.get("sharedSections", []) or [])
        == ",".join(right.get("sharedSections", []) or [])
        and str(left.get("classroom", "")).strip() == str(right.get("classroom", "")).strip()
        and str(left.get("labRoom", "")).strip() == str(right.get("labRoom", "")).strip()
        and str(left.get("fallbackLab", "")).strip() == str(right.get("fallbackLab", "")).strip()
    )


def _format_non_lab_room_label(entry: dict | None) -> str:
    if not entry:
        return ""
    classroom = str(entry.get("classroom", "")).strip()
    fallback_lab = str(entry.get("fallbackLab", "")).strip()
    if fallback_lab and classroom:
        return f"{fallback_lab}/{classroom}"
    if fallback_lab:
        return fallback_lab
    return classroom


def _section_cell_text(entry: dict | None) -> str:
    if not entry:
        return ""
    subject = str(entry.get("subjectName") or entry.get("subject") or "").strip()
    main_line = subject
    if entry.get("isLab"):
        lab_room = str(entry.get("labRoom") or entry.get("venue") or "").strip()
        return f"{main_line}\n({lab_room})" if lab_room else main_line
    room_label = _format_non_lab_room_label(entry)
    if room_label:
        return f"{main_line}\n({room_label})"
    return main_line


def _merge_section_day_row(
    worksheet,
    row_idx: int,
    section_schedule: dict[str, dict[int, dict | None]],
    day: str,
    period_config: list[dict[str, str]],
) -> tuple[bool, bool]:
    display_columns = {}
    col = 2
    break_cols = set()
    lunch_cols = set()
    for entry in period_config:
        p = str(entry.get("period", "")).strip()
        if p.isdigit():
            display_columns[int(p)] = col
        elif "break" in p.lower():
            break_cols.add(col)
        elif "lunch" in p.lower():
            lunch_cols.add(col)
        col += 1
    
    instructional_periods = sorted(display_columns.keys())
    
    idx = 0
    overlaps_break = False
    overlaps_lunch = False
    for c in break_cols:
        if worksheet.cell(row=row_idx, column=c).value:
            overlaps_break = True
    for c in lunch_cols:
        if worksheet.cell(row=row_idx, column=c).value:
            overlaps_lunch = True

    while idx < len(instructional_periods):
        p = instructional_periods[idx]
        entry = section_schedule[day].get(p)
        worksheet.cell(row=row_idx, column=display_columns[p], value=_section_cell_text(entry))
        
        end_idx = idx
        while end_idx + 1 < len(instructional_periods):
            next_p = instructional_periods[end_idx + 1]
            next_entry = section_schedule[day].get(next_p)
            
            is_adjacent = display_columns[next_p] == display_columns[instructional_periods[end_idx]] + 1
            is_lab = isinstance(entry, dict) and entry.get("isLab")
            is_same = _same_section_entry(entry, next_entry)
            
            can_merge = False
            if is_same:
                if is_adjacent:
                    can_merge = True
                elif is_lab:
                    mid_cols = range(display_columns[instructional_periods[end_idx]] + 1, display_columns[next_p])
                    if all(c in break_cols or c in lunch_cols for c in mid_cols):
                        can_merge = True
            
            if can_merge:
                end_idx += 1
            else:
                break
        
        if end_idx > idx:
            start_col = display_columns[p]
            end_col = display_columns[instructional_periods[end_idx]]
            for c in break_cols:
                if start_col <= c <= end_col:
                    overlaps_break = True
            for c in lunch_cols:
                if start_col <= c <= end_col:
                    overlaps_lunch = True
            worksheet.merge_cells(
                start_row=row_idx,
                start_column=start_col,
                end_row=row_idx,
                end_column=end_col,
            )
        idx = end_idx + 1
    
    return overlaps_break, overlaps_lunch


def _normalize_faculty_sheet_name(value: str) -> str:
    invalid = '\\/*?:[]'
    name = "".join("_" if char in invalid else char for char in value).strip()
    return name[:31] or "Faculty"


def _ordinal_year(year: str) -> str:
    y_clean = str(year).strip().lower()
    if "1st" in y_clean or "first" in y_clean or y_clean.startswith("1") or y_clean.startswith("i ") or y_clean == "i":
        return "1st Year"
    if "2nd" in y_clean or "second" in y_clean or y_clean.startswith("2") or y_clean.startswith("ii ") or y_clean == "ii":
        return "2nd Year"
    if "3rd" in y_clean or "third" in y_clean or y_clean.startswith("3") or y_clean.startswith("iii ") or y_clean == "iii":
        return "3rd Year"
    if "4th" in y_clean or "fourth" in y_clean or y_clean.startswith("4") or y_clean.startswith("iv ") or y_clean == "iv":
        return "4th Year"
    return f"{year} Year"


def _format_sections_ordinal(year: str, sections: list[str]) -> str:
    ordinal_yr = _ordinal_year(year)
    sec_lbls = []
    for sec in sections:
        sec_clean = str(sec).strip()
        if "-" in sec_clean:
            sec_clean = sec_clean.split("-")[-1].strip()
        if "[" in sec_clean and "]" in sec_clean:
            import re
            m = re.search(r'\[([^]]+)\]', sec_clean)
            if m:
                sec_clean = m.group(1).strip()
                if "-" in sec_clean:
                    sec_clean = sec_clean.split("-")[-1].strip()
        sec_lbls.append(f"{ordinal_yr} {sec_clean}")
    return ", ".join(sec_lbls)


def _roman_year(year: str) -> str:
    y_clean = str(year).strip().lower()
    if "1st" in y_clean or "first" in y_clean or y_clean.startswith("1") or y_clean.startswith("i "):
        return "I"
    if "2nd" in y_clean or "second" in y_clean or y_clean.startswith("2") or y_clean.startswith("ii "):
        return "II"
    if "3rd" in y_clean or "third" in y_clean or y_clean.startswith("3") or y_clean.startswith("iii "):
        return "III"
    if "4th" in y_clean or "fourth" in y_clean or y_clean.startswith("4") or y_clean.startswith("iv "):
        return "IV"
    if "i" in y_clean:
        if "iii" in y_clean: return "III"
        if "ii" in y_clean: return "II"
        if "iv" in y_clean: return "IV"
        return "I"
    return year


def _load_global_faculty_map() -> dict[str, str]:
    from storage.memory_store import store
    faculty_id_to_name = {}
    fac_map_payload = store.get_scoped_mapping("faculty_id_map", "global")
    if fac_map_payload:
        for row in fac_map_payload.get("rows", []):
            fid = str(row.get("faculty_id") or row.get("faculty id") or row.get("id") or "").strip()
            if fid.endswith(".0"):
                fid = fid[:-2]
            fname = str(row.get("faculty_name") or row.get("faculty name") or row.get("name") or "").strip()
            if fid:
                faculty_id_to_name[fid] = fname
    return faculty_id_to_name


def _build_faculty_schedule_details(
    sessions: list[dict],
    faculty_id_to_name: dict[str, str],
    instructional_periods: list[int],
) -> dict[str, dict[str, list[list[dict] | None]]]:
    name_to_id = {v.strip().lower(): k for k, v in faculty_id_to_name.items()}
    workloads: dict[str, dict[str, list[list[dict] | None]]] = {}
    for session in sessions:
        day = str(session.get("day", "")).strip()
        periods = [int(period) for period in session.get("periods", []) if int(period) in instructional_periods]
        if day not in DAYS or not periods:
            continue

        sections = ",".join(session.get("sections", []))
        detail = {
            "subject": str(session.get("subject_name", "")).strip() or str(session.get("subject_id", "")).strip(),
            "year": str(session.get("year", "")).strip(),
            "section": sections,
            "classroom": str(session.get("classroom", "")).strip(),
            "lab_room": str(session.get("lab_room") or session.get("venue") or "").strip(),
            "fallback_lab": str(session.get("fallbackLab", "")).strip(),
            "is_lab": bool(session.get("isLab")),
        }

        faculty_ids = [str(fid).strip() for fid in session.get("faculty_ids", []) if str(fid).strip()]
        if not faculty_ids:
            fallback_id = str(session.get("faculty_id", "")).strip()
            if fallback_id:
                faculty_ids = [part.strip() for part in fallback_id.split(",") if part.strip()]

        if not faculty_ids:
            faculty_names = [str(name).strip() for name in session.get("faculty_names", []) if str(name).strip()]
            if not faculty_names:
                fallback_name = str(session.get("faculty_name", "")).strip()
                if fallback_name:
                    faculty_names = [part.strip() for part in fallback_name.split(",") if part.strip()]
            for fname in faculty_names:
                fid = _find_faculty_id(fname, name_to_id)
                if fid:
                    faculty_ids.append(fid)
                else:
                    faculty_ids.append(fname)

        clean_ids = []
        for fid in faculty_ids:
            if fid.endswith(".0"):
                fid = fid[:-2]
            clean_ids.append(fid)

        for fid in clean_ids:
            schedule = workloads.setdefault(
                fid,
                {day_name: [None] * len(instructional_periods) for day_name in DAYS},
            )
            for period in periods:
                try:
                    index = instructional_periods.index(period)
                    if schedule[day][index] is None:
                        schedule[day][index] = []
                    existing = schedule[day][index]
                    if detail not in existing:
                        existing.append(detail.copy())
                except ValueError:
                    continue
    return workloads


def _faculty_slot_text(entries: list[dict] | None) -> str:
    if not entries:
        return ""
    rendered: list[str] = []
    for entry in entries:
        lines = [
            str(entry.get("subject", "")).strip(),
            f"{entry.get('year', '')} {entry.get('section', '')}".strip(),
        ]
        if entry.get("is_lab"):
            lab_room = str(entry.get("lab_room", "")).strip()
            if lab_room:
                lines.append(f"({lab_room})")
        else:
            room_label = _format_non_lab_room_label({
                "classroom": entry.get("classroom", ""),
                "fallbackLab": entry.get("fallback_lab", ""),
            })
            if room_label:
                lines.append(f"({room_label})")
        rendered.append("\n".join(line for line in lines if line))
    return "\n\n".join(rendered)


def _same_faculty_slot(left: list[dict] | None, right: list[dict] | None) -> bool:
    if not left or not right or len(left) != len(right):
        return False
    return all(left[idx] == right[idx] for idx in range(len(left)))


def _merge_faculty_run(
    worksheet,
    row_idx: int,
    entries: list[list[dict] | None],
    columns: list[int],
) -> None:
    idx = 0
    while idx < len(entries):
        current = entries[idx]
        worksheet.cell(row=row_idx, column=columns[idx], value=_faculty_slot_text(current))
        end = idx
        while end + 1 < len(entries) and _same_faculty_slot(current, entries[end + 1]):
            end += 1
        if end > idx:
            worksheet.merge_cells(
                start_row=row_idx,
                start_column=columns[idx],
                end_row=row_idx,
                end_column=columns[end],
            )
        idx = end + 1


def _build_faculty_legend(schedule: dict[str, list[list[dict] | None]]) -> list[str]:
    labels: set[str] = set()
    for day in DAYS:
        for entries in schedule.get(day, []):
            if not entries:
                continue
            for entry in entries:
                labels.add(f"{entry['year']} {entry['section']} - {entry['subject']}")
    return sorted(labels)


def _allocate_classrooms_to_schedule(
    year: str,
    sections: list[str],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    session_log: list[dict],
    classrooms: list[str],
    constraint_violations: list[dict],
    instructional_periods: list[int],
    academic_sessions: list[tuple[int, ...]],
    prior_room_busy: dict[str, set[tuple[str, int]]] | None = None,
    lab_room_names: set[str] | None = None,
    room_capacity_map: dict[str, int | None] | None = None,
    section_strength_map: dict[str, int] | None = None,
    fixed_classroom_blocks: dict[tuple[str, str, int], str] | None = None,
    section_home_room_map: dict[str, str] | None = None,
) -> None:
    normalized_rooms = [str(room).strip() for room in classrooms if str(room).strip()]
    lab_room_names = {str(room).strip() for room in (lab_room_names or set()) if str(room).strip()}
    room_capacity_map = {str(room).strip(): capacity for room, capacity in (room_capacity_map or {}).items() if str(room).strip()}
    section_strength_map = {str(section).strip(): int(value) for section, value in (section_strength_map or {}).items() if str(section).strip()}
    fixed_classroom_blocks = {
        (str(section).strip(), str(day).strip(), int(period)): str(classroom).strip()
        for (section, day, period), classroom in (fixed_classroom_blocks or {}).items()
        if str(section).strip() and str(day).strip()
    }
    section_home_room_map = {str(k).strip(): str(v).strip() for k, v in (section_home_room_map or {}).items() if str(k).strip()}
    preferred_rooms = [room for room in normalized_rooms if room not in lab_room_names]
    overflow_rooms = [room for room in normalized_rooms if room in lab_room_names]

    def _section_home_room(section: str) -> str:
        normalized_section = str(section).strip()
        if not normalized_section:
            return ""
        return (
            section_home_room_map.get(f"{year}|{normalized_section}")
            or section_home_room_map.get(normalized_section)
            or section_home_room_map.get(normalized_section.upper())
            or ""
        )

    def _section_strength(section: str) -> int:
        normalized_section = str(section).strip()
        if not normalized_section:
            return 0
        return int(
            section_strength_map.get(f"{year}|{normalized_section}")
            or section_strength_map.get(normalized_section)
            or section_strength_map.get(normalized_section.upper())
            or 0
        )

    def _reserve_room(room_name: str, day: str, periods: list[int]) -> None:
        if not room_name:
            return
        room_usage.setdefault(room_name, set()).update((day, period) for period in periods)

    room_usage: dict[str, set[tuple[str, int]]] = {}
    for room in normalized_rooms:
        prior = (prior_room_busy or {}).get(room, set())
        room_usage[room] = prior.copy()

    # First, handle labs - they already have their room (venue) locked in
    for session in session_log:
        if not bool(session.get("isLab")):
            continue
        room_name = _normalize_id_token(session.get("lab_room") or session.get("venue") or "")
        day = str(session.get("day", "")).strip()
        periods = [int(period) for period in session.get("periods", []) if int(period) in instructional_periods]
        if day in DAYS and periods and room_name:
            # Check if any period is already reserved
            for p in periods:
                if (day, p) in room_usage.setdefault(room_name, set()):
                    constraint_violations.append({
                        "year": year,
                        "sections": list(session.get("sections", [])),
                        "subject_id": str(session.get("subject_id", "")),
                        "faculty_id": "",
                        "constraint": "lab room conflict",
                        "detail": f"Lab room {room_name} is already occupied.",
                        "room": room_name,
                    })
            _reserve_room(room_name, day, periods)
            # Also ensure schedules grid is updated
            for section in session.get("sections", []):
                for period in periods:
                    cell = schedules[(year, section)][day].get(period)
                    if isinstance(cell, dict):
                        cell["labRoom"] = room_name

    for (section, day, period), classroom in fixed_classroom_blocks.items():
        cell = schedules.get((year, section), {}).get(day, {}).get(period)
        if not isinstance(cell, dict) or cell.get("isLab"):
            continue
        cell["fixedClassroomBlock"] = True
        cell["classroom"] = classroom
        if not classroom:
            continue
        if any((day, p) in room_usage.setdefault(classroom, set()) for p in [period]):
            constraint_violations.append({
                "year": year,
                "sections": [section],
                "subject_id": str(cell.get("subjectId", "")).strip(),
                "faculty_id": str(cell.get("facultyId", "")).strip(),
                "constraint": "fixed classroom block conflict",
                "detail": (
                    f"Fixed classroom {classroom} for section {section} on {day} period {period} "
                    f"conflicts with an already occupied room slot."
                ),
            })
        _reserve_room(classroom, day, [period])

    # Dynamic session-based classroom allocation helper
    def _allocate_session_rooms(day: str, session_idx: int, academic_session: tuple[int, ...]) -> bool:
        session_periods = list(academic_session)
        
        # Check active sections
        active_sections: list[str] = []
        occupied_periods_by_sec: dict[str, list[int]] = {}
        has_shared_by_sec: dict[str, bool] = {}
        fixed_rooms_by_sec_p: dict[tuple[str, int], str] = {}
        
        for section in sections:
            sec_occupied = []
            has_shared = False
            for period in session_periods:
                cell = schedules[(year, section)][day].get(period)
                if isinstance(cell, dict):
                    sec_occupied.append(period)
                    if cell.get("isLab"):
                        fixed_rooms_by_sec_p[(section, period)] = _normalize_id_token(cell.get("labRoom") or cell.get("venue") or "")
                    elif cell.get("fixedClassroomBlock"):
                        fixed_rooms_by_sec_p[(section, period)] = _normalize_id_token(cell.get("classroom") or "")
                    
                    shared_secs = [str(s).strip() for s in (cell.get("sharedSections") or []) if str(s).strip() in sections]
                    if len(shared_secs) > 1 or (len(shared_secs) == 1 and shared_secs[0] != section):
                        has_shared = True
                        
            if sec_occupied:
                active_sections.append(section)
                occupied_periods_by_sec[section] = sec_occupied
                has_shared_by_sec[section] = has_shared

        if not active_sections:
            return True

        import itertools
        candidate_rooms = preferred_rooms + overflow_rooms
        sequences_by_sec: dict[str, list[list[str]]] = {}
        for section in active_sections:
            occ = occupied_periods_by_sec[section]
            has_sh = has_shared_by_sec[section]
            
            home_room = _section_home_room(section)
            sorted_rooms = []
            if home_room and home_room in candidate_rooms:
                sorted_rooms.append(home_room)
            for r in candidate_rooms:
                if r not in sorted_rooms:
                    sorted_rooms.append(r)
            if "" not in sorted_rooms:
                sorted_rooms.append("")
                    
            allowed_rooms = []
            for period in occ:
                if (section, period) in fixed_rooms_by_sec_p:
                    allowed_rooms.append([fixed_rooms_by_sec_p[(section, period)]])
                else:
                    allowed_rooms.append(sorted_rooms)
                    
            seqs = []
            for seq in itertools.product(*allowed_rooms):
                seq_list = list(seq)
                seqs.append(seq_list)
            
            # Sort sequences to prefer physical classrooms first (fewer empty slots)
            # and then prefer fewer room transitions.
            def seq_sort_key(s):
                unassigned = s.count("")
                non_empty = [r for r in s if r]
                transitions = 0
                for idx in range(len(non_empty) - 1):
                    if non_empty[idx] != non_empty[idx + 1]:
                        transitions += 1
                return (unassigned, transitions)
            
            seqs.sort(key=seq_sort_key)
            sequences_by_sec[section] = seqs

        assignment: dict[str, list[str]] = {}
        room_at_period_secs: dict[tuple[str, int], list[str]] = {}
        
        def is_valid_partial(sec: str, seq: list[str]) -> bool:
            occ = occupied_periods_by_sec[sec]
            for idx, period in enumerate(occ):
                r = seq[idx]
                if not r:
                    continue
                cell = schedules[(year, sec)][day].get(period)
                is_lab = cell.get("isLab") if cell else False
                
                # Enforce that shared classes must use the SAME room
                if cell and not is_lab:
                    shared_secs = [str(s).strip() for s in (cell.get("sharedSections") or []) if str(s).strip() in active_sections]
                    for other in shared_secs:
                        if other == sec:
                            continue
                        if other in assignment:
                            other_occ = occupied_periods_by_sec[other]
                            if period in other_occ:
                                other_room = assignment[other][other_occ.index(period)]
                                if other_room != r:
                                    return False
                
                other_secs = room_at_period_secs.get((r, period), [])
                if other_secs:
                    if is_lab:
                        continue
                        
                    for other in other_secs:
                        cell_other = schedules[(year, other)][day].get(period)
                        if not cell or not cell_other:
                            return False
                        if cell_other.get("isLab"):
                            return False
                        if cell.get("subjectId") != cell_other.get("subjectId") or cell.get("facultyId") != cell_other.get("facultyId"):
                            return False
                        
                        shared_with_other = str(other).strip() in [str(s).strip() for s in (cell.get("sharedSections") or [])]
                        shared_with_sec = str(sec).strip() in [str(s).strip() for s in (cell_other.get("sharedSections") or [])]
                        if not shared_with_other or not shared_with_sec:
                            return False
                    
                    strength = _section_strength(sec)
                    combined_strength = strength + sum(_section_strength(o) for o in other_secs)
                    capacity = room_capacity_map.get(r)
                    if capacity is not None and capacity < combined_strength:
                        return False
                else:
                    if is_lab:
                        continue
                    is_fixed = (sec, period) in fixed_rooms_by_sec_p
                    if not is_fixed:
                        if (day, period) in room_usage.get(r, set()):
                            return False
                    
                    strength = _section_strength(sec)
                    capacity = room_capacity_map.get(r)
                    if capacity is not None and capacity < strength:
                        return False
            return True

        def register_sequence(sec: str, seq: list[str]):
            occ = occupied_periods_by_sec[sec]
            for idx, period in enumerate(occ):
                r = seq[idx]
                if r:
                    room_at_period_secs.setdefault((r, period), []).append(sec)

        def unregister_sequence(sec: str, seq: list[str]):
            occ = occupied_periods_by_sec[sec]
            for idx, period in enumerate(occ):
                r = seq[idx]
                if r:
                    if (r, period) in room_at_period_secs:
                        if sec in room_at_period_secs[(r, period)]:
                            room_at_period_secs[(r, period)].remove(sec)

        capacity_error_flag = [False]

        def backtrack(sec_idx: int) -> bool:
            if sec_idx == len(active_sections):
                return True
                
            sec = active_sections[sec_idx]
            seqs = sequences_by_sec[sec]
            
            for seq in seqs:
                if is_valid_partial(sec, seq):
                    register_sequence(sec, seq)
                    assignment[sec] = seq
                    if backtrack(sec_idx + 1):
                        return True
                    unregister_sequence(sec, seq)
                    assignment.pop(sec, None)
                else:
                    occ = occupied_periods_by_sec[sec]
                    for idx, period in enumerate(occ):
                        r = seq[idx]
                        if not r:
                            continue
                        capacity = room_capacity_map.get(r)
                        other_secs = room_at_period_secs.get((r, period), [])
                        if other_secs:
                            cell = schedules[(year, sec)][day].get(period)
                            all_sharing = True
                            for other in other_secs:
                                cell_other = schedules[(year, other)][day].get(period)
                                if not cell or not cell_other or cell.get("subjectId") != cell_other.get("subjectId") or cell.get("facultyId") != cell_other.get("facultyId"):
                                    all_sharing = False
                                    break
                            if all_sharing:
                                combined_strength = _section_strength(sec) + sum(_section_strength(o) for o in other_secs)
                                if capacity is not None and capacity < combined_strength:
                                    capacity_error_flag[0] = True
            return False

        if backtrack(0):
            for sec in active_sections:
                occ = occupied_periods_by_sec[sec]
                seq = assignment[sec]
                
                # Check if this section has any lab period in this session
                session_lab_room = None
                for period in occ:
                    cell = schedules[(year, sec)][day].get(period)
                    if isinstance(cell, dict) and cell.get("isLab"):
                        session_lab_room = _normalize_id_token(cell.get("labRoom") or cell.get("venue") or "")
                        if session_lab_room:
                            break
                            
                for idx, period in enumerate(occ):
                    r = seq[idx]
                    cell = schedules[(year, sec)][day].get(period)
                    if isinstance(cell, dict):
                        if cell.get("isLab"):
                            cell["labRoom"] = r
                        else:
                            if session_lab_room:
                                cell["fallbackLab"] = session_lab_room
                                if r != session_lab_room:
                                    cell["classroom"] = r
                                else:
                                    cell["classroom"] = ""
                            else:
                                cell["classroom"] = r
                                cell["fallbackLab"] = ""
                    is_fixed = (sec, period) in fixed_rooms_by_sec_p
                    if not is_fixed and r:
                        _reserve_room(r, day, [period])
                        
            for log_entry in session_log:
                if log_entry.get("day") != day or log_entry.get("isLab"):
                    continue
                log_periods = [int(p) for p in log_entry.get("periods", []) if int(p) in session_periods]
                if log_periods:
                    log_sec = None
                    for s in log_entry.get("sections", []):
                        if str(s).strip() in assignment:
                            log_sec = str(s).strip()
                            break
                    if log_sec:
                        occ = occupied_periods_by_sec[log_sec]
                        rooms_for_periods = []
                        for p in log_periods:
                            if p in occ:
                                idx = occ.index(p)
                                rooms_for_periods.append(assignment[log_sec][idx])
                        if rooms_for_periods:
                            log_entry["classroom"] = "/".join(dict.fromkeys(r for r in rooms_for_periods if r))

            # Report constraint violations for periods that ended up with no physical classroom assigned
            for sec in active_sections:
                occ = occupied_periods_by_sec[sec]
                seq = assignment[sec]
                unassigned_periods = []
                for idx, period in enumerate(occ):
                    r = seq[idx]
                    cell = schedules[(year, sec)][day].get(period)
                    is_lab = cell.get("isLab") if cell else False
                    is_fixed = (sec, period) in fixed_rooms_by_sec_p
                    if not is_lab and not is_fixed and not r:
                        unassigned_periods.append(period)
                if unassigned_periods:
                    periods_str = ", ".join(str(p) for p in unassigned_periods)
                    detail = f"No classroom can accommodate Section {sec} on {day} for period(s) {periods_str}."
                    constraint_violations.append({
                        "year": year,
                        "sections": [sec],
                        "subject_id": "",
                        "faculty_id": "",
                        "constraint": "classroom allocation constraint",
                        "detail": detail,
                    })

            return True
        else:
            for sec in active_sections:
                if capacity_error_flag[0] and has_shared_by_sec[sec]:
                    detail = "Shared class room capacity is insufficient."
                else:
                    detail = f"No classroom can accommodate Section {sec}."
                constraint_violations.append({
                    "year": year,
                    "sections": [sec],
                    "subject_id": "",
                    "faculty_id": "",
                    "constraint": "classroom allocation constraint",
                    "detail": detail,
                })
            return False

    for day in DAYS:
        for session_idx, academic_session in enumerate(academic_sessions):
            _allocate_session_rooms(day, session_idx, academic_session)

    if not normalized_rooms:
        return


def _apply_formatted_sheet_layout(
    worksheet,
    *,
    legend_row_count: int,
    legend_start_row: int,
    day_start_row: int,
    top_rows: int,
    max_col: int = 10,
) -> None:
    # Use standard widths for main columns, then adjust
    for column_idx in range(1, max_col + 1):
        worksheet.column_dimensions[chr(64 + column_idx)].width = 16
    worksheet.column_dimensions["A"].width = 7

    for row_idx in range(1, worksheet.max_row + 1):
        if day_start_row <= row_idx < day_start_row + len(DAYS):
            worksheet.row_dimensions[row_idx].height = 42
        elif row_idx <= top_rows:
            worksheet.row_dimensions[row_idx].height = 22
        elif row_idx > worksheet.max_row - legend_row_count - 2:
            worksheet.row_dimensions[row_idx].height = 22
        else:
            worksheet.row_dimensions[row_idx].height = 20

    _style_range(
        worksheet,
        1,
        worksheet.max_row,
        1,
        max_col,
        alignment=CENTER_ALIGNMENT,
        border=None,
    )
    
    grid_start_row = day_start_row - 2
    grid_end_row = day_start_row + len(DAYS) - 1
    thin_side = Side(style="thin", color="000000")
    
    for row_idx in range(grid_start_row, grid_end_row + 1):
        for col_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )
    
    # Add Bold Fonts to Headers, Day Column, and Signatures
    _style_range(worksheet, 1, top_rows, 1, max_col, font=BOLD_FONT)
    if day_start_row > 2:
        _style_range(worksheet, day_start_row - 2, day_start_row - 1, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, day_start_row, day_start_row + len(DAYS) - 1, 1, 1, font=BOLD_FONT)
    _style_range(worksheet, worksheet.max_row, worksheet.max_row, 1, max_col, font=BOLD_FONT)
    
    for merged_range in list(worksheet.merged_cells.ranges):
        anchor = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
        anchor.alignment = CENTER_ALIGNMENT
        if grid_start_row <= merged_range.min_row <= grid_end_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = worksheet.cell(row=r, column=c)
                    _clear_cell_border(cell)
                    
                    left_side = thin_side if c == merged_range.min_col else None
                    right_side = thin_side if c == merged_range.max_col else None
                    
                    top_side = thin_side if r == merged_range.min_row else None
                    bottom_side = thin_side if r == merged_range.max_row else None
                    
                    _apply_border_sides(
                        cell,
                        left=left_side,
                        right=right_side,
                        top=top_side,
                        bottom=bottom_side,
                    )
        else:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    worksheet.cell(row=r, column=c).border = Border()

    if legend_row_count > 0:
        for row_idx in range(legend_start_row, legend_start_row + legend_row_count):
            worksheet.cell(row=row_idx, column=1).alignment = LEFT_ALIGNMENT
            worksheet.cell(row=row_idx, column=6).alignment = LEFT_ALIGNMENT


def _build_section_timetables_workbook_from_schedule_map(
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    timetable_metadata: dict[str, Any] | None = None,
    period_config: list[dict[str, str]] | None = None,
) -> Workbook:
    if period_config is None:
        period_config = DEFAULT_PERIOD_CONFIG
    workbook = Workbook()
    metadata = _resolve_timetable_metadata(timetable_metadata)
    workbook.remove(workbook.active)
    
    header1 = [str(row.get("period", "")).strip() for row in period_config]
    header2 = [str(row.get("time", "")).strip() for row in period_config]
    max_col = 1 + len(period_config)

    for year, section in sorted(schedules.keys(), key=lambda item: (item[0], item[1])):
        worksheet = workbook.create_sheet(title=f"{year}_{section}"[:31])
        worksheet.append([ACADEMIC_METADATA["college"]] + [""] * (max_col - 1))
        worksheet.append(["(AUTONOMOUS)"] + [""] * (max_col - 1))
        worksheet.append([ACADEMIC_METADATA["department"]] + [""] * (max_col - 1))
        worksheet.append([f"ACADEMIC YEAR : {metadata['academicYear']} {metadata['semester']}"] + [""] * (max_col - 1))
        worksheet.append([_class_title(year, section, metadata["semester"])] + [""] * (max_col - 1))
        worksheet.append([f"With effect from : {metadata['withEffectFromDisplay']}"] + [""] * (max_col - 1))
        worksheet.append(["DAY"] + header1)
        worksheet.append([""] + header2)

        day_start_row = 9
        section_schedule = schedules[(year, section)]
        break_overlap_rows: list[int] = []
        lunch_overlap_rows: list[int] = []
        
        break_cols = []
        lunch_cols = []
        for col_idx, row in enumerate(period_config, start=2):
            p = str(row.get("period", "")).strip().lower()
            if "break" in p:
                break_cols.append(col_idx)
            elif "lunch" in p:
                lunch_cols.append(col_idx)

        for offset, day in enumerate(DAYS):
            row_idx = day_start_row + offset
            worksheet.cell(row=row_idx, column=1, value=DAY_SHORT_LABELS[day])
            overlaps_break, overlaps_lunch = _merge_section_day_row(worksheet, row_idx, section_schedule, day, period_config)
            if overlaps_break:
                break_overlap_rows.append(row_idx)
            if overlaps_lunch:
                lunch_overlap_rows.append(row_idx)

        worksheet.append([""] * max_col)
        legend_separator_row = worksheet.max_row
        legend = _build_subject_legend_for_section(section_schedule)
        for idx in range(0, len(legend), 2):
            left = legend[idx]
            right = legend[idx + 1] if idx + 1 < len(legend) else None
            worksheet.append(
                [f"{left[0]} : {left[1]}" if left else ""]
                + [""] * (max_col // 2 - 1)
                + [f"{right[0]} : {right[1]}" if right else ""]
                + [""] * (max_col // 2)
            )
        for _ in range(3):
            worksheet.append([""] * max_col)
        signature_separator_row = worksheet.max_row
        sig_row = signature_separator_row + 1
        worksheet.cell(row=sig_row, column=1, value="HEAD OF THE DEPARTMENT")
        worksheet.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=max_col // 2)
        worksheet.cell(row=sig_row, column=max_col // 2 + 1, value="PRINCIPAL")
        worksheet.merge_cells(start_row=sig_row, start_column=max_col // 2 + 1, end_row=sig_row, end_column=max_col)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
        worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=max_col)
        worksheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=max_col)
        worksheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=max_col)
        worksheet.merge_cells(start_row=7, start_column=1, end_row=8, end_column=1)

        def merge_vertical_marker(column: int, label: str, blocked_rows: list[int]) -> None:
            all_rows = [day_start_row + i for i in range(len(DAYS))]
            segments: list[tuple[int, int]] = []
            segment_start = None
            for r in all_rows:
                if r in blocked_rows:
                    if segment_start is not None:
                        segments.append((segment_start, r - 1))
                        segment_start = None
                else:
                    if segment_start is None:
                        segment_start = r
            if segment_start is not None:
                segments.append((segment_start, all_rows[-1]))

            for start, end in segments:
                worksheet.merge_cells(start_row=start, start_column=column, end_row=end, end_column=column)
                worksheet.cell(row=start, column=column, value=label)
                worksheet.cell(row=start, column=column).font = BOLD_FONT

        for bc in break_cols:
            merge_vertical_marker(bc, "BREAK", break_overlap_rows)
        for lc in lunch_cols:
            merge_vertical_marker(lc, "LUNCH", lunch_overlap_rows)

        worksheet.merge_cells(start_row=legend_separator_row, start_column=1, end_row=legend_separator_row, end_column=max_col)
        legend_start_row = day_start_row + len(DAYS) + 1
        for legend_row in range(legend_start_row, legend_start_row + (len(legend) + 1) // 2):
            worksheet.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=max_col // 2)
            worksheet.merge_cells(start_row=legend_row, start_column=max_col // 2 + 1, end_row=legend_row, end_column=max_col)
        
        _apply_formatted_sheet_layout(
            worksheet,
            legend_row_count=(len(legend) + 1) // 2,
            legend_start_row=legend_start_row,
            day_start_row=day_start_row,
            top_rows=6,
            max_col=max_col,
        )
    return workbook


def _build_section_timetables_workbook(
    year: str,
    sections: list[str],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    timetable_metadata: dict[str, Any] | None = None,
    period_config: list[dict[str, str]] | None = None,
) -> Workbook:
    filtered = {
        (year, section): schedules[(year, section)]
        for section in sections
        if (year, section) in schedules
    }
    return _build_section_timetables_workbook_from_schedule_map(filtered, timetable_metadata, period_config)


def _find_faculty_id(faculty_name: str, name_to_id: dict[str, str]) -> str | None:
    name_clean = faculty_name.strip()
    if name_clean in name_to_id:
        return name_to_id[name_clean]
    name_lower = name_clean.lower()
    for fname, fid in name_to_id.items():
        if fname.lower() == name_lower:
            return fid
    for fname, fid in name_to_id.items():
        if fname.lower() in name_lower or name_lower in fname.lower():
            return fid
    return None


def _write_single_workload_sheet(
    worksheet,
    start_row: int,
    faculty_id: str,
    faculty_name: str,
    schedule: dict,
    header1: list[str],
    header2: list[str],
    max_col: int,
    display_columns: dict,
    break_cols: list[int],
    lunch_cols: list[int],
    metadata: dict,
) -> None:
    # 1. College Header
    worksheet.cell(row=start_row, column=1, value=ACADEMIC_METADATA["college"])
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    
    # 2. Autonomous line
    worksheet.cell(row=start_row+1, column=1, value="(AUTONOMOUS)")
    worksheet.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=max_col)
    
    # 3. Department
    worksheet.cell(row=start_row+2, column=1, value=ACADEMIC_METADATA["department"])
    worksheet.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+2, end_column=max_col)
    
    # 4. Academic year metadata
    worksheet.cell(row=start_row+3, column=1, value=_faculty_workload_academic_line(metadata))
    worksheet.merge_cells(start_row=start_row+3, start_column=1, end_row=start_row+3, end_column=max_col)
    
    # 5. Faculty Workload Title Row
    worksheet.cell(row=start_row+4, column=1, value="Faculty Workload")
    worksheet.merge_cells(start_row=start_row+4, start_column=1, end_row=start_row+4, end_column=max_col)
    
    # 6. Faculty ID and Name Placement
    worksheet.cell(row=start_row+5, column=1, value=f"Faculty ID : {faculty_id}")
    worksheet.cell(row=start_row+5, column=max_col // 2 + 1, value=f"Faculty Name : {faculty_name}")
    worksheet.merge_cells(start_row=start_row+5, start_column=1, end_row=start_row+5, end_column=max_col // 2)
    worksheet.merge_cells(start_row=start_row+5, start_column=max_col // 2 + 1, end_row=start_row+5, end_column=max_col)
    
    # 7. Headers
    worksheet.cell(row=start_row+6, column=1, value="DAY")
    for c_idx, h in enumerate(header1, start=2):
        worksheet.cell(row=start_row+6, column=c_idx, value=h)
    for c_idx, h in enumerate(header2, start=2):
        worksheet.cell(row=start_row+7, column=c_idx, value=h)
    worksheet.merge_cells(start_row=start_row+6, start_column=1, end_row=start_row+7, end_column=1)
    
    # 8. Grid Data
    day_start_row = start_row + 8
    break_overlap_rows = []
    lunch_overlap_rows = []
    
    for offset, day in enumerate(DAYS):
        row_idx = day_start_row + offset
        worksheet.cell(row=row_idx, column=1, value=DAY_SHORT_LABELS[day])
        for bc in break_cols:
            worksheet.cell(row=row_idx, column=bc, value="BREAK" if offset == 0 else "")
        for lc in lunch_cols:
            worksheet.cell(row=row_idx, column=lc, value="LUNCH" if offset == 0 else "")
        
        day_data = schedule.get(day, [])
        instructional_periods = sorted(display_columns.keys())
        for p_idx, entries in enumerate(day_data):
            if p_idx < len(instructional_periods) and entries:
                p_num = instructional_periods[p_idx]
                col_idx = display_columns[p_num]
                
                # Group sections by (subject, room) to combine them into one cell
                groups = {}
                for e in entries:
                    if not e: continue
                    subj = str(e.get('subject') or '').strip()
                    room = str(e.get("classroom") or e.get("lab_room") or e.get("fallback_lab") or "").strip()
                    yr = str(e.get('year', '')).strip()
                    sec = str(e.get('section', '')).strip()
                    
                    key = (subj, room, yr)
                    groups.setdefault(key, []).append(sec)
                
                parts = []
                for (subj, room, yr), secs in groups.items():
                    sec_text = _format_sections_ordinal(yr, secs)
                    cell_text = f"{subj}\n{sec_text}"
                    room_clean = room.strip().strip("()")
                    if room_clean:
                        cell_text += f"\n({room_clean})"
                    parts.append(cell_text)
                
                text = "\n\n".join(parts)
                worksheet.cell(row=row_idx, column=col_idx, value=text)

        # Run horizontal merge loop
        idx = 0
        while idx < len(instructional_periods):
            p = instructional_periods[idx]
            col_start = display_columns[p]
            text = worksheet.cell(row=row_idx, column=col_start).value
            
            end_idx = idx
            while end_idx + 1 < len(instructional_periods):
                next_p = instructional_periods[end_idx + 1]
                col_next = display_columns[next_p]
                next_text = worksheet.cell(row=row_idx, column=col_next).value
                
                is_adjacent = col_next == display_columns[instructional_periods[end_idx]] + 1
                
                curr_entries = day_data[idx] if idx < len(day_data) else None
                is_curr_lab = any(e.get("is_lab") for e in curr_entries if e) if curr_entries else False
                
                is_same = (text == next_text) and (text is not None)
                
                can_merge = False
                if is_same:
                    if is_adjacent:
                        can_merge = True
                    elif is_curr_lab:
                        mid_cols = range(display_columns[instructional_periods[end_idx]] + 1, col_next)
                        if all(c in break_cols or c in lunch_cols for c in mid_cols):
                            can_merge = True
                            
                if can_merge:
                    end_idx += 1
                else:
                    break
            
            if end_idx > idx:
                col_end = display_columns[instructional_periods[end_idx]]
                for c in break_cols:
                    if col_start <= c <= col_end:
                        break_overlap_rows.append(row_idx)
                for c in lunch_cols:
                    if col_start <= c <= col_end:
                        lunch_overlap_rows.append(row_idx)
                worksheet.merge_cells(start_row=row_idx, start_column=col_start, end_row=row_idx, end_column=col_end)
            idx = end_idx + 1
    
    # Merge breaks/lunches vertically in segments
    def merge_vertical_marker(column: int, label: str, blocked_rows: list[int]) -> None:
        all_rows = [day_start_row + i for i in range(len(DAYS))]
        segments: list[tuple[int, int]] = []
        segment_start = None
        for r in all_rows:
            if r in blocked_rows:
                if segment_start is not None:
                    segments.append((segment_start, r - 1))
                    segment_start = None
            else:
                if segment_start is None: segment_start = r
        if segment_start is not None: segments.append((segment_start, all_rows[-1]))

        for start, end in segments:
            worksheet.merge_cells(start_row=start, start_column=column, end_row=end, end_column=column)
            worksheet.cell(row=start, column=column, value=label)
            worksheet.cell(row=start, column=column).font = BOLD_FONT

    for bc in break_cols:
        merge_vertical_marker(bc, "BREAK", break_overlap_rows)
    for lc in lunch_cols:
        merge_vertical_marker(lc, "LUNCH", lunch_overlap_rows)
    
    # 9. Signatures
    sig_row = start_row + 15
    worksheet.cell(row=sig_row, column=1, value="HEAD OF THE DEPARTMENT")
    worksheet.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=max_col // 2)
    worksheet.cell(row=sig_row, column=max_col // 2 + 1, value="PRINCIPAL")
    worksheet.merge_cells(start_row=sig_row, start_column=max_col // 2 + 1, end_row=sig_row, end_column=max_col)
    
    grid_s = start_row + 6
    grid_e = start_row + 13
    thin_side = Side(style="thin", color="000000")
    
    for column_idx in range(1, max_col + 1):
        worksheet.column_dimensions[chr(64 + column_idx)].width = 16
    worksheet.column_dimensions["A"].width = 7
    
    for r_idx in range(start_row, sig_row + 1):
        if day_start_row <= r_idx < day_start_row + len(DAYS):
            worksheet.row_dimensions[r_idx].height = 42
        elif r_idx <= start_row + 5:
            worksheet.row_dimensions[r_idx].height = 22
        elif r_idx == sig_row:
            worksheet.row_dimensions[r_idx].height = 22
        else:
            worksheet.row_dimensions[r_idx].height = 20
    
    _style_range(worksheet, start_row, sig_row, 1, max_col, alignment=CENTER_ALIGNMENT)
    _style_range(worksheet, start_row, start_row + 5, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, start_row + 6, start_row + 7, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, day_start_row, day_start_row + len(DAYS) - 1, 1, 1, font=BOLD_FONT)
    _style_range(worksheet, sig_row, sig_row, 1, max_col, font=BOLD_FONT)
    
    for r_idx in range(grid_s, grid_e + 1):
        for c_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )
    
    for merged_range in list(worksheet.merged_cells.ranges):
        if grid_s <= merged_range.min_row <= grid_e and merged_range.min_row >= start_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = worksheet.cell(row=r, column=c)
                    _clear_cell_border(cell)
                    left_s = thin_side if c == merged_range.min_col else None
                    right_s = thin_side if c == merged_range.max_col else None
                    top_s = thin_side if r == merged_range.min_row else None
                    bottom_s = thin_side if r == merged_range.max_row else None
                    _apply_border_sides(
                        cell,
                        left=left_s,
                        right=right_s,
                        top=top_s,
                        bottom=bottom_s,
                    )
        elif merged_range.min_row >= start_row and merged_range.max_row <= sig_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    worksheet.cell(row=r, column=c).border = Border()


def _write_printable_workload_block(
    worksheet,
    start_row: int,
    faculty_id: str,
    faculty_name: str,
    schedule: dict,
    header1: list[str],
    header2: list[str],
    max_col: int,
    display_columns: dict,
    break_cols: list[int],
    lunch_cols: list[int],
    metadata: dict,
) -> None:
    # 1. College Header
    worksheet.cell(row=start_row, column=1, value=ACADEMIC_METADATA["college"])
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    
    # 2. Autonomous line
    worksheet.cell(row=start_row+1, column=1, value="(AUTONOMOUS)")
    worksheet.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=max_col)
    
    # 3. Department
    worksheet.cell(row=start_row+2, column=1, value=ACADEMIC_METADATA["department"])
    worksheet.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+2, end_column=max_col)
    
    # 4. Academic year metadata
    worksheet.cell(row=start_row+3, column=1, value=_faculty_workload_academic_line(metadata))
    worksheet.merge_cells(start_row=start_row+3, start_column=1, end_row=start_row+3, end_column=max_col)
    
    # 5. Faculty Workload Title Row
    worksheet.cell(row=start_row+4, column=1, value="Faculty Workload")
    worksheet.merge_cells(start_row=start_row+4, start_column=1, end_row=start_row+4, end_column=max_col)
    
    # 6. Faculty ID and Name Placement
    worksheet.cell(row=start_row+5, column=1, value=f"Faculty ID : {faculty_id}")
    worksheet.cell(row=start_row+5, column=max_col // 2 + 1, value=f"Faculty Name : {faculty_name}")
    worksheet.merge_cells(start_row=start_row+5, start_column=1, end_row=start_row+5, end_column=max_col // 2)
    worksheet.merge_cells(start_row=start_row+5, start_column=max_col // 2 + 1, end_row=start_row+5, end_column=max_col)
    
    # 7. Headers
    worksheet.cell(row=start_row+6, column=1, value="DAY")
    for c_idx, h in enumerate(header1, start=2):
        worksheet.cell(row=start_row+6, column=c_idx, value=h)
    for c_idx, h in enumerate(header2, start=2):
        worksheet.cell(row=start_row+7, column=c_idx, value=h)
    worksheet.merge_cells(start_row=start_row+6, start_column=1, end_row=start_row+7, end_column=1)
    
    # 8. Grid Data
    day_start_row = start_row + 8
    break_overlap_rows = []
    lunch_overlap_rows = []
    
    for offset, day in enumerate(DAYS):
        row_idx = day_start_row + offset
        worksheet.cell(row=row_idx, column=1, value=DAY_SHORT_LABELS[day])
        for bc in break_cols:
            worksheet.cell(row=row_idx, column=bc, value="BREAK" if offset == 0 else "")
        for lc in lunch_cols:
            worksheet.cell(row=row_idx, column=lc, value="LUNCH" if offset == 0 else "")
        
        day_data = schedule.get(day, [])
        instructional_periods = sorted(display_columns.keys())
        for p_idx, entries in enumerate(day_data):
            if p_idx < len(instructional_periods) and entries:
                p_num = instructional_periods[p_idx]
                col_idx = display_columns[p_num]
                
                # Group sections by (subject, room) to combine them into one cell
                groups = {}
                for e in entries:
                    if not e: continue
                    subj = str(e.get('subject') or '').strip()
                    room = str(e.get("classroom") or e.get("lab_room") or e.get("fallback_lab") or "").strip()
                    yr = str(e.get('year', '')).strip()
                    sec = str(e.get('section', '')).strip()
                    
                    key = (subj, room, yr)
                    groups.setdefault(key, []).append(sec)
                
                parts = []
                for (subj, room, yr), secs in groups.items():
                    sec_text = _format_sections_ordinal(yr, secs)
                    cell_text = f"{subj}\n{sec_text}"
                    room_clean = room.strip().strip("()")
                    if room_clean:
                        cell_text += f"\n({room_clean})"
                    parts.append(cell_text)
                
                text = "\n\n".join(parts)
                worksheet.cell(row=row_idx, column=col_idx, value=text)

        # Run horizontal merge loop
        idx = 0
        while idx < len(instructional_periods):
            p = instructional_periods[idx]
            col_start = display_columns[p]
            text = worksheet.cell(row=row_idx, column=col_start).value
            
            end_idx = idx
            while end_idx + 1 < len(instructional_periods):
                next_p = instructional_periods[end_idx + 1]
                col_next = display_columns[next_p]
                next_text = worksheet.cell(row=row_idx, column=col_next).value
                
                is_adjacent = col_next == display_columns[instructional_periods[end_idx]] + 1
                
                curr_entries = day_data[idx] if idx < len(day_data) else None
                is_curr_lab = any(e.get("is_lab") for e in curr_entries if e) if curr_entries else False
                
                is_same = (text == next_text) and (text is not None)
                
                can_merge = False
                if is_same:
                    if is_adjacent:
                        can_merge = True
                    elif is_curr_lab:
                        mid_cols = range(display_columns[instructional_periods[end_idx]] + 1, col_next)
                        if all(c in break_cols or c in lunch_cols for c in mid_cols):
                            can_merge = True
                            
                if can_merge:
                    end_idx += 1
                else:
                    break
            
            if end_idx > idx:
                col_end = display_columns[instructional_periods[end_idx]]
                for c in break_cols:
                    if col_start <= c <= col_end:
                        break_overlap_rows.append(row_idx)
                for c in lunch_cols:
                    if col_start <= c <= col_end:
                        lunch_overlap_rows.append(row_idx)
                worksheet.merge_cells(start_row=row_idx, start_column=col_start, end_row=row_idx, end_column=col_end)
            idx = end_idx + 1
    
    # Merge breaks/lunches vertically in segments
    def merge_vertical_marker(column: int, label: str, blocked_rows: list[int]) -> None:
        all_rows = [day_start_row + i for i in range(len(DAYS))]
        segments: list[tuple[int, int]] = []
        segment_start = None
        for r in all_rows:
            if r in blocked_rows:
                if segment_start is not None:
                    segments.append((segment_start, r - 1))
                    segment_start = None
            else:
                if segment_start is None: segment_start = r
        if segment_start is not None: segments.append((segment_start, all_rows[-1]))

        for start, end in segments:
            worksheet.merge_cells(start_row=start, start_column=column, end_row=end, end_column=column)
            worksheet.cell(row=start, column=column, value=label)
            worksheet.cell(row=start, column=column).font = BOLD_FONT

    for bc in break_cols:
        merge_vertical_marker(bc, "BREAK", break_overlap_rows)
    for lc in lunch_cols:
        merge_vertical_marker(lc, "LUNCH", lunch_overlap_rows)
    
    # 9. Signatures
    sig_row = start_row + 15
    worksheet.cell(row=sig_row, column=1, value="HEAD OF THE DEPARTMENT")
    worksheet.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=max_col // 2)
    worksheet.cell(row=sig_row, column=max_col // 2 + 1, value="PRINCIPAL")
    worksheet.merge_cells(start_row=sig_row, start_column=max_col // 2 + 1, end_row=sig_row, end_column=max_col)
    
    grid_s = start_row + 6
    grid_e = start_row + 13
    thin_side = Side(style="thin", color="000000")
    
    for column_idx in range(1, max_col + 1):
        worksheet.column_dimensions[chr(64 + column_idx)].width = 16
    worksheet.column_dimensions["A"].width = 7
    
    for r_idx in range(start_row, sig_row + 1):
        if day_start_row <= r_idx < day_start_row + len(DAYS):
            worksheet.row_dimensions[r_idx].height = 42
        elif r_idx <= start_row + 5:
            worksheet.row_dimensions[r_idx].height = 22
        elif r_idx == sig_row:
            worksheet.row_dimensions[r_idx].height = 22
        else:
            worksheet.row_dimensions[r_idx].height = 20
    
    _style_range(worksheet, start_row, sig_row, 1, max_col, alignment=CENTER_ALIGNMENT)
    _style_range(worksheet, start_row, start_row + 5, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, start_row + 6, start_row + 7, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, day_start_row, day_start_row + len(DAYS) - 1, 1, 1, font=BOLD_FONT)
    _style_range(worksheet, sig_row, sig_row, 1, max_col, font=BOLD_FONT)
    
    for r_idx in range(grid_s, grid_e + 1):
        for c_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )
    
    for merged_range in list(worksheet.merged_cells.ranges):
        if grid_s <= merged_range.min_row <= grid_e and merged_range.min_row >= start_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = worksheet.cell(row=r, column=c)
                    _clear_cell_border(cell)
                    left_s = thin_side if c == merged_range.min_col else None
                    right_s = thin_side if c == merged_range.max_col else None
                    top_s = thin_side if r == merged_range.min_row else None
                    bottom_s = thin_side if r == merged_range.max_row else None
                    _apply_border_sides(
                        cell,
                        left=left_s,
                        right=right_s,
                        top=top_s,
                        bottom=bottom_s,
                    )
        elif merged_range.min_row >= start_row and merged_range.max_row <= sig_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    worksheet.cell(row=r, column=c).border = Border()


def _build_faculty_workload_workbook_from_details(
    faculty_schedules: dict[str, dict[str, list[list[dict] | None]]],
    timetable_metadata: dict[str, Any] | None = None,
    period_config: list[dict[str, str]] | None = None,
    faculty_id_to_name: dict[str, str] | None = None,
    printable: bool = False,
) -> Workbook:
    if period_config is None:
        period_config = DEFAULT_PERIOD_CONFIG
    workbook = Workbook()
    if not faculty_schedules:
        worksheet = workbook.active
        worksheet.title = "FacultyWorkload"
        worksheet.append(["No faculty workload data available."])
        return workbook
    workbook.remove(workbook.active)
    metadata = _resolve_timetable_metadata(timetable_metadata)

    header1 = [str(row.get("period", "")).strip() for row in period_config]
    header2 = [str(row.get("time", "")).strip() for row in period_config]
    max_col = 1 + len(period_config)
    
    display_columns = {}
    col = 2
    break_cols = []
    lunch_cols = []
    for entry in period_config:
        p = str(entry.get("period", "")).strip()
        if p.isdigit():
            display_columns[int(p)] = col
        elif "break" in p.lower():
            break_cols.append(col)
        elif "lunch" in p.lower():
            lunch_cols.append(col)
        col += 1

    global_map = _load_global_faculty_map()
    combined_map = {**global_map}
    if faculty_id_to_name:
        combined_map.update(faculty_id_to_name)
    faculty_id_to_name = combined_map

    if not printable:
        for faculty_id in sorted(faculty_schedules.keys()):
            schedule = faculty_schedules[faculty_id]
            faculty_name = faculty_id_to_name.get(faculty_id, faculty_id)
            
            sheet_title = _normalize_faculty_sheet_name(faculty_name)[:31]
            base_title = sheet_title[:28]
            counter = 1
            while sheet_title in workbook.sheetnames:
                sheet_title = f"{base_title}_{counter}"
                counter += 1
                
            worksheet = workbook.create_sheet(title=sheet_title)
            
            _write_single_workload_sheet(
                worksheet, 1, faculty_id, faculty_name, schedule,
                header1, header2, max_col, display_columns, break_cols, lunch_cols, metadata
            )
    else:
        faculties = sorted(faculty_schedules.keys())
        worksheet = workbook.create_sheet(title="Printable Workloads")
        for idx, faculty_id in enumerate(faculties):
            start_row = 1 + idx * 21
            faculty_name = faculty_id_to_name.get(faculty_id, faculty_id)
            schedule = faculty_schedules[faculty_id]
            _write_printable_workload_block(
                worksheet, start_row, faculty_id, faculty_name, schedule,
                header1, header2, max_col, display_columns, break_cols, lunch_cols, metadata
            )

    return workbook


def _build_faculty_workload_workbook(
    sessions: list[dict],
    faculty_id_to_name: dict[str, str],
    timetable_metadata: dict[str, Any] | None = None,
    period_config: list[dict[str, str]] | None = None,
    printable: bool = False,
) -> Workbook:
    instructional_periods = sorted({int(r["period"]) for r in period_config if str(r.get("period", "")).strip().isdigit()})
    faculty_schedules = _build_faculty_schedule_details(sessions, faculty_id_to_name, instructional_periods)
    return _build_faculty_workload_workbook_from_details(faculty_schedules, timetable_metadata, period_config, faculty_id_to_name, printable=printable)


def _write_single_saved_workload_sheet(
    worksheet,
    start_row: int,
    faculty_id: str,
    faculty_name: str,
    schedule: dict,
    header1: list[str],
    header2: list[str],
    max_col: int,
    display_columns: dict,
    break_cols: list[int],
    lunch_cols: list[int],
    metadata: dict,
) -> None:
    # 1. College Header
    worksheet.cell(row=start_row, column=1, value=ACADEMIC_METADATA["college"])
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    
    # 2. Autonomous line
    worksheet.cell(row=start_row+1, column=1, value="(AUTONOMOUS)")
    worksheet.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=max_col)
    
    # 3. Department
    worksheet.cell(row=start_row+2, column=1, value=ACADEMIC_METADATA["department"])
    worksheet.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+2, end_column=max_col)
    
    # 4. Academic year metadata
    worksheet.cell(row=start_row+3, column=1, value=_faculty_workload_academic_line(metadata))
    worksheet.merge_cells(start_row=start_row+3, start_column=1, end_row=start_row+3, end_column=max_col)
    
    # 5. Faculty Workload Title Row
    worksheet.cell(row=start_row+4, column=1, value="Faculty Workload")
    worksheet.merge_cells(start_row=start_row+4, start_column=1, end_row=start_row+4, end_column=max_col)
    
    # 6. Faculty ID and Name Placement
    worksheet.cell(row=start_row+5, column=1, value=f"Faculty ID : {faculty_id}")
    worksheet.cell(row=start_row+5, column=max_col // 2 + 1, value=f"Faculty Name : {faculty_name}")
    worksheet.merge_cells(start_row=start_row+5, start_column=1, end_row=start_row+5, end_column=max_col // 2)
    worksheet.merge_cells(start_row=start_row+5, start_column=max_col // 2 + 1, end_row=start_row+5, end_column=max_col)
    
    # 7. Headers
    worksheet.cell(row=start_row+6, column=1, value="DAY")
    for c_idx, h in enumerate(header1, start=2):
        worksheet.cell(row=start_row+6, column=c_idx, value=h)
    for c_idx, h in enumerate(header2, start=2):
        worksheet.cell(row=start_row+7, column=c_idx, value=h)
    worksheet.merge_cells(start_row=start_row+6, start_column=1, end_row=start_row+7, end_column=1)
    
    # 8. Grid Data
    day_start_row = start_row + 8
    break_overlap_rows = []
    lunch_overlap_rows = []
    
    for offset, day in enumerate(DAYS):
        row_idx = day_start_row + offset
        worksheet.cell(row=row_idx, column=1, value=DAY_SHORT_LABELS[day])
        for bc in break_cols:
            worksheet.cell(row=row_idx, column=bc, value="BREAK" if offset == 0 else "")
        for lc in lunch_cols:
            worksheet.cell(row=row_idx, column=lc, value="LUNCH" if offset == 0 else "")
        
        day_values = schedule.get(day, [])
        instructional_periods = sorted(display_columns.keys())
        for p_idx, val in enumerate(day_values):
            if p_idx < len(instructional_periods) and val:
                p_num = instructional_periods[p_idx]
                col_idx = display_columns[p_num]
                worksheet.cell(row=row_idx, column=col_idx, value=str(val).strip())

        # Run horizontal merge loop
        idx = 0
        while idx < len(instructional_periods):
            p = instructional_periods[idx]
            col_start = display_columns[p]
            text = worksheet.cell(row=row_idx, column=col_start).value
            
            end_idx = idx
            while end_idx + 1 < len(instructional_periods):
                next_p = instructional_periods[end_idx + 1]
                col_next = display_columns[next_p]
                next_text = worksheet.cell(row=row_idx, column=col_next).value
                
                is_adjacent = col_next == display_columns[instructional_periods[end_idx]] + 1
                is_curr_lab = text and ("LAB" in text.upper() or "PRACTICAL" in text.upper() or "PROJECT" in text.upper())
                is_same = (text == next_text) and (text is not None)
                
                can_merge = False
                if is_same:
                    if is_adjacent:
                        can_merge = True
                    elif is_curr_lab:
                        mid_cols = range(display_columns[instructional_periods[end_idx]] + 1, col_next)
                        if all(c in break_cols or c in lunch_cols for c in mid_cols):
                            can_merge = True
                            
                if can_merge:
                    end_idx += 1
                else:
                    break
            
            if end_idx > idx:
                col_end = display_columns[instructional_periods[end_idx]]
                for c in break_cols:
                    if col_start <= c <= col_end:
                        break_overlap_rows.append(row_idx)
                for c in lunch_cols:
                    if col_start <= c <= col_end:
                        lunch_overlap_rows.append(row_idx)
                worksheet.merge_cells(start_row=row_idx, start_column=col_start, end_row=row_idx, end_column=col_end)
            idx = end_idx + 1
    
    # Merge breaks/lunches vertically in segments
    def merge_vertical_marker(column: int, label: str, blocked_rows: list[int]) -> None:
        all_rows = [day_start_row + i for i in range(len(DAYS))]
        segments: list[tuple[int, int]] = []
        segment_start = None
        for r in all_rows:
            if r in blocked_rows:
                if segment_start is not None:
                    segments.append((segment_start, r - 1))
                    segment_start = None
            else:
                if segment_start is None: segment_start = r
        if segment_start is not None: segments.append((segment_start, all_rows[-1]))

        for start, end in segments:
            worksheet.merge_cells(start_row=start, start_column=column, end_row=end, end_column=column)
            worksheet.cell(row=start, column=column, value=label)
            worksheet.cell(row=start, column=column).font = BOLD_FONT

    for bc in break_cols:
        merge_vertical_marker(bc, "BREAK", break_overlap_rows)
    for lc in lunch_cols:
        merge_vertical_marker(lc, "LUNCH", lunch_overlap_rows)
    
    # 9. Signatures
    sig_row = start_row + 15
    worksheet.cell(row=sig_row, column=1, value="HEAD OF THE DEPARTMENT")
    worksheet.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=max_col // 2)
    worksheet.cell(row=sig_row, column=max_col // 2 + 1, value="PRINCIPAL")
    worksheet.merge_cells(start_row=sig_row, start_column=max_col // 2 + 1, end_row=sig_row, end_column=max_col)
    
    grid_s = start_row + 6
    grid_e = start_row + 13
    thin_side = Side(style="thin", color="000000")
    
    for column_idx in range(1, max_col + 1):
        worksheet.column_dimensions[chr(64 + column_idx)].width = 16
    worksheet.column_dimensions["A"].width = 7
    
    for r_idx in range(start_row, sig_row + 1):
        if day_start_row <= r_idx < day_start_row + len(DAYS):
            worksheet.row_dimensions[r_idx].height = 42
        elif r_idx <= start_row + 5:
            worksheet.row_dimensions[r_idx].height = 22
        elif r_idx == sig_row:
            worksheet.row_dimensions[r_idx].height = 22
        else:
            worksheet.row_dimensions[r_idx].height = 20
    
    _style_range(worksheet, start_row, sig_row, 1, max_col, alignment=CENTER_ALIGNMENT)
    _style_range(worksheet, start_row, start_row + 5, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, start_row + 6, start_row + 7, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, day_start_row, day_start_row + len(DAYS) - 1, 1, 1, font=BOLD_FONT)
    _style_range(worksheet, sig_row, sig_row, 1, max_col, font=BOLD_FONT)
    
    for r_idx in range(grid_s, grid_e + 1):
        for c_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )
    
    for merged_range in list(worksheet.merged_cells.ranges):
        if grid_s <= merged_range.min_row <= grid_e and merged_range.min_row >= start_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = worksheet.cell(row=r, column=c)
                    _clear_cell_border(cell)
                    left_s = thin_side if c == merged_range.min_col else None
                    right_s = thin_side if c == merged_range.max_col else None
                    top_s = thin_side if r == merged_range.min_row else None
                    bottom_s = thin_side if r == merged_range.max_row else None
                    _apply_border_sides(
                        cell,
                        left=left_s,
                        right=right_s,
                        top=top_s,
                        bottom=bottom_s,
                    )
        elif merged_range.min_row >= start_row and merged_range.max_row <= sig_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    worksheet.cell(row=r, column=c).border = Border()


def _write_printable_saved_workload_block(
    worksheet,
    start_row: int,
    faculty_id: str,
    faculty_name: str,
    schedule: dict,
    header1: list[str],
    header2: list[str],
    max_col: int,
    display_columns: dict,
    break_cols: list[int],
    lunch_cols: list[int],
    metadata: dict,
) -> None:
    # 1. College Header
    worksheet.cell(row=start_row, column=1, value=ACADEMIC_METADATA["college"])
    worksheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max_col)
    
    # 2. Autonomous line
    worksheet.cell(row=start_row+1, column=1, value="(AUTONOMOUS)")
    worksheet.merge_cells(start_row=start_row+1, start_column=1, end_row=start_row+1, end_column=max_col)
    
    # 3. Department
    worksheet.cell(row=start_row+2, column=1, value=ACADEMIC_METADATA["department"])
    worksheet.merge_cells(start_row=start_row+2, start_column=1, end_row=start_row+2, end_column=max_col)
    
    # 4. Academic year metadata
    worksheet.cell(row=start_row+3, column=1, value=_faculty_workload_academic_line(metadata))
    worksheet.merge_cells(start_row=start_row+3, start_column=1, end_row=start_row+3, end_column=max_col)
    
    # 5. Faculty Workload Title Row
    worksheet.cell(row=start_row+4, column=1, value="Faculty Workload")
    worksheet.merge_cells(start_row=start_row+4, start_column=1, end_row=start_row+4, end_column=max_col)
    
    # 6. Faculty ID and Name Placement
    worksheet.cell(row=start_row+5, column=1, value=f"Faculty ID : {faculty_id}")
    worksheet.cell(row=start_row+5, column=max_col // 2 + 1, value=f"Faculty Name : {faculty_name}")
    worksheet.merge_cells(start_row=start_row+5, start_column=1, end_row=start_row+5, end_column=max_col // 2)
    worksheet.merge_cells(start_row=start_row+5, start_column=max_col // 2 + 1, end_row=start_row+5, end_column=max_col)
    
    # 7. Headers
    worksheet.cell(row=start_row+6, column=1, value="DAY")
    for c_idx, h in enumerate(header1, start=2):
        worksheet.cell(row=start_row+6, column=c_idx, value=h)
    for c_idx, h in enumerate(header2, start=2):
        worksheet.cell(row=start_row+7, column=c_idx, value=h)
    worksheet.merge_cells(start_row=start_row+6, start_column=1, end_row=start_row+7, end_column=1)
    
    # 8. Grid Data
    day_start_row = start_row + 8
    break_overlap_rows = []
    lunch_overlap_rows = []
    
    for offset, day in enumerate(DAYS):
        row_idx = day_start_row + offset
        worksheet.cell(row=row_idx, column=1, value=DAY_SHORT_LABELS[day])
        for bc in break_cols:
            worksheet.cell(row=row_idx, column=bc, value="BREAK" if offset == 0 else "")
        for lc in lunch_cols:
            worksheet.cell(row=row_idx, column=lc, value="LUNCH" if offset == 0 else "")
        
        day_values = schedule.get(day, [])
        instructional_periods = sorted(display_columns.keys())
        for p_idx, val in enumerate(day_values):
            if p_idx < len(instructional_periods) and val:
                p_num = instructional_periods[p_idx]
                col_idx = display_columns[p_num]
                worksheet.cell(row=row_idx, column=col_idx, value=str(val).strip())

        # Run horizontal merge loop
        idx = 0
        while idx < len(instructional_periods):
            p = instructional_periods[idx]
            col_start = display_columns[p]
            text = worksheet.cell(row=row_idx, column=col_start).value
            
            end_idx = idx
            while end_idx + 1 < len(instructional_periods):
                next_p = instructional_periods[end_idx + 1]
                col_next = display_columns[next_p]
                next_text = worksheet.cell(row=row_idx, column=col_next).value
                
                is_adjacent = col_next == display_columns[instructional_periods[end_idx]] + 1
                is_curr_lab = text and ("LAB" in text.upper() or "PRACTICAL" in text.upper() or "PROJECT" in text.upper())
                is_same = (text == next_text) and (text is not None)
                
                can_merge = False
                if is_same:
                    if is_adjacent:
                        can_merge = True
                    elif is_curr_lab:
                        mid_cols = range(display_columns[instructional_periods[end_idx]] + 1, col_next)
                        if all(c in break_cols or c in lunch_cols for c in mid_cols):
                            can_merge = True
                            
                if can_merge:
                    end_idx += 1
                else:
                    break
            
            if end_idx > idx:
                col_end = display_columns[instructional_periods[end_idx]]
                for c in break_cols:
                    if col_start <= c <= col_end:
                        break_overlap_rows.append(row_idx)
                for c in lunch_cols:
                    if col_start <= c <= col_end:
                        lunch_overlap_rows.append(row_idx)
                worksheet.merge_cells(start_row=row_idx, start_column=col_start, end_row=row_idx, end_column=col_end)
            idx = end_idx + 1
    
    # Merge breaks/lunches vertically in segments
    def merge_vertical_marker(column: int, label: str, blocked_rows: list[int]) -> None:
        all_rows = [day_start_row + i for i in range(len(DAYS))]
        segments: list[tuple[int, int]] = []
        segment_start = None
        for r in all_rows:
            if r in blocked_rows:
                if segment_start is not None:
                    segments.append((segment_start, r - 1))
                    segment_start = None
            else:
                if segment_start is None: segment_start = r
        if segment_start is not None: segments.append((segment_start, all_rows[-1]))

        for start, end in segments:
            worksheet.merge_cells(start_row=start, start_column=column, end_row=end, end_column=column)
            worksheet.cell(row=start, column=column, value=label)
            worksheet.cell(row=start, column=column).font = BOLD_FONT

    for bc in break_cols:
        merge_vertical_marker(bc, "BREAK", break_overlap_rows)
    for lc in lunch_cols:
        merge_vertical_marker(lc, "LUNCH", lunch_overlap_rows)
    
    # 8. Signatures
    sig_row = start_row + 15
    worksheet.cell(row=sig_row, column=1, value="HEAD OF THE DEPARTMENT")
    worksheet.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=max_col // 2)
    worksheet.cell(row=sig_row, column=max_col // 2 + 1, value="PRINCIPAL")
    worksheet.merge_cells(start_row=sig_row, start_column=max_col // 2 + 1, end_row=sig_row, end_column=max_col)
    
    grid_s = start_row + 6
    grid_e = start_row + 13
    thin_side = Side(style="thin", color="000000")
    
    for column_idx in range(1, max_col + 1):
        worksheet.column_dimensions[chr(64 + column_idx)].width = 16
    worksheet.column_dimensions["A"].width = 7
    
    for r_idx in range(start_row, sig_row + 1):
        if day_start_row <= r_idx < day_start_row + len(DAYS):
            worksheet.row_dimensions[r_idx].height = 42
        elif r_idx <= start_row + 5:
            worksheet.row_dimensions[r_idx].height = 22
        elif r_idx == sig_row:
            worksheet.row_dimensions[r_idx].height = 22
        else:
            worksheet.row_dimensions[r_idx].height = 20
    
    _style_range(worksheet, start_row, sig_row, 1, max_col, alignment=CENTER_ALIGNMENT)
    _style_range(worksheet, start_row, start_row + 5, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, start_row + 6, start_row + 7, 1, max_col, font=BOLD_FONT)
    _style_range(worksheet, day_start_row, day_start_row + len(DAYS) - 1, 1, 1, font=BOLD_FONT)
    _style_range(worksheet, sig_row, sig_row, 1, max_col, font=BOLD_FONT)
    
    for r_idx in range(grid_s, grid_e + 1):
        for c_idx in range(1, max_col + 1):
            cell = worksheet.cell(row=r_idx, column=c_idx)
            cell.border = Border(
                left=thin_side,
                right=thin_side,
                top=thin_side,
                bottom=thin_side,
            )
    
    for merged_range in list(worksheet.merged_cells.ranges):
        if grid_s <= merged_range.min_row <= grid_e and merged_range.min_row >= start_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    cell = worksheet.cell(row=r, column=c)
                    _clear_cell_border(cell)
                    left_s = thin_side if c == merged_range.min_col else None
                    right_s = thin_side if c == merged_range.max_col else None
                    top_s = thin_side if r == merged_range.min_row else None
                    bottom_s = thin_side if r == merged_range.max_row else None
                    _apply_border_sides(
                        cell,
                        left=left_s,
                        right=right_s,
                        top=top_s,
                        bottom=bottom_s,
                    )
        elif merged_range.min_row >= start_row and merged_range.max_row <= sig_row:
            for r in range(merged_range.min_row, merged_range.max_row + 1):
                for c in range(merged_range.min_col, merged_range.max_col + 1):
                    worksheet.cell(row=r, column=c).border = Border()


def _build_faculty_workload_workbook_from_saved_workloads(
    faculty_workloads: dict[str, dict[str, list[str | None]]],
    timetable_metadata: dict[str, Any] | None = None,
    faculty_id_to_name: dict[str, str] | None = None,
    period_config: list[dict[str, str]] | None = None,
    printable: bool = False,
) -> Workbook:
    if period_config is None:
        period_config = DEFAULT_PERIOD_CONFIG
    workbook = Workbook()
    if not faculty_workloads:
        worksheet = workbook.active
        worksheet.title = "FacultyWorkload"
        worksheet.append(["No faculty workload data available."])
        return workbook
    workbook.remove(workbook.active)
    metadata = _resolve_timetable_metadata(timetable_metadata)

    header1 = [str(row.get("period", "")).strip() for row in period_config]
    header2 = [str(row.get("time", "")).strip() for row in period_config]
    max_col = 1 + len(period_config)
    
    display_columns = {}
    col = 2
    break_cols = []
    lunch_cols = []
    for entry in period_config:
        p = str(entry.get("period", "")).strip()
        if p.isdigit():
            display_columns[int(p)] = col
        elif "break" in p.lower():
            break_cols.append(col)
        elif "lunch" in p.lower():
            lunch_cols.append(col)
        col += 1

    global_map = _load_global_faculty_map()
    combined_map = {**global_map}
    if faculty_id_to_name:
        combined_map.update(faculty_id_to_name)
    faculty_id_to_name = combined_map

    name_to_id = {}
    if faculty_id_to_name:
        for fid, fname in faculty_id_to_name.items():
            name_to_id[fname.strip().lower()] = fid.strip()

    # Group workloads by faculty_id
    grouped_workloads = {}
    for key, schedule in faculty_workloads.items():
        key_clean = key.strip()
        if faculty_id_to_name and key_clean in faculty_id_to_name:
            fid = key_clean
        else:
            fid = _find_faculty_id(key_clean, name_to_id) or key_clean
        
        if fid not in grouped_workloads:
            grouped_workloads[fid] = {day: [None] * len(PERIODS) for day in DAYS}
        
        for day in DAYS:
            day_values = schedule.get(day, [])
            for p_idx, val in enumerate(day_values):
                if val:
                    if p_idx < len(grouped_workloads[fid][day]):
                        curr_val = grouped_workloads[fid][day][p_idx]
                        if curr_val is None:
                            grouped_workloads[fid][day][p_idx] = str(val).strip()
                        else:
                            val_str = str(val).strip()
                            curr_str = str(curr_val).strip()
                            if val_str not in curr_str:
                                grouped_workloads[fid][day][p_idx] = f"{curr_str}\n{val_str}"

    if not printable:
        for faculty_id in sorted(grouped_workloads.keys()):
            schedule = grouped_workloads[faculty_id]
            faculty_name = faculty_id_to_name.get(faculty_id, faculty_id)
            
            sheet_title = _normalize_faculty_sheet_name(faculty_name)[:31]
            base_title = sheet_title[:28]
            counter = 1
            while sheet_title in workbook.sheetnames:
                sheet_title = f"{base_title}_{counter}"
                counter += 1
                
            worksheet = workbook.create_sheet(title=sheet_title)
            
            _write_single_saved_workload_sheet(
                worksheet, 1, faculty_id, faculty_name, schedule,
                header1, header2, max_col, display_columns, break_cols, lunch_cols, metadata
            )
    else:
        faculties = sorted(grouped_workloads.keys())
        worksheet = workbook.create_sheet(title="Printable Workloads")
        for idx, faculty_id in enumerate(faculties):
            start_row = 1 + idx * 21
            faculty_name = faculty_id_to_name.get(faculty_id, faculty_id)
            schedule = grouped_workloads[faculty_id]
            _write_printable_saved_workload_block(
                worksheet, start_row, faculty_id, faculty_name, schedule,
                header1, header2, max_col, display_columns, break_cols, lunch_cols, metadata
            )

    return workbook


def _build_faculty_schedule_details_from_section_grids(
    section_grids: dict[tuple[str, str], dict[str, list[dict | None]]],
) -> tuple[dict[str, dict[str, list[list[dict] | None]]], dict[str, str]]:
    name_to_id = {}
    faculty_id_to_name = _load_global_faculty_map()
    if faculty_id_to_name:
        for fid, fname in faculty_id_to_name.items():
            name_to_id[fname.strip().lower()] = fid.strip()

    local_fac_map = {}
    workloads: dict[str, dict[str, list[list[dict] | None]]] = {}
    for (year, section), grid in section_grids.items():
        for day in DAYS:
            slots = list(grid.get(day, []))
            for idx, cell in enumerate(slots[: len(PERIODS)]):
                if not isinstance(cell, dict):
                    continue
                faculty_id = str(cell.get("facultyId") or cell.get("faculty_id") or "").strip()
                faculty_name = str(cell.get("facultyName") or cell.get("faculty") or "").strip()
                subject_name = str(cell.get("subjectName") or cell.get("subject") or "").strip()
                if not subject_name:
                    continue
                if not faculty_id and not faculty_name:
                    continue
                
                if faculty_id and faculty_name:
                    ids = [fid.strip() for fid in faculty_id.split(",") if fid.strip()]
                    names = [fname.strip() for fname in faculty_name.split(",") if fname.strip()]
                    for i, fid in enumerate(ids):
                        clean_fid = fid[:-2] if fid.endswith(".0") else fid
                        if i < len(names):
                            local_fac_map[clean_fid] = names[i]
                
                faculty_ids = []
                if faculty_id:
                    faculty_ids = [fid.strip() for fid in faculty_id.split(",") if fid.strip()]
                
                if not faculty_ids and faculty_name:
                    faculty_names = [name.strip() for name in faculty_name.split(",") if name.strip()]
                    for fname in faculty_names:
                        fid = _find_faculty_id(fname, name_to_id)
                        if fid:
                            faculty_ids.append(fid)
                            local_fac_map[fid] = fname
                        else:
                            faculty_ids.append(fname)
                            local_fac_map[fname] = fname
                
                clean_ids = []
                for fid in faculty_ids:
                    if fid.endswith(".0"):
                        fid = fid[:-2]
                    clean_ids.append(fid)
                
                sections = cell.get("sharedSections") or []
                section_label = ",".join(str(item).strip() for item in sections if str(item).strip()) or section
                detail = {
                    "subject": subject_name,
                    "year": year,
                    "section": section_label,
                    "classroom": str(cell.get("classroom", "")).strip(),
                    "lab_room": str(cell.get("labRoom") or cell.get("venue") or "").strip(),
                    "is_lab": bool(cell.get("isLab")),
                }
                
                for fid in clean_ids:
                    schedule = workloads.setdefault(
                        fid,
                        {day_name: [None] * len(PERIODS) for day_name in DAYS},
                    )
                    if schedule[day][idx] is None:
                        schedule[day][idx] = []
                    existing = schedule[day][idx] or []
                    if detail not in existing:
                        existing.append(detail.copy())
                    schedule[day][idx] = existing
    return workloads, local_fac_map


def _build_shared_classes_workbook(shared_sessions: list[dict]) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "SharedClasses"
    worksheet.append(["YEAR", "SUBJECT", "FACULTY", "SECTIONS", "DAY", "PERIODS"])
    for session in shared_sessions:
        worksheet.append(
            [
                session.get("year", ""),
                session.get("subject_name", "") or session.get("subject_id", ""),
                session.get("faculty_name", "") or session.get("faculty_id", ""),
                ",".join(session.get("sections", [])),
                session.get("day", ""),
                ",".join(str(period) for period in session.get("periods", [])),
            ]
        )
    return workbook


def _build_constraint_report_workbook(violations: list[dict], unscheduled: list[dict], shared_classes: list[dict]) -> Workbook:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "ConstraintViolations"
    worksheet.append(["YEAR", "SECTIONS", "SUBJECT", "FACULTY", "CONSTRAINT", "DETAIL"])
    for violation in violations:
        worksheet.append(
            [
                violation.get("year", ""),
                ",".join(violation.get("sections", [])),
                violation.get("subject_name", "") or violation.get("subject_id", ""),
                violation.get("faculty_name", "") or violation.get("faculty_id", ""),
                violation.get("constraint", ""),
                violation.get("detail", ""),
            ]
        )
    for item in unscheduled:
        worksheet.append(
            [
                item.get("year", ""),
                ",".join(item.get("sections", [])),
                item.get("subject_name", "") or item.get("subject_id", ""),
                item.get("faculty_name", "") or item.get("faculty_id", ""),
                "unscheduled subject",
                item.get("detail", ""),
            ]
        )
        
    shared_sheet = workbook.create_sheet(title="SharedClasses")
    shared_sheet.append(["DAY", "PERIODS", "YEAR", "SECTIONS", "SUBJECT", "FACULTY", "ROOM"])
    for sc in shared_classes:
        shared_sheet.append([
            sc.get("day", ""),
            ",".join(str(p) for p in sc.get("periods", [])),
            sc.get("year", ""),
            ",".join(sc.get("sections", [])),
            sc.get("subject_name", "") or sc.get("subject_id", ""),
            ", ".join(sc.get("faculty_names", [])),
            sc.get("classroom", "") or sc.get("labRoom", "") or sc.get("venue", ""),
        ])
        
    return workbook


def _check_faculty_workload_conflicts(
    year: str,
    sections: list[str],
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    instructional_periods: list[int],
    existing_workloads_payload: dict | None,
    faculty_id_to_name: dict[str, str],
) -> list[dict]:
    conflicts = []

    # 1. Double Booking check (same slot, multiple sections, not shared)
    for day in DAYS:
        for period in instructional_periods:
            faculty_assignments = {}  # faculty_id -> list of (section, cell)
            for section in sections:
                cell = schedules.get((year, section), {}).get(day, {}).get(period)
                if not isinstance(cell, dict):
                    continue

                # Extract faculty IDs
                faculty_raw = str(cell.get("facultyId") or "").strip()
                if not faculty_raw:
                    continue

                # Split multiple faculty tokens
                faculty_ids = _split_faculty_tokens(faculty_raw) or (faculty_raw,)
                for fid in faculty_ids:
                    fid_clean = _normalize_id_token(fid)
                    if fid_clean:
                        faculty_assignments.setdefault(fid_clean, []).append((section, cell))

            # Check for conflicts
            for fid_clean, assignments in faculty_assignments.items():
                if len(assignments) <= 1:
                    continue

                # If there are multiple assignments, check if they are shared classes
                for i in range(len(assignments)):
                    for j in range(i + 1, len(assignments)):
                        sec1, cell1 = assignments[i]
                        sec2, cell2 = assignments[j]

                        # Are they shared?
                        shared1 = str(sec2).strip() in [str(s).strip() for s in (cell1.get("sharedSections") or [])]
                        shared2 = str(sec1).strip() in [str(s).strip() for s in (cell2.get("sharedSections") or [])]
                        same_subject = cell1.get("subjectId") == cell2.get("subjectId")

                        if not (shared1 and shared2 and same_subject):
                            fac_name = faculty_id_to_name.get(fid_clean, fid_clean)
                            sub1 = cell1.get("subjectName") or cell1.get("subject") or cell1.get("subjectId")
                            sub2 = cell2.get("subjectName") or cell2.get("subject") or cell2.get("subjectId")
                            conflicts.append({
                                "year": year,
                                "sections": [sec1, sec2],
                                "subject_id": cell1.get("subjectId") or "",
                                "faculty_id": fid_clean,
                                "faculty_name": fac_name,
                                "constraint": "faculty workload conflict",
                                "detail": f"Faculty {fac_name} is double-booked on {day} Period {period} in Section {sec1} ({sub1}) and Section {sec2} ({sub2})."
                            })

    # 2. Existing Workload collision check
    if existing_workloads_payload and existing_workloads_payload.get("rows"):
        name_to_id = {v.strip().lower(): k for k, v in faculty_id_to_name.items() if v.strip()}
        for row in existing_workloads_payload["rows"]:
            fid = _normalize_id_token(row.get("faculty_id", ""))
            fname = str(row.get("faculty_name", "")).strip().lower()
            if not fid and fname in name_to_id:
                fid = name_to_id[fname]
            day = _normalize_day(row.get("day", ""))
            period = int(row.get("period", 0))

            if fid and day and period in instructional_periods:
                # Check if this faculty is scheduled in the generated timetable in this slot
                for section in sections:
                    cell = schedules.get((year, section), {}).get(day, {}).get(period)
                    if isinstance(cell, dict):
                        # Extract faculty IDs from generated cell
                        faculty_raw = str(cell.get("facultyId") or "").strip()
                        faculty_ids = _split_faculty_tokens(faculty_raw) or (faculty_raw,)
                        if any(_normalize_id_token(f) == fid for f in faculty_ids):
                            fac_name = faculty_id_to_name.get(fid, row.get("faculty_name", fid))
                            sub_name = cell.get("subjectName") or cell.get("subject") or cell.get("subjectId")
                            conflicts.append({
                                "year": year,
                                "sections": [section],
                                "subject_id": cell.get("subjectId") or "",
                                "faculty_id": fid,
                                "faculty_name": fac_name,
                                "constraint": "faculty workload conflict",
                                "detail": (
                                    f"Faculty {fac_name} is scheduled on {day} Period {period} in Section {section} ({sub_name}) "
                                    f"but has a conflict with an existing workload mapping value: '{row.get('cell_value', 'Already Scheduled')}'."
                                )
                            })

    return conflicts


def generate_timetable(
    request_data: GenerateTimetableRequest,
    store: MemoryStore,
    precheck_only: bool = False,
) -> dict:
    request_data.year = normalize_year(request_data.year)
    year = request_data.year
    timetable_metadata = request_data.timetableMetadata.model_dump()
    preferred_daily_limit = max(
        1,
        min(
            PREFERRED_SUBJECT_PERIODS_PER_DAY,
            int(request_data.dailySubjectLimit or PREFERRED_SUBJECT_PERIODS_PER_DAY),
        ),
    )

    subject_id_to_name, compulsory_continuous = _build_subject_maps(request_data, store)
    room_inventory = _build_room_inventory(request_data, store)
    section_strength_map = _build_section_strength_map(request_data, store)
    section_home_room_map = _build_section_home_room_map(request_data, store)
    classrooms = [str(item.get("name", "")).strip() for item in room_inventory if str(item.get("name", "")).strip()]
    lab_room_names = {
        str(item.get("name", "")).strip()
        for item in room_inventory
        if bool(item.get("is_lab")) and str(item.get("name", "")).strip()
    }
    room_capacity_map = {
        str(item.get("name", "")).strip(): (
            int(item.get("capacity")) if isinstance(item.get("capacity"), int) else None
        )
        for item in room_inventory
        if str(item.get("name", "")).strip()
    }
    all_fixed_classroom_blocks = _build_fixed_classroom_blocks(request_data, store)
    fixed_classroom_blocks = {
        (section, day, period): classroom
        for (block_year, section, day, period), classroom in all_fixed_classroom_blocks.items()
        if block_year == year
    }
    
    period_config = _build_period_config(request_data, store)
    instructional_periods, sessions = _derive_sessions(period_config)
    session_adjacency = _build_session_adjacency(sessions)
    
    max_period = max(instructional_periods) if instructional_periods else 7
    capacity_per_section = len(instructional_periods) * 6

    # Re-normalize year and faculty id map
    faculty_id_to_name = _build_faculty_maps(request_data, store)
    # Strip .0 from numeric faculty IDs
    faculty_id_to_name = {
        (k[:-2] if k.endswith(".0") else k): v
        for k, v in faculty_id_to_name.items()
    }

    # Fetch payloads
    main_payload = store.get_scoped_mapping("main_timetable_config", "global")
    lab_payload = store.get_scoped_mapping("lab_timetable_config", "global")
    shared_payload = store.get_scoped_mapping("shared_classes", "global")

    all_main_rows = list(main_payload.get("rows", []) if main_payload else [])
    all_lab_rows = list(lab_payload.get("rows", []) if lab_payload else [])
    raw_main_rows = [
        row
        for row in all_main_rows
        if normalize_year(str(row.get("year", ""))) == year
    ]
    raw_lab_rows = [
        row
        for row in all_lab_rows
        if normalize_year(str(row.get("year", ""))) == year
    ]

    for entry in request_data.manualEntries:
        entry_year = normalize_year(entry.year)
        if entry_year != year:
            continue
        fid = ",".join(
            _canonicalize_faculty_token(token, faculty_id_to_name)
            for token in _split_faculty_tokens(str(entry.facultyId))
            if _canonicalize_faculty_token(token, faculty_id_to_name)
        )
        subject_token = _normalize_id_token(entry.subjectId)
        
        raw_main_rows.append(
            {
                "year": entry_year,
                "section": str(entry.section).strip(),
                "subject_id": subject_token,
                "faculty_id": fid,
                "hours": int(entry.noOfHours),
                "continuous_hours": int(entry.continuousHours or 1),
                "min_consecutive_hours": max(1, int(entry.compulsoryContinuousHours or 1)),
                "max_consecutive_hours": max(
                    max(1, int(entry.compulsoryContinuousHours or 1)),
                    int(entry.continuousHours or 1),
                ),
            }
        )
        all_main_rows.append(raw_main_rows[-1])
        if entry.compulsoryContinuousHours:
            compulsory_continuous[subject_token] = max(1, int(entry.compulsoryContinuousHours))

    for lab in request_data.manualLabEntries:
        if normalize_year(lab.year) != year:
            continue
        raw_lab_rows.append(
            {
                "year": year,
                "section": str(lab.section).strip(),
                "subject_id": _normalize_id_token(lab.subjectId),
                "day": int(lab.day),
                "hours": [int(hour) for hour in lab.hours],
                "venue": str(lab.venue).strip(),
            }
        )
        all_lab_rows.append(raw_lab_rows[-1])

    if not raw_main_rows:
        raise HTTPException(status_code=400, detail="No configurable sections found for the specified year.")

    section_total_hours: dict[str, int] = {}
    main_rows_by_section_subject: dict[tuple[str, str], dict] = {}
    section_subject_faculty: dict[tuple[str, str], str] = {}
    for row in raw_main_rows:
        section = str(row.get("section", "")).strip()
        subject_id = _normalize_id_token(row.get("subject_id", ""))
        faculty_id = ",".join(
            _canonicalize_faculty_token(token, faculty_id_to_name)
            for token in _split_faculty_tokens(str(row.get("faculty_id", "")))
            if _canonicalize_faculty_token(token, faculty_id_to_name)
        )
            
        hours = int(row.get("hours", 0) or 0)
        continuous_hours = max(1, int(row.get("continuous_hours", 1) or 1))
        compulsory_hours = max(
            1,
            int(
                row.get("min_consecutive_hours")
                or row.get("compulsory_continuous_hours")
                or compulsory_continuous.get(subject_id, 1)
                or 1
            ),
        )
        row["min_consecutive_hours"] = min(hours, compulsory_hours) if hours > 0 else compulsory_hours
        row["max_consecutive_hours"] = max(
            row["min_consecutive_hours"],
            int(row.get("max_consecutive_hours", continuous_hours) or continuous_hours),
            continuous_hours,
        )
        if not section or not subject_id:
            continue
        section_total_hours[section] = section_total_hours.get(section, 0) + hours
        main_rows_by_section_subject[(section, subject_id)] = row
        section_subject_faculty[(section, subject_id)] = faculty_id

    all_sections = sorted(section_total_hours.keys())
    validation_errors = []
    for section, total in section_total_hours.items():
        if total != capacity_per_section:
            validation_errors.append(
                {
                    "year": year,
                    "section": section,
                    "hours": total,
                    "expected": capacity_per_section,
                    "detail": f"Section {section} needs exactly {capacity_per_section} hours but has {total} hour(s) configured.",
                }
            )

    all_sections = sorted(section_total_hours.keys())
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]] = {
        (year, section): {day: {period: None for period in instructional_periods} for day in DAYS} for section in all_sections
    }

    all_faculties = {
        token
        for item in main_rows_by_section_subject.values()
        for token in _split_faculty_tokens(str(item.get("faculty_id", "")).strip())
        if token
    }
    faculty_availability = _build_faculty_availability(request_data, store, all_faculties, faculty_id_to_name, instructional_periods)
    
    prior_timetable_ids = [
        str(timetable_id).strip()
        for timetable_id in request_data.priorTimetableIds
        if str(timetable_id).strip()
    ]

    global_lab_busy = _build_global_lab_occupancy(
        all_main_rows,
        all_lab_rows,
        year,
        faculty_id_to_name,
        faculty_availability,
        instructional_periods,
    )
    prior_faculty_busy = _build_prior_faculty_occupancy(prior_timetable_ids, store, faculty_id_to_name)
    prior_faculty_section_slots = _build_prior_faculty_section_assignments(prior_timetable_ids, store, faculty_id_to_name)
    faculty_busy: dict[str, set[tuple[str, int]]] = {
        faculty_id: slots.copy() for faculty_id, slots in prior_faculty_busy.items()
    }
    faculty_section_slots: dict[str, dict[tuple[str, int], tuple[str, ...]]] = {
        faculty_id: slot_map.copy() for faculty_id, slot_map in prior_faculty_section_slots.items()
    }
    for faculty_id, slots in global_lab_busy.items():
        faculty_busy.setdefault(faculty_id, set()).update(slots)

    # Global Room Occupancy: Prevent different years from using same room at same time
    prior_room_busy = _build_prior_room_occupancy(prior_timetable_ids, store)
    global_lab_room_busy = _build_global_lab_room_occupancy(all_lab_rows, year, instructional_periods)
    for room, slots in global_lab_room_busy.items():
        prior_room_busy.setdefault(room, set()).update(slots)

    # Load and integrate existing faculty workloads from workbook (Requirement 10, 11, 12, 16, 17)
    existing_workloads_payload = store.get_scoped_mapping("existing_faculty_workloads", "global")
    if existing_workloads_payload and existing_workloads_payload.get("rows"):
        name_to_id = {v.strip().lower(): k for k, v in faculty_id_to_name.items() if v.strip()}
        for row in existing_workloads_payload["rows"]:
            fid = _normalize_id_token(row.get("faculty_id", ""))
            fname = str(row.get("faculty_name", "")).strip().lower()
            if not fid and fname in name_to_id:
                fid = name_to_id[fname]
            day = _normalize_day(row.get("day", ""))
            period = int(row.get("period", 0))
            if fid and day and period in instructional_periods:
                faculty_busy.setdefault(fid, set()).add((day, period))

    # Load and integrate existing classroom timetables from workbook (Requirement 10, 13, 14, 15, 16, 17)
    existing_classrooms_payload = store.get_scoped_mapping("existing_classroom_timetables", "global")
    if existing_classrooms_payload and existing_classrooms_payload.get("rows"):
        for row in existing_classrooms_payload["rows"]:
            room = _normalize_id_token(row.get("classroom", ""))
            day = _normalize_day(row.get("day", ""))
            period = int(row.get("period", 0))
            if room and day and period in instructional_periods:
                prior_room_busy.setdefault(room, set()).add((day, period))
        
    session_log: list[dict] = []
    constraint_violations: list[dict] = []
    lab_assigned_hours: dict[tuple[str, str], int] = {}
    locked_hours_by_section: dict[str, int] = {}

    for item in validation_errors:
        constraint_violations.append(
            {
                "year": item.get("year", year),
                "sections": [str(item.get("section", "")).strip()] if str(item.get("section", "")).strip() else [],
                "subject_id": "",
                "faculty_id": "",
                "constraint": "section capacity constraint",
                "detail": str(item.get("detail", "")).strip()
                or f"Section totals should be {capacity_per_section} hours.",
            }
        )

    shared_constraints: dict[str, list[tuple[str, ...]]] = {}
    for row in (shared_payload.get("rows", []) if shared_payload else []):
        if normalize_year(str(row.get("year", ""))) != year:
            continue
        sub_id = _normalize_id_token(row.get("subject_id", "") or row.get("subject", ""))
        secs = tuple(sorted(_normalize_id_token(s) for s in row.get("sections", []) if _normalize_id_token(s)))
        if sub_id and secs:
            shared_constraints.setdefault(sub_id, []).append(secs)
    for entry in request_data.sharedClasses:
        if normalize_year(entry.year) != year:
            continue
        subject_id = _normalize_id_token(entry.subject)
        sections = tuple(sorted(_normalize_id_token(section) for section in entry.sections if _normalize_id_token(section)))
        if subject_id and sections:
            shared_constraints.setdefault(subject_id, []).append(sections)

    # Process Lab Rows
    sorted_lab_rows = sorted(
        raw_lab_rows,
        key=lambda row: (
            int(row.get("day", 0) or 0),
            tuple(sorted(int(period) for period in row.get("hours", []) if int(period) in instructional_periods)),
            str(row.get("subject_id", "")).strip(),
            str(row.get("section", "")).strip(),
        ),
    )
    
    lab_groups: dict[tuple[str, str, str, tuple[int, ...]], dict] = {}
    for row in sorted_lab_rows:
        section = str(row.get("section", "")).strip()
        subject_id = _normalize_id_token(row.get("subject_id", ""))
        explicit_sections = [str(item).strip() for item in row.get("sections", []) if str(item).strip()]
        sections = sorted(set(explicit_sections or ([section] if section else [])))
        faculty_id = section_subject_faculty.get((section, subject_id), "")
        day = _normalize_day(int(row.get("day", 0) or 0))
        periods = tuple(sorted(int(period) for period in row.get("hours", []) if int(period) in instructional_periods))
        
        venue = str(row.get("venue", "")).strip()
        if not sections or any(sec not in all_sections for sec in sections):
            continue
        if not day or not periods:
            continue
            
        group_key = (faculty_id, subject_id, day, periods)
        current_group = lab_groups.setdefault(group_key, {"sections": set(), "venue": venue})
        current_group["sections"].update(sections)
        if venue and not current_group["venue"]:
            current_group["venue"] = venue

    for (faculty_id, subject_id, day, periods), group in sorted(lab_groups.items()):
        sections = sorted(group["sections"])
        venue = group.get("venue", "")
        subject_id_resolved, subject_name = _resolve_subject_output(subject_id, subject_id_to_name)
        faculty_options_for_lab = _resolve_faculty_pool(faculty_id, faculty_id_to_name, faculty_availability)
        selected_faculty = _pick_best_faculty_option_for_locked_session(
            faculty_options_for_lab,
            day,
            periods,
            faculty_busy,
            faculty_availability,
            instructional_periods,
            tuple(sections),
            faculty_section_slots,
            session_adjacency,
        )
        if selected_faculty:
            faculty_ids_resolved = _split_faculty_tokens(selected_faculty)
        else:
            fallback_tokens = _split_faculty_tokens(faculty_id)
            faculty_ids_resolved = fallback_tokens
        
        faculty_id_resolved, faculty_name = _resolve_faculty_display(faculty_ids_resolved, faculty_id_to_name)
        
        placed_sections: set[str] = set()
        placed_periods: list[int] = []
        for period in periods:
            period_sections: list[str] = []
            for section in sections:
                if schedules[(year, section)][day][period] is not None:
                    continue
                schedules[(year, section)][day][period] = {
                    "subject": subject_name,
                    "subjectName": subject_name,
                    "subjectId": subject_id_resolved,
                    "faculty": faculty_name,
                    "facultyName": faculty_name,
                    "facultyId": faculty_id_resolved,
                    "isLab": True,
                    "locked": True,
                    "venue": venue,
                    "sharedSections": sections if len(sections) > 1 else [],
                }
                lab_assigned_hours[(section, subject_id)] = lab_assigned_hours.get((section, subject_id), 0) + 1
                locked_hours_by_section[section] = locked_hours_by_section.get(section, 0) + 1
                period_sections.append(section)
                placed_sections.add(section)
            if period_sections:
                placed_periods.append(period)
                for faculty_token in faculty_ids_resolved:
                    faculty_busy.setdefault(faculty_token, set()).add((day, period))
        
        if placed_periods:
            _mark_faculty_section_assignment(
                faculty_ids_resolved,
                day,
                placed_periods,
                sorted(placed_sections),
                faculty_section_slots,
            )
            session_log.append({
                "year": year,
                "subject_id": subject_id_resolved,
                "subject_name": subject_name,
                "faculty_id": faculty_id_resolved,
                "faculty_name": faculty_name,
                "faculty_ids": list(faculty_ids_resolved),
                "sections": sorted(placed_sections),
                "day": day,
                "periods": placed_periods,
                "venue": venue,
                "isLab": True,
                "source": "lab",
            })

    daily_subject_counts = _build_daily_subject_counts(
        schedules,
        year,
        all_sections,
        instructional_periods,
    )

    # Prepare Requirements
    requirements: list[Requirement] = []
    covered_shared_subjects: set[tuple[str, str]] = set()

    def build_requirement_from_rows(subject_id: str, sections: tuple[str, ...], rows: list[dict], shared: bool) -> Requirement | None:
        raw_faculty_ids = {_normalize_faculty_field(r.get("faculty_id", "")) for r in rows if r}
        faculty_pools = {
            tuple(sorted(set(_resolve_faculty_pool(fid, faculty_id_to_name, faculty_availability))))
            for fid in raw_faculty_ids if fid
        }
        remaining_hours = [
            max(0, int(r["hours"]) - lab_assigned_hours.get((str(r["section"]), subject_id), 0))
            for r in rows if r
        ]
        if not remaining_hours or len(set(remaining_hours)) != 1 or remaining_hours[0] <= 0:
            return None
            
        hours = remaining_hours[0]
        faculty_options = next(iter(sorted(faculty_pools)), ())
        faculty_label = sorted(raw_faculty_ids)[0] if raw_faculty_ids else ""
        
        min_consecutive = max(int(r.get("min_consecutive_hours", 1) or 1) for r in rows if r)
        max_consecutive = min(int(r.get("max_consecutive_hours", 1) or 1) for r in rows if r)
        
        faculty_team = ()
        if len(faculty_options) > 1:
            faculty_team = faculty_options
            faculty_options = ()

        return Requirement(
            subject_id=subject_id,
            faculty_id=faculty_label,
            faculty_options=faculty_options,
            faculty_team=faculty_team,
            sections=sections,
            hours=hours,
            min_consecutive_hours=min_consecutive,
            max_consecutive_hours=max_consecutive,
            shared=shared,
            preferred_daily_limit=preferred_daily_limit,
            hard_daily_limit=HARD_SUBJECT_PERIODS_PER_DAY,
            daily_limit_exempt=(
                min_consecutive > 1
                or hours > HARD_SUBJECT_PERIODS_PER_DAY * len(DAYS)
            ),
            phase=_determine_requirement_phase(shared, min_consecutive, hours, faculty_options, faculty_availability),
        )

    # ... Shared Classes processing ...
    for subject_id, groups in sorted(shared_constraints.items()):
        merged_groups = _merge_overlapping_section_groups(list(set(groups)))
        for sections in sorted(merged_groups):
            section_rows = [main_rows_by_section_subject.get((sec, subject_id)) for sec in sections]
            if any(r is None for r in section_rows): continue
            req = build_requirement_from_rows(subject_id, sections, [r for r in section_rows if r], True)
            if req:
                requirements.append(req)
                for sec in sections: covered_shared_subjects.add((sec, subject_id))

    for (section, subject_id), row in sorted(main_rows_by_section_subject.items()):
        if (section, subject_id) in covered_shared_subjects: continue
        req = build_requirement_from_rows(subject_id, (section,), [row], False)
        if req: requirements.append(req)

    # ── FAST FEASIBILITY PRE-CHECK ──────────────────────────────────────────
    # When called with precheck_only=True (the /feasibility endpoint), we do
    # NOT run the solver.  Instead we perform O(R + S + F) approximate checks
    # that finish in < 100 ms even for large inputs.
    #
    # The checks are intentionally *over-approximate*: if they say "feasible"
    # the solver might still fail (tight faculty conflicts, etc.), but if they
    # say "infeasible" it is a guaranteed hard failure that the solver cannot
    # overcome either.
    if precheck_only:
        issues: list[dict] = list(constraint_violations)  # includes capacity errors
        total_instructional_slots = len(DAYS) * len(instructional_periods)

        # ── (a) Section capacity: required_hours ≤ free_slots ─────────────
        # Safe because every section-period can hold at most 1 subject.
        section_required: dict[str, int] = {}
        for req in requirements:
            for sec in req.sections:
                section_required[sec] = section_required.get(sec, 0) + req.hours

        for section in all_sections:
            free = sum(
                1 for day in DAYS for p in instructional_periods
                if schedules[(year, section)][day].get(p) is None
            )
            needed = section_required.get(section, 0)
            if needed > free:
                issues.append({
                    "year": year,
                    "sections": [section],
                    "subject_id": "",
                    "faculty_id": "",
                    "constraint": "section capacity overload",
                    "detail": (
                        f"Section {section} needs {needed} theory hour(s) "
                        f"but only {free} free slot(s) remain after labs."
                    ),
                })

        # ── (b) Faculty capacity: assigned_hours ≤ available_periods ──────
        # Safe because a faculty member cannot teach in a period they are
        # unavailable or already busy in.
        faculty_required: dict[str, int] = {}
        for req in requirements:
            tokens = req.faculty_options or ((req.faculty_id,) if req.faculty_id else ())
            # Distribute to the FIRST (primary) option; if pool, the solver
            # can redistribute, so we use max-available as an optimistic bound.
            if tokens:
                faculty_required[tokens[0]] = faculty_required.get(tokens[0], 0) + req.hours

        for fid, needed in faculty_required.items():
            avail = faculty_availability.get(fid, _default_day_availability(instructional_periods))
            busy = faculty_busy.get(fid, set())
            total_free = sum(
                1 for day in DAYS for p in avail.get(day, set())
                if (day, p) not in busy
            )
            if needed > total_free:
                fname = faculty_id_to_name.get(fid, fid)
                issues.append({
                    "year": year,
                    "sections": [],
                    "subject_id": "",
                    "faculty_id": fid,
                    "constraint": "faculty capacity overload",
                    "detail": (
                        f"Faculty {fname} ({fid}) is assigned {needed} hour(s) "
                        f"but only has {total_free} free period(s) this week."
                    ),
                })

        # ── (c) Shared-class check: hours ≤ min(free slots of sections) ───
        # For shared classes all involved sections must be free at the SAME
        # time, so the bottleneck is the section with the fewest free slots.
        for req in requirements:
            if not req.shared or len(req.sections) <= 1:
                continue
            min_free = min(
                sum(
                    1 for day in DAYS for p in instructional_periods
                    if schedules[(year, sec)][day].get(p) is None
                )
                for sec in req.sections
            )
            if req.hours > min_free:
                issues.append({
                    "year": year,
                    "sections": list(req.sections),
                    "subject_id": req.subject_id,
                    "faculty_id": req.faculty_id,
                    "constraint": "shared class capacity",
                    "detail": (
                        f"Shared subject {req.subject_id} needs {req.hours}h "
                        f"but the tightest section only has {min_free} free slot(s)."
                    ),
                })

        # ── (d) Continuous-block approximation ────────────────────────────
        # If a subject requires N consecutive periods but no session has N
        # contiguous instructional periods, placement is impossible.
        max_session_len = max((len(s) for s in sessions), default=0)
        for req in requirements:
            if req.min_consecutive_hours <= 1:
                continue
            needed_block = req.min_consecutive_hours
            # 4-hour blocks can span two adjacent sessions
            effective_max = max_session_len
            if needed_block == 4 and len(sessions) >= 2:
                for i in range(len(sessions) - 1):
                    effective_max = max(effective_max, len(sessions[i]) + len(sessions[i + 1]))
            if needed_block > effective_max:
                issues.append({
                    "year": year,
                    "sections": list(req.sections),
                    "subject_id": req.subject_id,
                    "faculty_id": req.faculty_id,
                    "constraint": "continuous hours impossible",
                    "detail": (
                        f"Subject {req.subject_id} requires {needed_block} "
                        f"consecutive period(s) but longest session is {max_session_len}."
                    ),
                })

        precheck_free_slots_tracker = {
            section: sum(
                1
                for day in DAYS
                for p in instructional_periods
                if schedules[(year, section)][day].get(p) is None
            )
            for section in all_sections
        }
        precheck_faculty_daily_loads = _build_faculty_daily_loads(faculty_busy)
        precheck_subject_edge_counts = _build_subject_edge_counts(
            schedules,
            year,
            all_sections,
            instructional_periods,
        )
        for req in requirements:
            candidates = _enumerate_slot_candidates(
                req,
                req.hours,
                schedules,
                daily_subject_counts,
                faculty_busy,
                faculty_availability,
                faculty_section_slots,
                year,
                list(DAYS),
                instructional_periods,
                instructional_periods,
                sessions,
                session_adjacency,
                1,
                precheck_free_slots_tracker,
                precheck_faculty_daily_loads,
                precheck_subject_edge_counts,
            )
            if candidates:
                continue
            issues.append({
                "year": year,
                "sections": list(req.sections),
                "subject_id": req.subject_id,
                "faculty_id": req.faculty_id,
                "constraint": _infer_failure_reason(
                    req,
                    schedules,
                    daily_subject_counts,
                    faculty_busy,
                    faculty_availability,
                    faculty_section_slots,
                    year,
                    instructional_periods,
                    sessions,
                    session_adjacency,
                ),
                "detail": f"Subject {req.subject_id} has no feasible placement under the current hard constraints.",
            })

        section_summary = []
        for section in all_sections:
            free = sum(
                1 for day in DAYS for p in instructional_periods
                if schedules[(year, section)][day].get(p) is None
            )
            needed = section_required.get(section, 0)
            section_summary.append({
                "section": section,
                "requiredHours": needed,
                "freeSlots": free,
                "deficitHours": max(0, needed - free),
                "lockedSlots": max(0, total_instructional_slots - free),
            })

        feasible = len(issues) == 0
        blocking_sections = [
            section for section in section_summary if int(section.get("deficitHours", 0) or 0) > 0
        ]
        blocking_sections.sort(key=lambda item: (-int(item["deficitHours"]), item["section"]))
        section_summary.sort(key=lambda item: item["section"])
        return {
            "year": year,
            "feasible": feasible,
            "blockingSections": blocking_sections,
            "sectionSummary": section_summary,
            "issues": _group_issue_records(issues) if issues else [],
            "sections": all_sections,
            "requirementCount": len(requirements),
            "capacityPerSection": capacity_per_section,
        }

    # ── FULL SOLVER (only when precheck_only=False) ───────────────────────
    # Solve
    timeout_seconds = _compute_timeout_seconds(len(all_sections), len(requirements))
    retry_orders = [
        (list(DAYS), list(instructional_periods)),
        (list(DAYS), list(reversed(instructional_periods))),
        (list(reversed(DAYS)), list(instructional_periods)),
        (list(reversed(DAYS)), list(reversed(instructional_periods))),
        (list(DAYS[1:] + DAYS[:1]), list(instructional_periods)),
    ]
    strategy_orderings = [("shared-first", lambda item: _requirement_priority(item, faculty_availability, instructional_periods))]
    
    solved = False
    attempt = 0
    attempt_strategy_names = []
    
    while not solved and attempt < len(retry_orders):
        strategy_name, strategy_key = strategy_orderings[0]
        attempt_strategy_names.append(strategy_name)
        requirements.sort(key=strategy_key)
        days_order, periods_order = retry_orders[attempt]
        
        remaining_by_req = {i: req.hours for i, req in enumerate(requirements)}
        
        # Initialize free slots tracker
        free_slots_tracker = {}
        for section in all_sections:
            count = 0
            for d in DAYS:
                for p in instructional_periods:
                    if schedules[(year, section)][d].get(p) is None:
                        count += 1
            free_slots_tracker[section] = count
        faculty_daily_loads = _build_faculty_daily_loads(faculty_busy)
        subject_edge_counts = _build_subject_edge_counts(
            schedules,
            year,
            all_sections,
            instructional_periods,
        )

        def backtrack() -> bool:
            if _can_greedily_finish_remaining_requirements(requirements, remaining_by_req):
                return _finish_simple_section_requirements(
                    requirements,
                    remaining_by_req,
                    schedules,
                    daily_subject_counts,
                    faculty_busy,
                    faculty_availability,
                    faculty_section_slots,
                    year,
                    days_order,
                    periods_order,
                    instructional_periods,
                    sessions,
                    session_adjacency,
                    free_slots_tracker,
                    faculty_daily_loads,
                    subject_edge_counts,
                    subject_id_to_name,
                    faculty_id_to_name,
                    session_log,
                )
            next_req_idx, candidates = _select_next_requirement(
                requirements, remaining_by_req, schedules, daily_subject_counts, faculty_busy, faculty_availability,
                faculty_section_slots, year, days_order, periods_order, instructional_periods, sessions,
                session_adjacency, 8, free_slots_tracker, faculty_daily_loads, subject_edge_counts
            )
            if next_req_idx is None: return True
            if not candidates: return False
            
            req = requirements[next_req_idx]
            for cand in candidates:
                placements = _place_block(
                    req, cand.faculty_ids, cand.day, cand.start_period, cand.block_size,
                    schedules, faculty_busy, year, subject_id_to_name, faculty_id_to_name,
                    session_log, faculty_section_slots, free_slots_tracker, daily_subject_counts,
                    faculty_daily_loads, subject_edge_counts, instructional_periods, source="solver"
                )
                remaining_by_req[next_req_idx] -= cand.block_size
                if _section_capacity_is_feasible(
                    requirements,
                    remaining_by_req,
                    schedules,
                    year,
                    instructional_periods,
                ) and backtrack():
                    return True
                remaining_by_req[next_req_idx] += cand.block_size
                _undo_block(
                    req,
                    cand.faculty_ids,
                    placements,
                    schedules,
                    faculty_busy,
                    year,
                    session_log,
                    faculty_section_slots,
                    free_slots_tracker,
                    daily_subject_counts,
                    faculty_daily_loads,
                    subject_edge_counts,
                    instructional_periods,
                )
            return False

        if backtrack():
            solved = True
        attempt += 1

    if not solved:
        for i, req in enumerate(requirements):
            rem = remaining_by_req.get(i, 0)
            if rem > 0:
                reason = _infer_failure_reason(
                    req, schedules, daily_subject_counts, faculty_busy, faculty_availability,
                    faculty_section_slots, year, instructional_periods, sessions, session_adjacency
                )
                subject_id, subject_name = _resolve_subject_output(req.subject_id, subject_id_to_name)
                # Use faculty_options if available, otherwise fallback to faculty_id
                faculty_ids = req.faculty_options or ((req.faculty_id,) if req.faculty_id else ())
                faculty_id_display, _ = _resolve_faculty_display(faculty_ids, faculty_id_to_name)
                
                constraint_violations.append({
                    "year": year,
                    "sections": list(req.sections),
                    "subject_id": subject_id,
                    "faculty_id": faculty_id_display,
                    "constraint": reason,
                    "detail": f"Subject {subject_name} ({rem}h) could not be scheduled due to {reason}."
                })

    # Cleanup and Return
    _allocate_classrooms_to_schedule(
        year, all_sections, schedules, session_log, classrooms,
        constraint_violations, instructional_periods, sessions,
        prior_room_busy=prior_room_busy,
        lab_room_names=lab_room_names,
        room_capacity_map=room_capacity_map,
        section_strength_map=section_strength_map,
        fixed_classroom_blocks=fixed_classroom_blocks,
        section_home_room_map=section_home_room_map,
    )

    hard_limit_violations = _collect_subject_daily_hard_limit_violations(
        requirements,
        schedules,
        year,
        instructional_periods,
    )
    if hard_limit_violations:
        solved = False
        constraint_violations.extend(hard_limit_violations)

    quality_metrics, quality_warnings = _build_timetable_quality_metrics(
        requirements,
        schedules,
        year,
        instructional_periods,
        sessions,
    )
    if not classrooms:
        constraint_violations = [
            violation
            for violation in constraint_violations
            if violation.get("constraint") != "classroom allocation constraint"
        ]
    
    # Detect and append faculty workload conflicts
    workload_conflicts = _check_faculty_workload_conflicts(
        year,
        all_sections,
        schedules,
        instructional_periods,
        existing_workloads_payload,
        faculty_id_to_name,
    )
    constraint_violations.extend(workload_conflicts)
    
    all_grids = _serialize_section_grids(year, all_sections, schedules, instructional_periods)
    faculty_workloads = _build_faculty_workloads_from_sessions(session_log, instructional_periods)
    
    section_workbook = _build_section_timetables_workbook(year, all_sections, schedules, timetable_metadata, period_config)
    faculty_workbook = _build_faculty_workload_workbook(session_log, faculty_id_to_name, timetable_metadata, period_config)
    
    shared_classes = []
    for log_entry in session_log:
        if len(log_entry.get("sections", [])) > 1:
            shared_classes.append(log_entry)
            
    constraint_report_workbook = _build_constraint_report_workbook(constraint_violations, [], shared_classes)
    room_schedules = _build_room_schedule_map(schedules, classrooms)
    room_workbook = _build_room_timetables_workbook(classrooms, room_schedules, timetable_metadata, period_config)
    
    # Serialize room grids for the frontend JSON
    room_grids = {}
    for room, days in room_schedules.items():
        grid = {day: [None] * len(instructional_periods) for day in DAYS}
        for day in DAYS:
            for p_idx, p in enumerate(instructional_periods):
                cell = days.get(day, {}).get(p)
                if cell:
                    grid[day][p_idx] = cell
        room_grids[room] = grid
    
    generated_files = {
        "sectionTimetables": _encode_workbook("section_timetables.xlsx", section_workbook),
        "facultyWorkload": _encode_workbook("faculty_workload.xlsx", faculty_workbook),
        "constraintReport": _encode_workbook("constraint_report.xlsx", constraint_report_workbook),
        "sharedClassesReport": _encode_workbook("shared_classes_report.xlsx", constraint_report_workbook),
        "roomTimetables": _encode_workbook("room_timetables.xlsx", room_workbook),
    }
    
    timetable_id = store.next_timetable_id()
    selected_section = request_data.section if request_data.section in all_grids else all_sections[0]
    
    store.save_timetable(timetable_id, {
        "id": timetable_id,
        "year": year,
        "section": selected_section,
        "grid": all_grids.get(selected_section, {day: [None] * len(instructional_periods) for day in DAYS}),
        "allGrids": all_grids,
        "roomGrids": room_grids,
        "facultyWorkloads": faculty_workloads,
        "constraintViolations": constraint_violations,
        "qualityWarnings": _group_issue_records(quality_warnings),
        "unscheduledSubjects": [],
        "hasValidTimetable": solved,
        "hasConstraintViolations": not solved or bool(constraint_violations),
        "generatedFiles": generated_files,
        "timetableMetadata": timetable_metadata,
        "sharedClasses": shared_classes,
        "generationMeta": {
            "qualityMetrics": quality_metrics,
            "attemptStrategies": attempt_strategy_names,
            "retryStrategies": len(retry_orders),
            "timeoutSeconds": timeout_seconds,
        },
    })
    
    _persist_faculty_occupancy(store, timetable_id, session_log, instructional_periods)
    return {"timetableId": timetable_id, "message": "Timetable generated."}


def _build_room_schedule_map(
    schedules: dict[tuple[str, str], dict[str, dict[int, dict | None]]],
    classrooms: list[str]
) -> dict[str, dict[str, dict[int, dict | None]]]:
    all_rooms: list[str] = []
    seen_rooms: set[str] = set()

    for room in classrooms:
        normalized = str(room).strip()
        if not normalized or normalized in seen_rooms:
            continue
        seen_rooms.add(normalized)
        all_rooms.append(normalized)

    for _, grid in schedules.items():
        for day in DAYS:
            for p in PERIODS:
                cell = grid.get(day, {}).get(p)
                if not cell:
                    continue
                for candidate in (
                    cell.get("classroom"),
                    cell.get("labRoom"),
                    cell.get("venue"),
                    cell.get("fallbackLab"),
                ):
                    normalized = str(candidate or "").strip()
                    if not normalized or normalized in seen_rooms:
                        continue
                    seen_rooms.add(normalized)
                    all_rooms.append(normalized)

    room_schedules = {room: {d: {p: None for p in PERIODS} for d in DAYS} for room in all_rooms}
    
    def _store_room_entry(
        room: str,
        day: str,
        period: int,
        cell: dict,
        year: str,
        section: str,
    ) -> None:
        if room not in room_schedules:
            return
        cloned_cell = cell.copy()
        cloned_cell["year"] = year
        cloned_cell["section"] = section
        existing = room_schedules[room][day][period]
        if existing:
            existing_sections = [part.strip() for part in str(existing.get("section", "")).split(",") if part.strip()]
            if section not in existing_sections:
                cloned_cell["section"] = f"{existing.get('section', '')}, {section}".strip(", ")
            else:
                cloned_cell["section"] = str(existing.get("section", "")).strip()
        room_schedules[room][day][period] = cloned_cell

    for (year, section), grid in schedules.items():
        for day in DAYS:
            for p in PERIODS:
                cell = grid.get(day, {}).get(p)
                if not cell:
                    continue
                room_targets: list[str] = []
                for candidate in (
                    cell.get("fallbackLab"),
                    cell.get("labRoom"),
                    cell.get("venue"),
                    cell.get("classroom"),
                ):
                    room = str(candidate or "").strip()
                    if room and room not in room_targets:
                        room_targets.append(room)
                for room in room_targets:
                    _store_room_entry(room, day, p, cell, year, section)
    
    return room_schedules


def _build_room_timetables_workbook_from_schedule_map(
    room_schedules: dict[str, dict[str, dict[int, dict | None]]],
    timetable_metadata: dict[str, Any] | None = None,
    period_config: list[dict[str, str]] | None = None,
) -> Workbook:
    if period_config is None:
        period_config = DEFAULT_PERIOD_CONFIG
    workbook = Workbook()
    metadata = _resolve_timetable_metadata(timetable_metadata)
    if not room_schedules:
        worksheet = workbook.active
        worksheet.title = "Rooms"
        worksheet["A1"] = ACADEMIC_METADATA["college"]
        worksheet["A2"] = "(AUTONOMOUS)"
        worksheet["A3"] = ACADEMIC_METADATA["department"]
        worksheet["A4"] = f"ACADEMIC YEAR : {metadata['academicYear']} {metadata['semester']}"
        worksheet["A5"] = "ROOM TIMETABLES"
        worksheet["A6"] = "No classroom or lab inventory configured."
        return workbook
    workbook.remove(workbook.active)
    
    header1 = [str(row.get("period", "")).strip() for row in period_config]
    header2 = [str(row.get("time", "")).strip() for row in period_config]
    max_col = 1 + len(period_config)

    for room in sorted(room_schedules.keys()):
        worksheet = workbook.create_sheet(title=room[:31])
        worksheet.append([ACADEMIC_METADATA["college"]] + [""] * (max_col - 1))
        worksheet.append(["(AUTONOMOUS)"] + [""] * (max_col - 1))
        worksheet.append([ACADEMIC_METADATA["department"]] + [""] * (max_col - 1))
        worksheet.append([f"ACADEMIC YEAR : {metadata['academicYear']} {metadata['semester']}"] + [""] * (max_col - 1))
        worksheet.append([f"ROOM TIMETABLE: {room}"] + [""] * (max_col - 1))
        worksheet.append([""] * (max_col // 2) + [f"With effect from : {metadata['withEffectFromDisplay']}"] + [""] * (max_col // 2))
        worksheet.append(["DAY"] + header1)
        worksheet.append([""] + header2)

        day_start_row = 9
        section_schedule = room_schedules[room]
        break_overlap_rows: list[int] = []
        lunch_overlap_rows: list[int] = []
        
        break_cols = []
        lunch_cols = []
        for col_idx, row in enumerate(period_config, start=2):
            p = str(row.get("period", "")).strip().lower()
            if "break" in p:
                break_cols.append(col_idx)
            elif "lunch" in p:
                lunch_cols.append(col_idx)

        def _room_cell_text(entry: dict | None) -> str:
            if not entry:
                return ""
            subject = str(entry.get("subjectName") or entry.get("subject") or "").strip()
            year = str(entry.get("year", "")).strip()
            section = str(entry.get("section", "")).strip()
            return f"{subject}\n{year} {section}"

        for offset, day in enumerate(DAYS):
            row_idx = day_start_row + offset
            worksheet.cell(row=row_idx, column=1, value=DAY_SHORT_LABELS[day])
            
            display_columns = {}
            col = 2
            for entry in period_config:
                p = str(entry.get("period", "")).strip()
                if p.isdigit():
                    display_columns[int(p)] = col
                col += 1
            instructional_periods = sorted(display_columns.keys())
            
            idx = 0
            overlaps_break = False
            overlaps_lunch = False
            for c in break_cols:
                if worksheet.cell(row=row_idx, column=c).value:
                    overlaps_break = True
            for c in lunch_cols:
                if worksheet.cell(row=row_idx, column=c).value:
                    overlaps_lunch = True

            while idx < len(instructional_periods):
                p = instructional_periods[idx]
                entry = section_schedule[day].get(p)
                worksheet.cell(row=row_idx, column=display_columns[p], value=_room_cell_text(entry))
                
                end_idx = idx
                while end_idx + 1 < len(instructional_periods):
                    next_p = instructional_periods[end_idx + 1]
                    next_entry = section_schedule[day].get(next_p)
                    
                    is_adjacent = display_columns[next_p] == display_columns[instructional_periods[end_idx]] + 1
                    is_lab = isinstance(entry, dict) and entry.get("isLab")
                    is_same = (
                        _same_section_entry(entry, next_entry) 
                        and entry 
                        and next_entry
                        and entry.get("year") == next_entry.get("year") 
                        and entry.get("section") == next_entry.get("section")
                    )
                    
                    can_merge = False
                    if is_same:
                        if is_adjacent:
                            can_merge = True
                        elif is_lab:
                            mid_cols = range(display_columns[instructional_periods[end_idx]] + 1, display_columns[next_p])
                            if all(c in break_cols or c in lunch_cols for c in mid_cols):
                                can_merge = True
                    
                    if can_merge:
                        end_idx += 1
                    else:
                        break
                
                if end_idx > idx:
                    start_col = display_columns[p]
                    end_col = display_columns[instructional_periods[end_idx]]
                    for c in break_cols:
                        if start_col <= c <= end_col:
                            overlaps_break = True
                    for c in lunch_cols:
                        if start_col <= c <= end_col:
                            overlaps_lunch = True
                    worksheet.merge_cells(
                        start_row=row_idx, 
                        start_column=start_col, 
                        end_row=row_idx, 
                        end_column=end_col
                    )
                idx = end_idx + 1
                
            if overlaps_break: break_overlap_rows.append(row_idx)
            if overlaps_lunch: lunch_overlap_rows.append(row_idx)

        worksheet.append([""] * max_col)
        legend_separator_row = worksheet.max_row
        for _ in range(3):
            worksheet.append([""] * max_col)
        signature_separator_row = worksheet.max_row
        sig_row = signature_separator_row + 1
        worksheet.cell(row=sig_row, column=1, value="HEAD OF THE DEPARTMENT")
        worksheet.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=max_col // 2)
        worksheet.cell(row=sig_row, column=max_col // 2 + 1, value="PRINCIPAL")
        worksheet.merge_cells(start_row=sig_row, start_column=max_col // 2 + 1, end_row=sig_row, end_column=max_col)

        worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        worksheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_col)
        worksheet.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_col)
        worksheet.merge_cells(start_row=4, start_column=1, end_row=4, end_column=max_col)
        worksheet.merge_cells(start_row=5, start_column=1, end_row=5, end_column=max_col)
        worksheet.merge_cells(start_row=6, start_column=1, end_row=6, end_column=max_col // 2)
        worksheet.merge_cells(start_row=6, start_column=max_col // 2 + 1, end_row=6, end_column=max_col)
        worksheet.merge_cells(start_row=7, start_column=1, end_row=8, end_column=1)

        def merge_vertical_marker(column: int, label: str, blocked_rows: list[int]) -> None:
            all_rows = [day_start_row + i for i in range(len(DAYS))]
            segments: list[tuple[int, int]] = []
            segment_start = None
            for r in all_rows:
                if r in blocked_rows:
                    if segment_start is not None:
                        segments.append((segment_start, r - 1))
                        segment_start = None
                else:
                    if segment_start is None: segment_start = r
            if segment_start is not None: segments.append((segment_start, all_rows[-1]))

            for start, end in segments:
                worksheet.merge_cells(start_row=start, start_column=column, end_row=end, end_column=column)
                worksheet.cell(row=start, column=column, value=label)
                worksheet.cell(row=start, column=column).font = BOLD_FONT

        for bc in break_cols: merge_vertical_marker(bc, "BREAK", break_overlap_rows)
        for lc in lunch_cols: merge_vertical_marker(lc, "LUNCH", lunch_overlap_rows)

        _apply_formatted_sheet_layout(
            worksheet,
            legend_row_count=0,
            legend_start_row=legend_separator_row + 1,
            day_start_row=day_start_row,
            top_rows=6,
            max_col=max_col,
        )
    return workbook


def _build_room_timetables_workbook(
    rooms: list[str],
    room_schedules: dict[str, dict[str, dict[int, dict | None]]],
    timetable_metadata: dict[str, Any] | None = None,
    period_config: list[dict[str, str]] | None = None,
) -> Workbook:
    filtered = dict(room_schedules)
    return _build_room_timetables_workbook_from_schedule_map(filtered, timetable_metadata, period_config)

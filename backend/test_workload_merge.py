import sys
from pathlib import Path
ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

import unittest
from fastapi.testclient import TestClient
from io import BytesIO
import openpyxl
from main import app
from services.auth import get_current_user

# Override dependency to bypass auth during test
app.dependency_overrides[get_current_user] = lambda: {"username": "admin", "role": "admin"}

class TestWorkloadMerge(unittest.TestCase):
    def setUp(self):
        self.client = TestClient(app)

    def _create_mock_workbook(self, faculty_name_val, day_val, period_val, cell_val):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["COLLEGES NAME"])
        ws.append(["(AUTONOMOUS)"])
        ws.append(["DEPT"])
        ws.append(["ACADEMIC YEAR : 2024-2025 Even"])
        ws.append([f"FACULTY WORKLOAD : {faculty_name_val}"])
        ws.append(["Room No :", "", "", "", "", "With effect from : 18-03-2024"])
        ws.append(["DAY", "1", "2", "", "3", "4", "", "5", "6", "7"])
        ws.append(["", "9.10-10.00", "10.00-10.50", "10.50-11.00", "11.00-11.50", "11.50-12.40", "12.40-1.30", "1.30-2.20", "2.20-3.10", "3.10-4.00"])
        
        row = ["MON", "", "", "BREAK", "", "", "LUNCH", "", "", ""]
        if day_val == "MON":
            col_map = {1: 2, 2: 3, 3: 5, 4: 6, 5: 8, 6: 9, 7: 10}
            row[col_map[period_val] - 1] = cell_val
        ws.append(row)
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    def test_merge_no_conflict(self):
        wb1_bytes = self._create_mock_workbook("Ramesh (F001)", "MON", 1, "MATH (II CSE-A)\nRoom C101")
        wb2_bytes = self._create_mock_workbook("Ramesh (F001)", "MON", 2, "PHY (II CSE-A)\nRoom C101")
        
        files = [
            ("files", ("wb1.xlsx", wb1_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("wb2.xlsx", wb2_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ]
        
        response = self.client.post("/api/uploads/merge-workloads", files=files)
        self.assertEqual(200, response.status_code)
        
        import base64
        res_json = response.json()
        wb_bytes = base64.b64decode(res_json["facultyWorkloadWorkbook"]["contentBase64"])
        merged_wb = openpyxl.load_workbook(BytesIO(wb_bytes), data_only=True)
        self.assertIn("Ramesh", merged_wb.sheetnames)
        
        ws = merged_wb["Ramesh"]
        self.assertEqual("MATH (II CSE-A)\nRoom C101", ws.cell(row=9, column=2).value)
        self.assertEqual("PHY (II CSE-A)\nRoom C101", ws.cell(row=9, column=3).value)

    def test_merge_with_conflict(self):
        wb1_bytes = self._create_mock_workbook("Ramesh (F001)", "MON", 1, "MATH (II CSE-A)\nRoom C101")
        wb2_bytes = self._create_mock_workbook("Ramesh (F001)", "MON", 1, "PHY (II CSE-A)\nRoom C101")
        
        files = [
            ("files", ("wb1.xlsx", wb1_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("wb2.xlsx", wb2_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ]
        
        response = self.client.post("/api/uploads/merge-workloads", files=files)
        self.assertEqual(200, response.status_code)
        
        import base64
        res_json = response.json()
        wb_bytes = base64.b64decode(res_json["facultyWorkloadWorkbook"]["contentBase64"])
        merged_wb = openpyxl.load_workbook(BytesIO(wb_bytes), data_only=True)
        self.assertIn("Validation Report", merged_wb.sheetnames)
        ws_report = merged_wb["Validation Report"]
        self.assertEqual("Ramesh", ws_report.cell(row=4, column=1).value)
        self.assertEqual("F001", ws_report.cell(row=4, column=2).value)
        self.assertEqual("MON", ws_report.cell(row=4, column=3).value)
        self.assertEqual(1, ws_report.cell(row=4, column=4).value)

    def test_merge_stacked_workloads_on_single_sheet(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "StackedSheet"
        
        def append_workload(fac_name, day, period, val):
            ws.append(["COLLEGES NAME"])
            ws.append(["(AUTONOMOUS)"])
            ws.append(["DEPT"])
            ws.append(["ACADEMIC YEAR : 2024-2025 Even"])
            ws.append([f"FACULTY WORKLOAD : {fac_name}"])
            ws.append(["Room No :", "", "", "", "", "With effect from : 18-03-2024"])
            ws.append(["DAY", "1", "2", "", "3", "4", "", "5", "6", "7"])
            ws.append(["", "9.10-10.00", "10.00-10.50", "10.50-11.00", "11.00-11.50", "11.50-12.40", "12.40-1.30", "1.30-2.20", "2.20-3.10", "3.10-4.00"])
            
            row = ["MON", "", "", "BREAK", "", "", "LUNCH", "", "", ""]
            if day == "MON":
                col_map = {1: 2, 2: 3, 3: 5, 4: 6, 5: 8, 6: 9, 7: 10}
                row[col_map[period] - 1] = val
            ws.append(row)
            
            for _ in range(5):
                ws.append([])
                
        append_workload("Smith (F001)", "MON", 1, "MATH (II CSE-A)\nRoom C101")
        append_workload("Jones (F002)", "MON", 2, "PHY (II CSE-A)\nRoom C101")
        
        output = BytesIO()
        wb.save(output)
        wb_bytes = output.getvalue()
        
        files = [
            ("files", ("stacked.xlsx", wb_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
            ("files", ("dummy.xlsx", self._create_mock_workbook("Smith (F001)", "MON", 3, "OS (II CSE-A)\nRoom C102"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")),
        ]
        
        response = self.client.post("/api/uploads/merge-workloads", files=files)
        self.assertEqual(200, response.status_code)
        
        import base64
        res_json = response.json()
        out_bytes = base64.b64decode(res_json["facultyWorkloadWorkbook"]["contentBase64"])
        merged_wb = openpyxl.load_workbook(BytesIO(out_bytes), data_only=True)
        
        self.assertIn("Smith", merged_wb.sheetnames)
        self.assertIn("Jones", merged_wb.sheetnames)
        
        ws_smith = merged_wb["Smith"]
        self.assertEqual("MATH (II CSE-A)\nRoom C101", ws_smith.cell(row=9, column=2).value)
        self.assertEqual("OS (II CSE-A)\nRoom C102", ws_smith.cell(row=9, column=5).value)
        
        ws_jones = merged_wb["Jones"]
        self.assertEqual("PHY (II CSE-A)\nRoom C101", ws_jones.cell(row=9, column=3).value)

if __name__ == "__main__":
    unittest.main()

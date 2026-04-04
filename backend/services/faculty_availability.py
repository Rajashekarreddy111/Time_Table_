import base64
from datetime import datetime, time
from io import BytesIO
import logging
import re

from fastapi import HTTPException
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side

from services.utils import normalize_year
from storage.memory_store import MemoryStore

logger = logging.getLogger(__name__)

DAYS = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
PERIOD_TIME = {
    1: "9:10 - 10:00",
    2: "10:00 - 10:50",
    3: "11:00 - 11:50",
    4: "11:50 - 12:40",
    5: "1:30 - 2:20",
    6: "2:20 - 3:10",
    7: "3:10 - 4:00",
}
PERIOD_WINDOWS = {
    1: (9 * 60 + 10, 10 * 60),
    2: (10 * 60, 10 * 60 + 50),
    3: (11 * 60, 11 * 60 + 50),
    4: (11 * 60 + 50, 12 * 60 + 40),
    5: (13 * 60 + 30, 14 * 60 + 20),
    6: (14 * 60 + 20, 15 * 60 + 10),
    7: (15 * 60 + 10, 16 * 60),
}
THIN_SIDE = Side(style="thin", color="000000")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _validation_error(message: str, details: list | None = None) -> HTTPException:
    return HTTPException(
        status_code=400,
        detail={"error": "ValidationError", "message": message, "details": details or []},
    )


def _to_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_day(value) -> str:
    raw = _to_text(value)
    if not raw:
        return ""
    compact = re.sub(r"[^A-Z]+", "", raw.upper())
    day_map = {
        "MON": "Monday",
        "MONDAY": "Monday",
        "TUE": "Tuesday",
        "TUESDAY": "Tuesday",
        "WED": "Wednesday",
        "WEDNESDAY": "Wednesday",
        "THU": "Thursday",
        "THURSDAY": "Thursday",
        "FRI": "Friday",
        "FRIDAY": "Friday",
        "SAT": "Saturday",
        "SATURDAY": "Saturday",
    }
    return day_map.get(compact, raw.capitalize())


def _normalize_token(value: str) -> str:
    return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())


def _build_section_candidates(full_year: str, section: str) -> set[str]:
    section_value = str(section or "").strip()
    if not section_value:
        return set()

    raw_year = str(full_year or "").strip()
    year_number = "".join(ch for ch in raw_year if ch.isdigit())
    compact = f"{year_number}{section_value}" if year_number else section_value
    section_without_year_suffix = section_value
    if year_number:
        suffix_pattern = re.compile(rf"[-_/ ]?{re.escape(year_number)}$", re.IGNORECASE)
        section_without_year_suffix = suffix_pattern.sub("", section_value).strip()

    candidates = {
        _normalize_token(section_value),
        _normalize_token(section_without_year_suffix),
        _normalize_token(f"{raw_year}|{section_value}"),
        _normalize_token(f"{raw_year}|{section_without_year_suffix}"),
        _normalize_token(f"{raw_year} {section_value}"),
        _normalize_token(f"{raw_year} {section_without_year_suffix}"),
        _normalize_token(compact),
        _normalize_token(f"{year_number}{section_without_year_suffix}") if year_number else "",
    }
    return {candidate for candidate in candidates if candidate}


def _normalize_ignored_sections(ignored_sections: list[str]) -> set[str]:
    normalized: set[str] = set()
    for item in ignored_sections:
        text = str(item or "").strip()
        if not text:
            continue
        normalized.add(_normalize_token(text))
        if "|" in text:
            _, _, tail = text.partition("|")
            if tail.strip():
                normalized.add(_normalize_token(tail))
        else:
            normalized_year = normalize_year(text)
            if normalized_year != text:
                normalized.add(_normalize_token(normalized_year))
    return normalized


def _is_ignored(class_info: dict, ignored_years: list[str], ignored_sections: list[str]) -> bool:
    raw_year = str(class_info.get("year", "")).strip()
    if not raw_year:
        return False
    full_year = normalize_year(raw_year)
    if full_year in ignored_years:
        return True
    section = str(class_info.get("section", "")).strip()
    if not section:
        return False
    normalized_ignored = _normalize_ignored_sections(ignored_sections)
    for candidate in _build_section_candidates(full_year, section):
        if candidate in normalized_ignored:
            return True
    return False


def _availability_result(
    day_name: str,
    selected_periods: list[int],
    selected_faculty: list[str],
    available_faculty: list[str],
    faculty_required: int,
    available_count: int,
    *,
    start_time=None,
    end_time=None,
) -> dict:
    safe_required = max(1, int(faculty_required or 1))
    sufficient = available_count >= safe_required
    shortage = max(0, safe_required - available_count)
    if sufficient:
        message = (
            f"Sufficient faculty available. Selected {len(selected_faculty)} "
            f"out of {available_count} available."
        )
    elif available_count > 0:
        message = (
            f"Only {available_count} faculty available; {safe_required} requested. "
            "No sufficient faculty available."
        )
    else:
        message = (
            f"No faculty available for the selected slot(s); {safe_required} requested."
        )

    return {
        "day": day_name,
        "periods": [{"period": period, "time": PERIOD_TIME[period]} for period in selected_periods],
        "startTime": _to_text(start_time),
        "endTime": _to_text(end_time),
        "faculty": selected_faculty,
        "availableFaculty": available_faculty,
        "availableFacultyCount": available_count,
        "sufficientFaculty": sufficient,
        "shortageCount": shortage,
        "message": message,
    }


def _build_faculty_name_map(store: MemoryStore, faculty_id_map_file_id: str | None) -> dict[str, str]:
    if not faculty_id_map_file_id:
        payload = store.get_scoped_mapping("faculty_id_map", "global")
        if not payload:
            return {}
    else:
        payload = store.get_file_map(faculty_id_map_file_id)
        if not payload:
            raise _validation_error("Invalid facultyIdMapFileId", [])

    name_map: dict[str, str] = {}
    for row in payload.get("rows", []):
        faculty_id = _to_text(row.get("faculty_id"))
        faculty_name = _to_text(row.get("faculty_name"))
        if faculty_id and faculty_name:
            name_map[faculty_id] = faculty_name
    return name_map


def _get_availability_payload(store: MemoryStore, availability_file_id: str | None) -> dict:
    if availability_file_id:
        payload = store.get_file_map(availability_file_id)
        if not payload:
            raise _validation_error("Invalid availabilityFileId", [])
        return payload

    payload = store.get_scoped_mapping("faculty_availability", "global")
    if not payload:
        raise _validation_error("Invalid availabilityFileId", [])
    return payload


def _get_known_faculty_names(payload: dict, schedules: dict[str, dict[str, dict[int, dict]]]) -> list[str]:
    known_names: set[str] = set()
    for value in payload.get("facultyNames", []):
        text = _to_text(value)
        if text:
            known_names.add(text)
    known_names.update(schedules.keys())
    return sorted(known_names, key=lambda faculty: (faculty.lower(), faculty))


def _parse_clock_time(value) -> int | None:
    if isinstance(value, datetime):
        return value.hour * 60 + value.minute
    if isinstance(value, time):
        return value.hour * 60 + value.minute
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        numeric = float(value)
        if 0 <= numeric < 1:
            total_minutes = round(numeric * 24 * 60)
            return int(total_minutes % (24 * 60))

    text = _to_text(value)
    if not text:
        return None

    normalized = re.sub(r"\s+", "", text).upper().replace(".", ":")
    has_meridiem = normalized.endswith("AM") or normalized.endswith("PM")
    for fmt in ("%H:%M", "%H:%M:%S", "%I:%M%p", "%I:%M:%S%p", "%I%p"):
        try:
            parsed = datetime.strptime(normalized, fmt)
            minutes = parsed.hour * 60 + parsed.minute
            if not has_meridiem and fmt == "%H:%M":
                hour = parsed.hour
                if 1 <= hour <= 4:
                    minutes += 12 * 60
            elif not has_meridiem and fmt == "%H:%M:%S":
                hour = parsed.hour
                if 1 <= hour <= 4:
                    minutes += 12 * 60
            return minutes
        except ValueError:
            continue

    timestamp_match = re.search(r"(\d{1,2}):(\d{2})(?::(\d{2}))?$", text)
    if timestamp_match:
        hour = int(timestamp_match.group(1))
        minute = int(timestamp_match.group(2))
        if 1 <= hour <= 4 and "AM" not in normalized and "PM" not in normalized:
            hour += 12
        return hour * 60 + minute
    return None


def _resolve_periods_from_time_range(start_time, end_time) -> list[int]:
    start_minutes = _parse_clock_time(start_time)
    end_minutes = _parse_clock_time(end_time)
    if start_minutes is None or end_minutes is None:
        return []
    if end_minutes < start_minutes:
        raise _validation_error("End time must be greater than or equal to start time", [])

    periods = [
        period
        for period, (period_start, period_end) in PERIOD_WINDOWS.items()
        if period_start < end_minutes and period_end > start_minutes
    ]
    logger.debug(
        "Resolved time range to periods",
        extra={
            "start_time": _to_text(start_time),
            "end_time": _to_text(end_time),
            "start_minutes": start_minutes,
            "end_minutes": end_minutes,
            "selected_periods": periods,
        },
    )
    return periods


def _resolve_selected_periods(
    periods: list[int] | None,
    start_time=None,
    end_time=None,
) -> list[int]:
    raw_periods = list(periods or [])
    if not raw_periods and start_time and end_time:
        raw_periods = _resolve_periods_from_time_range(start_time, end_time)

    selected_periods = sorted({period for period in raw_periods if period in PERIOD_TIME})
    if selected_periods:
        return selected_periods

    if start_time or end_time:
        if not start_time or not end_time:
            raise _validation_error("Both start time and end time are required when periods are not provided", [])
        raise _validation_error(
            "No timetable periods overlap the provided start and end time",
            [{
                "startTime": _to_text(start_time),
                "endTime": _to_text(end_time),
            }],
        )

    raise _validation_error("At least one period must be selected", [])


def _fair_select_faculty(
    available_faculty: set[str],
    faculty_required: int,
    selection_counts: dict[str, int] | None = None,
) -> list[str]:
    safe_required = max(1, int(faculty_required or 1))
    ranked = sorted(
        available_faculty,
        key=lambda faculty: (selection_counts.get(faculty, 0), faculty.lower(), faculty)
        if selection_counts is not None
        else (0, faculty.lower(), faculty),
    )
    chosen = ranked[: min(safe_required, len(ranked))]
    if selection_counts is not None:
        for faculty in chosen:
            selection_counts[faculty] = selection_counts.get(faculty, 0) + 1
    return chosen


def _format_day_schedule(day_schedule: dict[int, dict]) -> list[dict]:
    return [
        {
            "period": period,
            "time": PERIOD_TIME.get(period, ""),
            "is_available": bool(info.get("is_available")),
            "year": _to_text(info.get("year")),
            "section": _to_text(info.get("section")),
            "subject": _to_text(info.get("subject")),
        }
        for period, info in sorted(day_schedule.items())
    ]


def _warn_for_missing_overlap_rows(
    faculty: str,
    day_name: str,
    selected_periods: list[int],
    day_mode: str | None,
    day_schedule: dict[int, dict],
) -> None:
    if day_mode != "occupancy":
        return
    missing = [period for period in selected_periods if period not in day_schedule]
    if not missing:
        return
    logger.warning(
        "Occupancy dataset is missing overlapping period rows; missing periods will be treated as free",
        extra={
            "faculty": faculty,
            "day_name": day_name,
            "selected_periods": selected_periods,
            "missing_periods": missing,
            "day_schedule": _format_day_schedule(day_schedule),
        },
    )


def _build_schedules_from_upload(
    store: MemoryStore,
    availability_file_id: str,
    faculty_name_map: dict[str, str],
) -> tuple[
    dict[str, dict[str, dict[int, dict]]],
    dict[str, set[str]],
    dict[str, dict[str, str]],
]:
    payload = _get_availability_payload(store, availability_file_id)

    schedules: dict[str, dict[str, dict[int, dict]]] = {}
    explicit_free_days: dict[str, set[str]] = {}
    schedule_modes: dict[str, dict[str, str]] = {}

    for row in payload.get("rows", []):
        day = _normalize_day(row.get("day"))
        try:
            period = int(float(_to_text(row.get("period"))))
        except ValueError:
            continue
        if day not in DAYS or period not in PERIOD_TIME:
            continue

        faculty_id = _to_text(row.get("faculty_id"))
        faculty_name = _to_text(row.get("faculty_name"))
        faculty_key = faculty_name or faculty_name_map.get(faculty_id) or faculty_id
        if not faculty_key:
            continue

        class_info = {
            "year": _to_text(row.get("year")),
            "section": _to_text(row.get("section")),
            "subject": _to_text(row.get("subject")),
            "is_available": bool(row.get("is_available")),
        }
        schedules.setdefault(faculty_key, {}).setdefault(day, {})[period] = class_info

        faculty_day_modes = schedule_modes.setdefault(faculty_key, {})
        if class_info["is_available"]:
            previous_mode = faculty_day_modes.get(day)
            if previous_mode == "occupancy":
                raise _validation_error(
                    "Mixed faculty availability dataset detected",
                    [{
                        "faculty": faculty_key,
                        "day": day,
                        "detail": "The same faculty/day contains both free-slot rows and busy-slot rows.",
                    }],
                )
            explicit_free_days.setdefault(faculty_key, set()).add(day)
            faculty_day_modes[day] = "availability"
        else:
            previous_mode = faculty_day_modes.get(day)
            if previous_mode == "availability":
                raise _validation_error(
                    "Mixed faculty availability dataset detected",
                    [{
                        "faculty": faculty_key,
                        "day": day,
                        "detail": "The same faculty/day contains both free-slot rows and busy-slot rows.",
                    }],
                )
            faculty_day_modes.setdefault(day, "occupancy")

    if not schedules:
        raise _validation_error("Uploaded faculty availability file has no valid schedule rows", [])
    return schedules, explicit_free_days, schedule_modes


def _is_faculty_free_for_period(
    faculty: str,
    day_name: str,
    period: int,
    schedules: dict[str, dict[str, dict[int, dict]]],
    explicit_free_days: dict[str, set[str]],
    schedule_modes: dict[str, dict[str, str]],
    ignored_years: list[str],
    ignored_sections: list[str],
) -> bool:
    day_schedule = schedules.get(faculty, {}).get(day_name, {})
    class_info = day_schedule.get(period)
    day_mode = schedule_modes.get(faculty, {}).get(day_name)

    logger.debug(
        "Checking faculty availability for period",
        extra={
            "faculty": faculty,
            "day_name": day_name,
            "period": period,
            "day_mode": day_mode,
            "class_info": class_info,
            "day_schedule": _format_day_schedule(day_schedule),
        },
    )

    if day_mode is None:
        faculty_modes = schedule_modes.get(faculty, {})
        if not faculty_modes:
            return False
        if any(mode == "availability" for mode in faculty_modes.values()):
            return False
        return True

    if day_mode == "availability" or day_name in explicit_free_days.get(faculty, set()):
        if class_info is None:
            return False
        return bool(class_info.get("is_available")) or _is_ignored(class_info, ignored_years, ignored_sections)

    if class_info is None:
        return True
    return bool(class_info.get("is_available")) or _is_ignored(class_info, ignored_years, ignored_sections)


def _faculty_is_free_for_all_selected_periods(
    faculty: str,
    day_name: str,
    selected_periods: list[int],
    schedules: dict[str, dict[str, dict[int, dict]]],
    explicit_free_days: dict[str, set[str]],
    schedule_modes: dict[str, dict[str, str]],
    ignored_years: list[str],
    ignored_sections: list[str],
) -> bool:
    day_schedule = schedules.get(faculty, {}).get(day_name, {})
    day_mode = schedule_modes.get(faculty, {}).get(day_name)
    _warn_for_missing_overlap_rows(faculty, day_name, selected_periods, day_mode, day_schedule)

    for period in selected_periods:
        if not _is_faculty_free_for_period(
            faculty,
            day_name,
            period,
            schedules,
            explicit_free_days,
            schedule_modes,
            ignored_years,
            ignored_sections,
        ):
            return False
    return True


def get_available_faculty_for_all_periods(
    store: MemoryStore,
    date_value: str,
    periods: list[int],
    start_time,
    end_time,
    faculty_required: int,
    ignored_years: list[str],
    ignored_sections: list[str],
    availability_file_id: str | None,
    faculty_id_map_file_id: str | None,
) -> dict:
    if not availability_file_id:
        raise _validation_error("availabilityFileId is required", [])

    try:
        day_name = datetime.strptime(date_value, "%Y-%m-%d").strftime("%A")
    except ValueError as exc:
        raise _validation_error("Invalid date format; expected YYYY-MM-DD", []) from exc

    if day_name not in DAYS:
        raise _validation_error("Selected date is Sunday; no classes scheduled", [])

    selected_periods = _resolve_selected_periods(periods, start_time, end_time)

    logger.debug(
        "Processing single faculty availability request",
        extra={
            "date_value": date_value,
            "day_name": day_name,
            "start_time": _to_text(start_time),
            "end_time": _to_text(end_time),
            "selected_periods": selected_periods,
            "faculty_required": faculty_required,
        },
    )

    faculty_name_map = _build_faculty_name_map(store, faculty_id_map_file_id)
    availability_payload = _get_availability_payload(store, availability_file_id)
    schedules, explicit_free_days, schedule_modes = _build_schedules_from_upload(
        store,
        availability_file_id,
        faculty_name_map,
    )
    faculty_names = _get_known_faculty_names(availability_payload, schedules)
    common_available: set[str] = set()

    for faculty in faculty_names:
        if _faculty_is_free_for_all_selected_periods(
            faculty,
            day_name,
            selected_periods,
            schedules,
            explicit_free_days,
            schedule_modes,
            ignored_years,
            ignored_sections,
        ):
            common_available.add(faculty)

    selected_faculty = _fair_select_faculty(common_available, faculty_required)
    return _availability_result(
        day_name=day_name,
        selected_periods=selected_periods,
        selected_faculty=selected_faculty,
        available_faculty=sorted(common_available, key=lambda faculty: (faculty.lower(), faculty)),
        faculty_required=faculty_required,
        available_count=len(common_available),
        start_time=start_time,
        end_time=end_time,
    )


def get_bulk_available_faculty(
    store: MemoryStore,
    availability_file_id: str,
    query_file_id: str,
    ignored_years: list[str],
    ignored_sections: list[str],
    faculty_id_map_file_id: str | None,
) -> dict:
    if not availability_file_id:
        raise _validation_error("availabilityFileId is required", [])
    if not query_file_id:
        raise _validation_error("queryFileId is required", [])

    query_payload = store.get_file_map(query_file_id)
    if not query_payload:
        raise _validation_error("Invalid queryFileId", [])

    query_rows = query_payload.get("rows", [])
    if not query_rows:
        raise _validation_error("Query file is empty", [])

    faculty_name_map = _build_faculty_name_map(store, faculty_id_map_file_id)
    availability_payload = _get_availability_payload(store, availability_file_id)
    schedules, explicit_free_days, schedule_modes = _build_schedules_from_upload(
        store,
        availability_file_id,
        faculty_name_map,
    )
    faculty_names = _get_known_faculty_names(availability_payload, schedules)
    results = []
    selection_counts: dict[str, int] = {}

    for row in query_rows:
        date_value = row.get("date")
        faculty_required = row.get("facultyRequired", 1)
        periods = row.get("periods", [])
        start_time = row.get("startTime")
        end_time = row.get("endTime")

        try:
            day_name = datetime.strptime(date_value, "%Y-%m-%d").strftime("%A")
        except ValueError:
            try:
                from dateutil import parser

                day_name = parser.parse(date_value).strftime("%A")
            except Exception:
                continue

        if day_name not in DAYS:
            continue

        selected_periods = _resolve_selected_periods(periods, start_time, end_time)
        logger.debug(
            "Processing bulk faculty availability row",
            extra={
                "date_value": date_value,
                "day_name": day_name,
                "start_time": start_time,
                "end_time": end_time,
                "selected_periods": selected_periods,
                "faculty_required": faculty_required,
            },
        )

        common_available: set[str] = set()
        for faculty in faculty_names:
            if _faculty_is_free_for_all_selected_periods(
                faculty,
                day_name,
                selected_periods,
                schedules,
                explicit_free_days,
                schedule_modes,
                ignored_years,
                ignored_sections,
            ):
                common_available.add(faculty)

        selected_faculty = _fair_select_faculty(common_available, faculty_required, selection_counts)
        results.append(
            {
                "date": date_value,
                "facultyRequired": faculty_required,
                **_availability_result(
                    day_name=day_name,
                    selected_periods=selected_periods,
                    selected_faculty=selected_faculty,
                    available_faculty=sorted(common_available, key=lambda faculty: (faculty.lower(), faculty)),
                    faculty_required=faculty_required,
                    available_count=len(common_available),
                    start_time=start_time,
                    end_time=end_time,
                ),
            }
        )

    return {"results": results}


def _format_export_date(value: str) -> str:
    text = _to_text(value)
    match = re.match(r"^(\d{4})-(\d{2})-(\d{2})", text)
    if match:
        return f"{match.group(3)}-{match.group(2)}-{match.group(1)}"
    return text


def _build_export_slot(item: dict) -> dict:
    start_text = _to_text(item.get("startTime"))
    end_text = _to_text(item.get("endTime"))
    if start_text and end_text:
        label = f"{start_text} - {end_text}"
        return {"key": f"{start_text}__{end_text}", "label": label, "start_minutes": _parse_clock_time(start_text) or 0}

    period_labels = [str(period.get("time", "")).strip() for period in item.get("periods", []) if str(period.get("time", "")).strip()]
    if not period_labels:
        return {"key": "Session", "label": "Session", "start_minutes": 0}
    first_start = period_labels[0].split("-", 1)[0].strip()
    last_end = period_labels[-1].rsplit("-", 1)[-1].strip()
    return {
        "key": f"{first_start}__{last_end}",
        "label": f"{first_start} - {last_end}",
        "start_minutes": _parse_clock_time(first_start) or 0,
    }


def _slot_meridiem(slot: dict) -> str:
    return "AM" if int(slot.get("start_minutes", 0)) < 12 * 60 else "PM"


def _build_export_groups(items: list[dict]) -> list[dict]:
    slots_by_date: dict[str, list[dict]] = {}
    for item in items:
        slot = _build_export_slot(item)
        date_key = _to_text(item.get("date"))
        current = slots_by_date.setdefault(date_key, [])
        if not any(existing.get("key") == slot["key"] for existing in current):
            current.append(slot)

    groups: list[dict] = []
    for date_key in sorted(slots_by_date):
        ordered = sorted(slots_by_date[date_key], key=lambda slot: (int(slot.get("start_minutes", 0)), str(slot.get("label", ""))))
        uses_meridiem = len(ordered) <= 2
        if uses_meridiem:
            slots = [
                next((slot for slot in ordered if _slot_meridiem(slot) == "AM"), {"key": "__am__", "label": "AM", "start_minutes": 0}),
                next((slot for slot in ordered if _slot_meridiem(slot) == "PM"), {"key": "__pm__", "label": "PM", "start_minutes": 12 * 60}),
            ]
        else:
            slots = ordered
        groups.append(
            {
                "date": date_key,
                "header_date": _format_export_date(date_key),
                "slots": slots,
                "uses_meridiem": uses_meridiem,
            }
        )
    return groups


def _build_timing_summary(groups: list[dict]) -> str:
    labels: list[str] = []
    for group in groups:
        for slot in group.get("slots", []):
            label = str(slot.get("label", "")).strip()
            if group.get("uses_meridiem"):
                meridiem = _slot_meridiem(slot)
                if meridiem not in {"AM", "PM"}:
                    continue
            if label and label not in labels and not (group.get("uses_meridiem") and label in {"AM", "PM"}):
                labels.append(label)
    return " & ".join(f"({label})" for label in labels)


def _workbook_bytes(workbook: Workbook) -> bytes:
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _encode_workbook(file_name: str, workbook: Workbook) -> dict:
    return {
        "fileName": file_name,
        "contentType": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "contentBase64": base64.b64encode(_workbook_bytes(workbook)).decode("ascii"),
    }


def build_bulk_faculty_availability_workbook(results: list[dict], *, mode: str) -> dict:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Fair Selection" if mode == "selected" else "All Available"

    groups = _build_export_groups(results)
    faculty_getter = (lambda item: item.get("faculty", [])) if mode == "selected" else (lambda item: item.get("availableFaculty", []))
    faculty_names = sorted(
        {
            _to_text(name)
            for item in results
            for name in faculty_getter(item)
            if _to_text(name)
        },
        key=lambda name: (name.lower(), name),
    )

    total_columns = 3 + sum(len(group["slots"]) for group in groups)
    rows: list[list[object]] = []

    def pad_row(values: list[object]) -> list[object]:
        return values + [""] * max(0, total_columns - len(values))

    department_short_name = "CSE"
    rows.append(pad_row([department_short_name, "", "", _build_timing_summary(groups)]))
    rows.append(pad_row(["S.No", "Faculty Name", "Total"]))
    rows.append(pad_row(["", "", ""]))

    column_cursor = 4
    slot_column_map: dict[str, int] = {}
    for group in groups:
        rows[1][column_cursor - 1] = group["header_date"]
        for slot_index, slot in enumerate(group["slots"]):
            rows[2][column_cursor + slot_index - 1] = ("AM" if slot_index == 0 else "PM") if group["uses_meridiem"] else slot["label"]
            map_key = f"{group['date']}__{('AM' if slot_index == 0 else 'PM') if group['uses_meridiem'] else slot['key']}"
            slot_column_map[map_key] = column_cursor + slot_index
        column_cursor += len(group["slots"])

    faculty_totals: dict[str, int] = {}
    for item in results:
        for name in faculty_getter(item):
            faculty_name = _to_text(name)
            if faculty_name:
                faculty_totals[faculty_name] = faculty_totals.get(faculty_name, 0) + 1

    faculty_row_map: dict[str, int] = {}
    for index, faculty_name in enumerate(faculty_names, start=1):
        rows.append(pad_row([index, faculty_name, faculty_totals.get(faculty_name, 0)]))
        faculty_row_map[faculty_name] = len(rows)

    slot_totals: dict[int, int] = {}
    for item in results:
        slot = _build_export_slot(item)
        group = next((entry for entry in groups if entry["date"] == _to_text(item.get("date"))), None)
        if not group:
            continue
        lookup_key = f"{group['date']}__{_slot_meridiem(slot) if group['uses_meridiem'] else slot['key']}"
        column_index = slot_column_map.get(lookup_key)
        if not column_index:
            continue
        current_names = [_to_text(name) for name in faculty_getter(item) if _to_text(name)]
        slot_totals[column_index] = len(current_names)
        for faculty_name in current_names:
            row_index = faculty_row_map.get(faculty_name)
            if row_index:
                rows[row_index - 1][column_index - 1] = "X"

    total_row = pad_row(["", "Total", sum(slot_totals.values())])
    for column_index, total in slot_totals.items():
        total_row[column_index - 1] = total
    rows.append(total_row)

    for row in rows:
        worksheet.append(row)

    column_cursor = 4
    for group in groups:
        worksheet.merge_cells(start_row=2, start_column=column_cursor, end_row=2, end_column=column_cursor + len(group["slots"]) - 1)
        column_cursor += len(group["slots"])
    worksheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=3, end_column=2)
    worksheet.merge_cells(start_row=1, start_column=3, end_row=3, end_column=3)
    if total_columns > 3:
        worksheet.merge_cells(start_row=1, start_column=4, end_row=1, end_column=total_columns)

    widths = [8, 30, 12] + [
        max(12, len(("AM" if group["uses_meridiem"] and idx == 0 else "PM" if group["uses_meridiem"] else slot["label"])) + 2)
        for group in groups
        for idx, slot in enumerate(group["slots"])
    ]
    for idx, width in enumerate(widths, start=1):
        worksheet.column_dimensions[worksheet.cell(row=1, column=idx).column_letter].width = width

    for row_idx in range(1, worksheet.max_row + 1):
        worksheet.row_dimensions[row_idx].height = 22
        for col_idx in range(1, total_columns + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            cell.alignment = CENTER_ALIGNMENT
            cell.border = THIN_BORDER
            if row_idx <= 3 or row_idx == worksheet.max_row:
                cell.font = Font(bold=True)

    file_name = "invisilation_fair_selection.xlsx" if mode == "selected" else "invisilation_all_available.xlsx"
    return _encode_workbook(file_name, workbook)

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
import re

import pandas as pd
from openpyxl import Workbook


DAYS: list[str] = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
PERIODS: list[int] = [1, 2, 3, 4, 5, 6, 7]
DAY_INDEX: dict[int, str] = {i + 1: d for i, d in enumerate(DAYS)}


class MainConfigValidationError(Exception):
    def __init__(self, year: str, section: str, actual_hours: int) -> None:
        super().__init__(f"Year={year} Section={section} actual_hours={actual_hours} expected_hours=42")
        self.year = year
        self.section = section
        self.actual_hours = actual_hours


def _norm_col(s: Any) -> str:
    text = str(s).strip().lower()
    text = text.replace("-", "_").replace(" ", "_")
    text = re.sub(r"_+", "_", text)
    return text


def _to_str(x: Any) -> str:
    if x is None:
        return ""
    return str(x).strip()


def _to_int(x: Any) -> int | None:
    if x is None:
        return None
    try:
        # Handles numeric-like cells (e.g. 1.0 in Excel)
        return int(float(str(x).strip()))
    except Exception:
        return None


def _parse_periods(value: Any) -> list[int]:
    """
    Accepts:
    - "1,2,3"
    - 1 (single)
    - "1 2 3"
    - lists/tuples of numeric-ish
    """
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        out: list[int] = []
        for v in value:
            iv = _to_int(v)
            if iv is not None:
                out.append(iv)
        return out
    text = _to_str(value)
    if not text:
        return []
    # Extract integers (robust against "P2" / "2")
    tokens: list[str] = []
    cur = ""
    for ch in text:
        if ch.isdigit() or (ch == "-" and not cur):
            cur += ch
        else:
            if cur:
                tokens.append(cur)
                cur = ""
    if cur:
        tokens.append(cur)
    out = []
    for t in tokens:
        iv = _to_int(t)
        if iv is not None:
            out.append(iv)
    return out


def _parse_day(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, (int, float)) and not pd.isna(value):
        dv = _to_int(value)
        if dv is None:
            return None
        return DAY_INDEX.get(dv)
    text = _to_str(value)
    if not text:
        return None
    # "Mon", "Monday", "1" etc.
    if text.isdigit():
        dv = _to_int(text)
        return DAY_INDEX.get(dv) if dv else None
    lowered = text.lower()
    for d in DAYS:
        if d.lower().startswith(lowered[:3]):
            return d
    return None


def _load_table(path: str | Path) -> pd.DataFrame:
    """
    Load a spreadsheet-like table.
    Uses the repo's flexible upload parser so grouped main-timetable templates work.
    """
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(str(p))
    suffix = p.suffix.lower()
    if suffix not in {".csv", ".xlsx", ".xls"}:
        raise ValueError(f"Unsupported file extension: {suffix}")

    # Reuse the existing repo parser for grouped main timetable Excel.
    from services.file_parser import parse_tabular_upload

    file_bytes = p.read_bytes()
    return parse_tabular_upload(p.name, file_bytes)


def _read_main_config(main_config_path: str | Path) -> list[dict[str, Any]]:
    """
    Returns rows in the unified form:
    {year, section, subject_id, faculty_id, total_hours, continuous_hours_optional}
    """
    df = _load_table(main_config_path)
    if df.empty:
        raise ValueError("Main config file is empty")
    cols = {_norm_col(c): c for c in df.columns}

    # Row-based template (prompt style)
    needed = ["year", "section", "subject_id", "faculty_id", "total_hours"]
    has_row_based = all(k in cols for k in needed)
    if has_row_based:
        year_col = cols["year"]
        section_col = cols["section"]
        subject_col = cols["subject_id"]
        faculty_col = cols["faculty_id"]
        total_hours_col = cols["total_hours"]
        cont_col = cols.get("continuous_hours")

        rows: list[dict[str, Any]] = []
        for _, r in df.iterrows():
            year = _to_str(r.get(year_col))
            section = _to_str(r.get(section_col))
            subject_id = _to_str(r.get(subject_col))
            faculty_id = _to_str(r.get(faculty_col))
            total_hours = _to_int(r.get(total_hours_col))
            continuous_hours = _to_int(r.get(cont_col)) if cont_col else None

            if not year or not section or not subject_id or not faculty_id or total_hours is None:
                continue
            if total_hours <= 0:
                continue
            rows.append(
                {
                    "year": year,
                    "section": section,
                    "subject_id": subject_id,
                    "faculty_id": faculty_id,
                    "total_hours": total_hours,
                    "continuous_hours_optional": continuous_hours,
                }
            )
        if not rows:
            raise ValueError("No valid rows in main config file")
        return rows

    # Grouped template (repo sample style): YEAR + SUBJECT_ID, then per section triplets.
    # Expected columns include: <SECTION>_HOURS, <SECTION>_FACULTY_ID, <SECTION>_CONTINUOUS_HOURS (optional).
    df.columns = [str(c).strip() for c in df.columns]
    lc_map = {_norm_col(c): c for c in df.columns}
    year_col = lc_map.get("year")
    subject_col = lc_map.get("subject_id") or lc_map.get("subject")
    if not year_col or not subject_col:
        raise ValueError("Main config columns not recognized (row-based or grouped)")

    # Find *_hours columns that are not continuous hours.
    hours_cols: dict[str, str] = {}
    for c in df.columns:
        orig = str(c).strip()
        cn = _norm_col(orig)
        if cn.endswith("_hours") and "continuous" not in cn:
            # Preserve the section name casing from the column header.
            lower_orig = orig.lower()
            section_name = orig[: -len("_hours")] if lower_orig.endswith("_hours") else orig
            hours_cols[section_name] = c

    if not hours_cols:
        raise ValueError("Main config grouped template not recognized (no <SECTION>_HOURS columns)")

    # For each section, discover faculty/continuous columns if present.
    rows: list[dict[str, Any]] = []
    for section_name, hours_column in hours_cols.items():
        faculty_column = None
        continuous_column = None
        sec_norm = _norm_col(section_name)
        for c in df.columns:
            cn = _norm_col(c)
            if cn == f"{sec_norm}_faculty_id":
                faculty_column = c
            if cn == f"{sec_norm}_continuous_hours":
                continuous_column = c

        if faculty_column is None:
            raise ValueError(f"Missing grouped faculty column for section {section_name}")

        for _, r in df.iterrows():
            year = _to_str(r.get(year_col))
            section = _to_str(section_name)
            subject_id = _to_str(r.get(subject_col))
            faculty_id = _to_str(r.get(faculty_column))
            total_hours = _to_int(r.get(hours_column))
            continuous_hours = _to_int(r.get(continuous_column)) if continuous_column else None
            if not year or not section or not subject_id or not faculty_id or total_hours is None:
                continue
            if total_hours <= 0:
                continue
            rows.append(
                {
                    "year": year,
                    "section": section,
                    "subject_id": subject_id,
                    "faculty_id": faculty_id,
                    "total_hours": total_hours,
                    "continuous_hours_optional": continuous_hours,
                }
            )
    if not rows:
        raise ValueError("No valid rows in grouped main config file")
    return rows


def _read_faculty_subject_maps(faculty_id_name_path: str | Path, subject_id_name_path: str | Path) -> tuple[dict[str, str], dict[str, str]]:
    faculty_df = _load_table(faculty_id_name_path)
    subject_df = _load_table(subject_id_name_path)
    if faculty_df.empty or subject_df.empty:
        raise ValueError("Mapping files are empty")

    fac_cols = {_norm_col(c): c for c in faculty_df.columns}
    fac_id_col = fac_cols.get("faculty_id") or fac_cols.get("id_assigned") or fac_cols.get("id") or fac_cols.get("idassigned")
    fac_name_col = fac_cols.get("faculty_name") or fac_cols.get("facultyname") or fac_cols.get("faculty_name".replace(" ", "")) or fac_cols.get("facultyname") or fac_cols.get("faculty_name")
    if not fac_id_col or not fac_name_col:
        # Try template headers: "faculty name", "id assigned"
        fac_id_col = fac_cols.get("id_assigned") or fac_cols.get("id") or fac_cols.get("idassigned")
        fac_name_col = fac_cols.get("faculty_name") or fac_cols.get("facultyname") or fac_cols.get("facultyname") or fac_cols.get("name")
    if not fac_id_col or not fac_name_col:
        raise ValueError("Faculty ID mapping columns not recognized")

    faculty_id_to_name: dict[str, str] = {}
    for _, r in faculty_df.iterrows():
        fid = _to_str(r.get(fac_id_col))
        fname = _to_str(r.get(fac_name_col))
        if fid and fname:
            faculty_id_to_name[str(fid)] = fname

    sub_cols = {_norm_col(c): c for c in subject_df.columns}
    sub_id_col = sub_cols.get("subject_id") or sub_cols.get("id")
    sub_name_col = sub_cols.get("subject_name") or sub_cols.get("subjectname") or sub_cols.get("subject_name".replace(" ", "")) or sub_cols.get("name") or sub_cols.get("subject")
    if not sub_id_col or not sub_name_col:
        raise ValueError("Subject ID mapping columns not recognized")

    subject_id_to_name: dict[str, str] = {}
    for _, r in subject_df.iterrows():
        sid = _to_str(r.get(sub_id_col))
        sname = _to_str(r.get(sub_name_col))
        if sid and sname:
            subject_id_to_name[str(sid)] = sname

    return faculty_id_to_name, subject_id_to_name


def _read_continuous_mapping(continuous_hours_mapping_path: str | Path) -> dict[str, int]:
    df = _load_table(continuous_hours_mapping_path)
    if df.empty:
        raise ValueError("Continuous hours mapping file is empty")
    cols = {_norm_col(c): c for c in df.columns}
    sid_col = cols.get("subject_id") or cols.get("subject")
    ch_col = cols.get("continuous_hours") or cols.get("compulsory_continuous_hours") or cols.get("compulsory_continuous_h") or cols.get("compulsory_continuous_hours".replace(" ", ""))
    if not sid_col or not ch_col:
        raise ValueError("Continuous hours mapping columns not recognized")

    mapping: dict[str, int] = {}
    for _, r in df.iterrows():
        sid = _to_str(r.get(sid_col))
        ch = _to_int(r.get(ch_col))
        if not sid or ch is None:
            continue
        if ch <= 0:
            continue
        mapping[sid] = ch
    if not mapping:
        raise ValueError("No continuous hours found")
    return mapping


def _read_faculty_availability(availability_path: str | Path) -> dict[str, dict[str, set[int]]]:
    """
    Supports:
    - Day-grid template: first column faculty id, then Monday..Saturday columns containing comma-separated periods.
    - Row-wise template: faculty_id/day/period
    """
    df = _load_table(availability_path)
    if df.empty:
        raise ValueError("Faculty availability file is empty")
    cols = {_norm_col(c): c for c in df.columns}

    day_cols = {d.lower(): cols.get(d.lower()[:3]) for d in DAYS}
    # Better: find any column whose normalized name matches a day.
    normalized_cols = {_norm_col(c): c for c in df.columns}
    day_col_by_day: dict[str, str] = {}
    for d in DAYS:
        key1 = d.lower()
        key2 = d.lower()[:3]
        for nc, orig in normalized_cols.items():
            if nc == key1.replace(" ", "").replace("-", "_") or nc == key2:
                day_col_by_day[d] = orig

    faculty_id_col = cols.get("faculty_id") or cols.get("facultyid") or cols.get("faculty")
    if not faculty_id_col:
        # Try day-grid template: "Faculty ID"
        for c in df.columns:
            if _norm_col(c) in {"facultyid", "faculty_id", "faculty"}:
                faculty_id_col = c
                break
    if not faculty_id_col:
        raise ValueError("Faculty availability column for FACULTY_ID not recognized")

    # If day-grid columns exist, parse that format.
    if any(d in day_col_by_day for d in DAYS):
        availability: dict[str, dict[str, set[int]]] = {}
        for _, r in df.iterrows():
            fid = _to_str(r.get(faculty_id_col))
            if not fid:
                continue
            day_map: dict[str, set[int]] = {d: set() for d in DAYS}
            for d in DAYS:
                c = day_col_by_day.get(d)
                if not c:
                    continue
                periods = _parse_periods(r.get(c))
                day_map[d] = {p for p in periods if p in PERIODS}
            availability[str(fid)] = day_map
        if not availability:
            raise ValueError("No faculty parsed from availability file")
        return availability

    # Row-wise: faculty_id, day, period
    fid_col = faculty_id_col
    day_col = cols.get("day")
    period_col = cols.get("period") or cols.get("periods")
    if not day_col or not period_col:
        raise ValueError("Faculty availability columns not recognized (need day grid or row-wise day/period)")
    availability: dict[str, dict[str, set[int]]] = {}
    for _, r in df.iterrows():
        fid = _to_str(r.get(fid_col))
        if not fid:
            continue
        day = _parse_day(r.get(day_col))
        p = _to_int(r.get(period_col))
        if not day or p is None or p not in PERIODS:
            continue
        availability.setdefault(str(fid), {d: set() for d in DAYS})
        availability[str(fid)][day].add(p)
    if not availability:
        raise ValueError("No faculty availability rows parsed")
    return availability


def _read_labs(lab_path: str | Path) -> list[dict[str, Any]]:
    df = _load_table(lab_path)
    if df.empty:
        raise ValueError("Lab timetable file is empty")
    cols = {_norm_col(c): c for c in df.columns}

    year_col = cols.get("year")
    section_col = cols.get("section")
    subject_col = cols.get("subject_id") or cols.get("subjectid") or cols.get("subject") or cols.get("subjectid")
    day_col = cols.get("day")
    periods_col = cols.get("periods") or cols.get("hours") or cols.get("period")
    faculty_col = cols.get("faculty_id") or cols.get("facultyid") or cols.get("faculty")
    if not year_col or not section_col or not subject_col or not day_col or not periods_col:
        raise ValueError("Lab timetable columns not recognized")

    labs: list[dict[str, Any]] = []
    for _, r in df.iterrows():
        year = _to_str(r.get(year_col))
        section = _to_str(r.get(section_col))
        subject_id = _to_str(r.get(subject_col))
        day = _parse_day(r.get(day_col))
        periods = _parse_periods(r.get(periods_col))
        faculty_id = _to_str(r.get(faculty_col)) if faculty_col else ""

        if not year or not section or not subject_id or not day or not periods:
            continue
        uniq_periods = sorted({p for p in periods if p in PERIODS})
        if not uniq_periods:
            continue
        labs.append(
            {
                "year": year,
                "section": section,
                "subject_id": subject_id,
                "day": day,
                "periods": uniq_periods,
                "faculty_id": faculty_id,
            }
        )
    if not labs:
        raise ValueError("No valid lab entries found")
    return labs


def _read_shared_classes(shared_classes_path: str | Path) -> list[dict[str, Any]]:
    df = _load_table(shared_classes_path)
    if df.empty:
        raise ValueError("Shared class file is empty")
    cols = {_norm_col(c): c for c in df.columns}
    year_col = cols.get("year")
    subject_col = cols.get("subject_id") or cols.get("subject")
    sections_col = cols.get("sections") or cols.get("section_list") or cols.get("section")
    faculty_col = cols.get("faculty_id") or cols.get("facultyid") or cols.get("faculty")
    if not year_col or not subject_col or not sections_col:
        raise ValueError("Shared class columns not recognized")

    shared: list[dict[str, Any]] = []
    for _, r in df.iterrows():
        year = _to_str(r.get(year_col))
        subject_id = _to_str(r.get(subject_col))
        sections_raw = _to_str(r.get(sections_col))
        faculty_id = _to_str(r.get(faculty_col)) if faculty_col else ""
        if not year or not subject_id or not sections_raw:
            continue
        sections = [s.strip() for s in sections_raw.replace(" ", "").split(",") if s.strip()]
        if not sections:
            continue
        shared.append(
            {
                "year": year,
                "subject_id": subject_id,
                "sections": sections,
                "faculty_id": faculty_id,
            }
        )
    if not shared:
        raise ValueError("No shared class entries found")
    return shared


@dataclass(frozen=True)
class LabSessionKey:
    year: str
    subject_id: str
    faculty_id: str
    day: str
    periods: tuple[int, ...]


@dataclass(frozen=True)
class SharedSessionKey:
    source: str  # "shared_class_file" | "lab_file"
    year: str
    subject_id: str
    faculty_id: str
    day: str
    periods: tuple[int, ...]


@dataclass
class Task:
    year: str
    subject_id: str
    faculty_id: str
    sections: tuple[str, ...]
    continuous_len: int


def _validate_main_totals(main_rows: list[dict[str, Any]]) -> None:
    totals: dict[tuple[str, str], int] = {}
    for r in main_rows:
        key = (r["year"], r["section"])
        totals[key] = totals.get(key, 0) + int(r["total_hours"])
    for (year, section), total in sorted(totals.items()):
        if total != 42:
            raise MainConfigValidationError(year=year, section=section, actual_hours=total)


def _make_output_workbook_section_timetables(
    section_timetables: dict[tuple[str, str], dict[str, list[str | None]]],
    faculty_id_to_name: dict[str, str],
    subject_id_to_name: dict[str, str],
) -> Workbook:
    """
    section_timetables[(year, section)] = {day: [cell for periods 1..7]}
    cell values are either None or "SUBJECT_ID|FACULTY_ID"
    """
    wb = Workbook()
    wb.remove(wb.active)
    for (year, section), day_map in sorted(section_timetables.items()):
        ws = wb.create_sheet(title=f"{year}_{section}"[:31])
        ws.append(["DAY", *PERIODS])
        for day in DAYS:
            row: list[Any] = [day]
            for period_idx in PERIODS:
                cell = day_map[day][period_idx - 1]
                if not cell:
                    row.append("")
                    continue
                sid, fid = cell.split("|", 1)
                subj_name = subject_id_to_name.get(sid, sid)
                fac_name = faculty_id_to_name.get(fid, fid)
                row.append(f"{subj_name} | {fac_name}")
            ws.append(row)
    return wb


def _make_output_workbook_faculty_workload(
    faculty_workload: dict[tuple[str, str], dict[str, list[str | None]]],
    faculty_id_to_name: dict[str, str],
    subject_id_to_name: dict[str, str],
) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    for (year, fid), day_map in sorted(faculty_workload.items(), key=lambda x: (x[0][0], x[0][1])):
        fname = faculty_id_to_name.get(fid, fid)
        ws = wb.create_sheet(title=f"{year}_{fname}"[:31])
        ws.append(["DAY", *PERIODS])
        for day in DAYS:
            row: list[Any] = [day]
            for pidx in PERIODS:
                cell = day_map[day][pidx - 1]
                if not cell:
                    row.append("")
                    continue
                # cell: "subject_id;sections_csv"
                sid, sections_csv = cell.split(";", 1)
                subj_name = subject_id_to_name.get(sid, sid)
                row.append(f"{subj_name} ({sections_csv})" if sections_csv else subj_name)
            ws.append(row)
    return wb


def _make_output_workbook_shared_report(shared_sessions: list[dict[str, Any]], faculty_id_to_name: dict[str, str], subject_id_to_name: dict[str, str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "SharedClasses"
    ws.append(["YEAR", "SUBJECT", "FACULTY", "SECTIONS", "DAY", "PERIODS"])
    for s in shared_sessions:
        ws.append(
            [
                s["year"],
                subject_id_to_name.get(s["subject_id"], s["subject_id"]),
                faculty_id_to_name.get(s["faculty_id"], s["faculty_id"]),
                ",".join(s["sections"]),
                s["day"],
                ",".join(str(p) for p in s["periods"]),
            ]
        )
    return wb


def _make_output_workbook_constraint_violations(violations: list[dict[str, Any]], subject_id_to_name: dict[str, str], faculty_id_to_name: dict[str, str]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "ConstraintViolations"
    ws.append(["YEAR", "SECTIONS", "SUBJECT", "FACULTY", "CONSTRAINT", "DETAIL"])
    for v in violations:
        ws.append(
            [
                v.get("year", ""),
                ",".join(v.get("sections", []) or []),
                subject_id_to_name.get(str(v.get("subject_id", "")), str(v.get("subject_id", ""))),
                faculty_id_to_name.get(str(v.get("faculty_id", "")), str(v.get("faculty_id", ""))),
                v.get("constraint", ""),
                v.get("detail", ""),
            ]
        )
    return wb


def _default_full_availability() -> dict[str, set[int]]:
    return {d: set(PERIODS) for d in DAYS}


def _faculty_allowed_periods(
    availability: dict[str, dict[str, set[int]]],
    faculty_id: str,
    day: str,
) -> set[int]:
    # Prompt rule: if not listed, fully available.
    if faculty_id not in availability:
        return set(PERIODS)
    return availability[faculty_id].get(day, set())


def generate_timetable_from_files(
    main_config_path: str | Path,
    lab_timetable_path: str | Path,
    shared_class_path: str | Path,
    faculty_id_name_path: str | Path,
    subject_id_name_path: str | Path,
    continuous_hours_mapping_path: str | Path,
    faculty_availability_path: str | Path,
    out_dir: str | Path,
) -> dict[str, Path]:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)

    main_rows = _read_main_config(main_config_path)
    _validate_main_totals(main_rows)  # Step-1 (mandatory)

    faculty_id_to_name, subject_id_to_name = _read_faculty_subject_maps(faculty_id_name_path, subject_id_name_path)
    continuous_map = _read_continuous_mapping(continuous_hours_mapping_path)
    availability = _read_faculty_availability(faculty_availability_path)
    labs = _read_labs(lab_timetable_path)
    shared_rows = _read_shared_classes(shared_class_path)

    # Index main config by (year, section, subject)
    main_index: dict[tuple[str, str, str], dict[str, Any]] = {}
    years: set[str] = set()
    for r in main_rows:
        years.add(r["year"])
        key = (r["year"], r["section"], r["subject_id"])
        main_index[key] = r

    # Remaining hours starts as main total hours; labs subtract hard-fixed hours.
    # Also keep a copy of each section's (subject -> hours).
    section_subject_hours: dict[tuple[str, str, str], int] = {}
    for (year, section, subject_id), r in main_index.items():
        section_subject_hours[(year, section, subject_id)] = int(r["total_hours"])

    # Prepare output containers
    section_timetables: dict[tuple[str, str], dict[str, list[str | None]]] = {}
    faculty_workload: dict[tuple[str, str], dict[str, list[str | None]]] = {}
    shared_report_sessions: list[dict[str, Any]] = []
    constraint_violations: list[dict[str, Any]] = []

    # Year-wise independent execution
    for year in sorted(years):
        # Step-2 initialization (IDs only)
        timetable: dict[str, dict[str, list[dict | None]]] = {}
        faculty_schedule: dict[str, dict[str, list[str | None]]] = {}
        remaining_hours: dict[tuple[str, str], int] = {}
        shared_sessions: list[SharedSessionKey] = []

        sections_for_year = sorted({sec for (y, sec, _) in main_index.keys() if y == year})

        for section in sections_for_year:
            timetable[section] = {day: [None for _ in PERIODS] for day in DAYS}
        for key in main_index.keys():
            if key[0] != year:
                continue
            fid = str(main_index[key]["faculty_id"]).strip()
            faculty_schedule.setdefault(fid, {day: [None for _ in PERIODS] for day in DAYS})

        # Remaining hours by (section, subject)
        for (y, section, subject_id), hrs in section_subject_hours.items():
            if y != year:
                continue
            remaining_hours[(section, subject_id)] = hrs

        # ---- Step-3/4: LAB assignment (hard constraint) ----
        # Group by (year, subject, faculty, day, periods) so shared labs are consistent.
        lab_group_map: dict[LabSessionKey, set[str]] = {}
        lab_entry_periods_by_section_subject: dict[tuple[str, str, str, str], list[int]] = {}

        for lab in labs:
            if lab["year"] != year:
                continue
            section = lab["section"]
            subject_id = lab["subject_id"]
            day = lab["day"]
            periods = tuple(lab["periods"])

            # Faculty: from lab file if present; else derived from main config (still file-driven).
            fid_from_lab = lab.get("faculty_id", "")
            if fid_from_lab:
                faculty_id = str(fid_from_lab).strip()
            else:
                main_key = (year, section, subject_id)
                faculty_id = str(main_index.get(main_key, {}).get("faculty_id", "")).strip()
            if not faculty_id:
                constraint_violations.append(
                    {
                        "year": year,
                        "sections": [section],
                        "subject_id": subject_id,
                        "faculty_id": "",
                        "constraint": "lab faculty missing",
                        "detail": "Lab entry has no FACULTY_ID and cannot be derived from main config.",
                    }
                )
                continue

            key = LabSessionKey(year=year, subject_id=subject_id, faculty_id=faculty_id, day=day, periods=periods)
            lab_group_map.setdefault(key, set()).add(section)

            # Track periods to subtract from remaining hours
            lab_entry_periods_by_section_subject.setdefault((year, section, subject_id, faculty_id), [])
            lab_entry_periods_by_section_subject[(year, section, subject_id, faculty_id)] = list(periods)

        # Place labs
        # session_id stored in faculty_schedule is a string key; timetable stores subject_id/faculty_id.
        for gkey, sections in sorted(lab_group_map.items(), key=lambda x: (x[0].day, x[0].periods, x[0].subject_id)):
            session_id = f"LAB|{gkey.subject_id}|{gkey.faculty_id}|{gkey.day}|{','.join(str(p) for p in gkey.periods)}"
            # Shared lab detection for report
            if len(sections) > 1:
                shared_sessions.append(SharedSessionKey(source="lab_file", year=year, subject_id=gkey.subject_id, faculty_id=gkey.faculty_id, day=gkey.day, periods=gkey.periods))

            for section in sorted(sections):
                for p in gkey.periods:
                    if timetable[section][gkey.day][p - 1] is not None:
                        constraint_violations.append(
                            {
                                "year": year,
                                "sections": [section],
                                "subject_id": gkey.subject_id,
                                "faculty_id": gkey.faculty_id,
                                "constraint": "lab overlap",
                                "detail": f"Lab overlap at {gkey.day} P{p} for section {section}.",
                            }
                        )
                        continue
                    timetable[section][gkey.day][p - 1] = {"subject_id": gkey.subject_id, "faculty_id": gkey.faculty_id, "source": "lab_file"}
                    faculty_schedule.setdefault(gkey.faculty_id, {day: [None for _ in PERIODS] for day in DAYS})
                    if faculty_schedule[gkey.faculty_id][gkey.day][p - 1] not in (None, session_id):
                        constraint_violations.append(
                            {
                                "year": year,
                                "sections": [section],
                                "subject_id": gkey.subject_id,
                                "faculty_id": gkey.faculty_id,
                                "constraint": "faculty conflict (lab)",
                                "detail": f"Faculty {gkey.faculty_id} already scheduled at {gkey.day} P{p} for a different session.",
                            }
                        )
                    faculty_schedule[gkey.faculty_id][gkey.day][p - 1] = session_id

                # Subtract lab hours from remaining_hours
                remaining_hours[(section, gkey.subject_id)] = max(
                    0, remaining_hours.get((section, gkey.subject_id), 0) - len(gkey.periods)
                )

        # ---- Step-6: apply Shared Class File ----
        # Build shared requirement descriptors.
        shared_groups: list[dict[str, Any]] = []
        # Track covered (year, section, subject) so we don't schedule them again.
        covered_by_shared: set[tuple[str, str, str]] = set()

        for row in shared_rows:
            if row["year"] != year:
                continue
            subject_id = row["subject_id"]
            sections = [s for s in row["sections"] if s]
            faculty_id_file = row.get("faculty_id", "")
            faculty_id_file = str(faculty_id_file).strip()
            if len(sections) < 2:
                # Shared session still valid but report should reflect file-driven definition.
                pass

            # Validate all sections exist and subject exists in main config
            missing_sections = [s for s in sections if (year, s, subject_id) not in main_index]
            if missing_sections:
                constraint_violations.append(
                    {
                        "year": year,
                        "sections": missing_sections,
                        "subject_id": subject_id,
                        "faculty_id": faculty_id_file,
                        "constraint": "shared class constraint",
                        "detail": "Shared class references sections missing from main config.",
                    }
                )
                continue

            # Determine the effective faculty_id for this shared group:
            # - if provided in shared file, enforce it matches main config faculty for every section.
            # - if missing, derive from main config (still file-driven).
            effective_faculty_id = faculty_id_file
            if not effective_faculty_id:
                s0 = sections[0]
                effective_faculty_id = str(main_index[(year, s0, subject_id)]["faculty_id"]).strip()

            mismatch_sections = [s for s in sections if str(main_index[(year, s, subject_id)]["faculty_id"]).strip() != effective_faculty_id]
            if mismatch_sections:
                constraint_violations.append(
                    {
                        "year": year,
                        "sections": mismatch_sections,
                        "subject_id": subject_id,
                        "faculty_id": effective_faculty_id,
                        "constraint": "shared class constraint",
                        "detail": "All shared sections must have the same faculty.",
                    }
                )
                continue

            # Validate equal remaining hours for all sections (required for same day/period schedule).
            remaining_list = [remaining_hours.get((s, subject_id), 0) for s in sections]
            if len(set(remaining_list)) != 1:
                constraint_violations.append(
                    {
                        "year": year,
                        "sections": sections,
                        "subject_id": subject_id,
                        "faculty_id": effective_faculty_id,
                        "constraint": "shared class constraint",
                        "detail": "Sections in a shared class must have equal remaining hours after lab subtraction.",
                    }
                )
                continue

            remaining = int(remaining_list[0])
            if remaining <= 0:
                continue

            # Continuous length for the subject
            cont_len = int(continuous_map.get(subject_id, 1))
            # Optional override from main config: if any section row provides continuous hours, enforce equality.
            overrides = []
            for s in sections:
                opt = main_index[(year, s, subject_id)].get("continuous_hours_optional")
                if opt is not None and int(opt) > 0:
                    overrides.append(int(opt))
            if overrides:
                if len(set(overrides)) != 1:
                    constraint_violations.append(
                        {
                            "year": year,
                            "sections": sections,
                            "subject_id": subject_id,
                            "faculty_id": effective_faculty_id,
                            "constraint": "continuous hours constraint",
                            "detail": "Shared class sections have conflicting continuous hours values in main config.",
                        }
                    )
                    continue
                cont_len = overrides[0]

            if cont_len <= 0:
                cont_len = 1

            # Create shared requirement descriptor
            shared_groups.append(
                {
                    "year": year,
                    "subject_id": subject_id,
                    "faculty_id": effective_faculty_id,
                    "sections": tuple(sorted(set(sections))),
                    "remaining_hours": remaining,
                    "continuous_len": cont_len,
                }
            )
            for s in sections:
                covered_by_shared.add((year, s, subject_id))

        # ---- Step-8: Assign remaining subjects with backtracking ----
        # Build task list (expand remaining hours into fixed consecutive blocks).
        tasks: list[Task] = []

        # Expand each subject's remaining hours into consecutive blocks.
        # Continuous-hours mapping defines the preferred block size; the final remainder block
        # may be shorter but must still be consecutive.
        for g in shared_groups:
            rem = int(g["remaining_hours"])
            k = int(g["continuous_len"])
            if k <= 0:
                k = 1
            while rem > 0:
                block = min(k, rem)
                tasks.append(
                    Task(
                        year=year,
                        subject_id=g["subject_id"],
                        faculty_id=g["faculty_id"],
                        sections=g["sections"],
                        continuous_len=block,
                    )
                )
                rem -= block

        # Non-shared remaining subjects
        for (y, section, subject_id), total in section_subject_hours.items():
            if y != year:
                continue
            if (y, section, subject_id) in covered_by_shared:
                continue
            rem = int(remaining_hours.get((section, subject_id), 0))
            if rem <= 0:
                continue
            cont_len = int(continuous_map.get(subject_id, 1))
            opt = main_index[(year, section, subject_id)].get("continuous_hours_optional")
            if opt is not None and int(opt) > 0:
                cont_len = int(opt)
            if cont_len <= 0:
                cont_len = 1
            faculty_id = str(main_index[(year, section, subject_id)]["faculty_id"]).strip()
            while rem > 0:
                block = min(cont_len, rem)
                tasks.append(Task(year=year, subject_id=subject_id, faculty_id=faculty_id, sections=(section,), continuous_len=block))
                rem -= block

        # Solve tasks via backtracking + slot scoring
        # We'll store placement info for later faculty workload + shared report.
        placed = 0
        placement_log: list[dict[str, Any]] = []

        def _score_slot(task: Task, day: str, start_period: int) -> int:
            # Spec weights:
            # +5 section free
            # +5 faculty free
            # +3 continuous possible
            # -5 conflict
            block_len = task.continuous_len
            periods = list(range(start_period, start_period + block_len))
            section_free = all(timetable[s][day][p - 1] is None for s in task.sections for p in periods)
            faculty_free = all(faculty_schedule.get(task.faculty_id, {}).get(day, [None for _ in PERIODS])[p - 1] is None for p in periods)
            continuous_possible = start_period + block_len - 1 <= PERIODS[-1]
            conflict = 0
            if not faculty_free:
                # If occupied, it's a conflict unless it is the same session id (we never re-place tasks).
                conflict = -5
            score = 0
            if section_free:
                score += 5
            if faculty_free:
                score += 5
            if continuous_possible:
                score += 3
            score += conflict
            return score

        def _enumerate_candidates(task: Task) -> list[tuple[str, int]]:
            cands: list[tuple[str, int]] = []
            L = task.continuous_len
            for day in DAYS:
                allowed = _faculty_allowed_periods(availability, task.faculty_id, day)
                if not allowed:
                    continue
                for start in range(PERIODS[0], PERIODS[-1] - L + 2):
                    if any(p not in allowed for p in range(start, start + L)):
                        continue
                    periods = list(range(start, start + L))
                    if not all(timetable[s][day][p - 1] is None for s in task.sections for p in periods):
                        continue
                    # Faculty conflict rule: not allowed if different subject same time.
                    # Since we place each task once, any occupied slot is a conflict.
                    if any(faculty_schedule.get(task.faculty_id, {}).get(day, [None for _ in PERIODS])[p - 1] is not None for p in periods):
                        continue
                    cands.append((day, start))
            # Highest score first; deterministic tie-break: day then start.
            cands.sort(key=lambda t: (-_score_slot(task, t[0], t[1]), DAYS.index(t[0]), t[1]))
            return cands

        def _apply_task(task: Task, day: str, start_period: int, session_source: str) -> str:
            periods = list(range(start_period, start_period + task.continuous_len))
            session_id = f"{session_source}|{task.subject_id}|{task.faculty_id}|{day}|{','.join(str(p) for p in periods)}"
            timetable_source = session_source
            for p in periods:
                faculty_schedule.setdefault(task.faculty_id, {d: [None for _ in PERIODS] for d in DAYS})
                faculty_schedule[task.faculty_id][day][p - 1] = session_id
                for s in task.sections:
                    timetable[s][day][p - 1] = {"subject_id": task.subject_id, "faculty_id": task.faculty_id, "source": timetable_source}
            return session_id

        def _undo_task(task: Task, day: str, start_period: int) -> None:
            periods = list(range(start_period, start_period + task.continuous_len))
            for p in periods:
                faculty_schedule[task.faculty_id][day][p - 1] = None
                for s in task.sections:
                    timetable[s][day][p - 1] = None

        # Determine source for each task: if task is from shared_groups, it should log shared class report.
        # We do that by matching task.sections length > 1 -> "shared_class_file" (as created from file).
        task_sources: list[str] = ["shared_class_file" if len(t.sections) > 1 else "solver" for t in tasks]

        # Subject order for retry: deterministic permutations based on attempt index.
        subject_order_base = []
        seen_subjects = set()
        for t in tasks:
            if t.subject_id not in seen_subjects:
                subject_order_base.append(t.subject_id)
                seen_subjects.add(t.subject_id)

        attempt_orders: list[list[str]] = []
        if subject_order_base:
            attempt_orders.append(subject_order_base)
            attempt_orders.append(list(reversed(subject_order_base)))
            # Rotate left by 1 and right by 1 deterministically.
            attempt_orders.append(subject_order_base[1:] + subject_order_base[:1])
            attempt_orders.append(subject_order_base[-1:] + subject_order_base[:-1])
            # Sort by decreasing blocks as a 5th deterministic order
            blocks_by_subject: dict[str, int] = {}
            for t in tasks:
                blocks_by_subject[t.subject_id] = blocks_by_subject.get(t.subject_id, 0) + 1
            attempt_orders.append(sorted(subject_order_base, key=lambda sid: (-blocks_by_subject[sid], sid)))
        else:
            attempt_orders = [list(subject_order_base) for _ in range(5)]

        solution_found = False

        def backtrack(task_list: list[Task], task_idx_map: dict[int, Task]) -> bool:
            nonlocal placed, solution_found
            if placed == len(tasks):
                return True
            # Choose next task with minimum candidates; deterministic tie-break by task order in current list.
            best_task_i: int | None = None
            best_cands: list[tuple[str, int]] = []
            best_len: int | None = None
            for i, task in enumerate(task_list):
                if task is None:  # type: ignore[comparison-overlap]
                    continue
                cands = _enumerate_candidates(task)
                if not cands:
                    return False
                if best_len is None or len(cands) < best_len:
                    best_len = len(cands)
                    best_task_i = i
                    best_cands = cands
                elif best_len is not None and len(cands) == best_len:
                    # Tie-break: earlier in original task list
                    if best_task_i is None or i < best_task_i:
                        best_task_i = i
                        best_cands = cands

            if best_task_i is None:
                return False
            task = task_list[best_task_i]
            task_list[best_task_i] = None  # type: ignore[assignment]
            source = task_sources[best_task_i]

            for day, start in best_cands:
                _apply_task(task, day, start, session_source=source)
                placement_log.append({"task": task, "day": day, "start": start, "source": source})
                placed += 1
                if backtrack(task_list, task_idx_map):
                    return True
                placed -= 1
                placement_log.pop()
                _undo_task(task, day, start)

            task_list[best_task_i] = task
            return False

        for attempt in range(5):
            # Reset mutable state for each attempt (only tasks placement differs).
            # We rebuild timetable + faculty_schedule from lab placements for each attempt.
            # Labs are hard-fixed, so we can re-apply them quickly by reusing stored timetable initial from labs.
            # Here we reconstruct from scratch from the current 'timetable' which already includes labs.
            # For correctness and determinism, we undo all solver tasks but keep labs.
            # Since we only used backtracking on the current timetable state, we need to reset.
            if solution_found:
                break

            # Reset solver-placed entries (clear anything with source=="solver" or "shared_class_file")
            for section in sections_for_year:
                for day in DAYS:
                    for pidx, period in enumerate(PERIODS):
                        cell = timetable[section][day][pidx]
                        if cell is not None and cell.get("source") in {"solver", "shared_class_file"}:
                            timetable[section][day][pidx] = None
            for fid in list(faculty_schedule.keys()):
                for day in DAYS:
                    for pidx in range(len(PERIODS)):
                        # Clear any session id created by solver/shared classes.
                        sid = faculty_schedule[fid][day][pidx]
                        if sid is not None and (sid.startswith("solver|") or sid.startswith("shared_class_file|")):
                            faculty_schedule[fid][day][pidx] = None

            placed = 0
            placement_log = []

            # Reorder tasks by subject_id according to attempt order.
            order = attempt_orders[min(attempt, len(attempt_orders) - 1)]
            rank = {sid: i for i, sid in enumerate(order)}
            task_ordered = sorted(tasks, key=lambda t: (rank.get(t.subject_id, 10**9), t.faculty_id, t.sections))
            task_sources = ["shared_class_file" if len(t.sections) > 1 else "solver" for t in task_ordered]
            # Backtracking mutates task_list, so pass a list we can set None in.
            task_idx_map: dict[int, Task] = {}
            task_list = list(task_ordered)
            if backtrack(task_list, task_idx_map):
                solution_found = True
                break

        if not solution_found:
            # Constraint report: for each unplaced task, attempt to infer reason.
            # We do a conservative reason inference by checking existence of at least one feasible slot.
            # Note: this reports at the granularity of (subject, faculty, sections).
            remaining_task_counts: dict[tuple[str, str, tuple[str, ...]], int] = {}
            for t in tasks:
                remaining_task_counts[(t.subject_id, t.faculty_id, t.sections)] = remaining_task_counts.get((t.subject_id, t.faculty_id, t.sections), 0) + 1

            # If solver failed, assume all tasks remain unscheduled.
            for (sid, fid, secs), count in sorted(remaining_task_counts.items(), key=lambda x: (x[0][0], len(x[0][2]))):
                reason = "no free slot"
                # Check faculty availability first (spec prefers that message).
                avail_any = False
                section_any = False
                for day in DAYS:
                    allowed = _faculty_allowed_periods(availability, fid, day)
                    if not allowed:
                        continue
                    # Continuous possible start range
                    for start in range(PERIODS[0], PERIODS[-1] -  int(next(t.continuous_len for t in tasks if t.subject_id==sid and t.faculty_id==fid and t.sections==secs)) + 2):
                        L = int(next(t.continuous_len for t in tasks if t.subject_id==sid and t.faculty_id==fid and t.sections==secs))
                        if start + L - 1 > PERIODS[-1]:
                            continue
                        periods = list(range(start, start + L))
                        if any(p not in allowed for p in periods):
                            continue
                        # Section emptiness
                        if all(timetable[s][day][p - 1] is None for s in secs for p in periods):
                            section_any = True
                            # Faculty availability with faculty schedule emptiness
                            if all(faculty_schedule.get(fid, {}).get(day, [None for _ in PERIODS])[p - 1] is None for p in periods):
                                avail_any = True
                                break
                    if avail_any:
                        break
                if avail_any:
                    reason = "solver failed to find placement"
                elif section_any:
                    reason = "faculty availability conflict"
                else:
                    reason = "section conflict"

                constraint_violations.append(
                    {
                        "year": year,
                        "sections": list(secs),
                        "subject_id": sid,
                        "faculty_id": fid,
                        "constraint": "unscheduled subject",
                        "detail": f"Unable to place remaining hour-block(s) ({count} block(s)). Reason: {reason}.",
                    }
                )

        # Populate section timetables + faculty workload from the final timetable state.
        # Only if any constraint violations were recorded for this year, we still export partial state.
        for section in sections_for_year:
            day_map: dict[str, list[str | None]] = {d: [None for _ in PERIODS] for d in DAYS}
            for day in DAYS:
                for i, p in enumerate(PERIODS):
                    cell = timetable[section][day][i]
                    if cell is None:
                        day_map[day][i] = None
                        continue
                    sid = str(cell["subject_id"])
                    fid = str(cell["faculty_id"])
                    day_map[day][i] = f"{sid}|{fid}"
            section_timetables[(year, section)] = day_map

        # Faculty workload is derived from timetable placements.
        for section in sections_for_year:
            for day in DAYS:
                for i, p in enumerate(PERIODS):
                    cell = timetable[section][day][i]
                    if cell is None:
                        continue
                    fid = str(cell["faculty_id"])
                    sid = str(cell["subject_id"])
                    faculty_workload.setdefault((year, fid), {d: [None for _ in PERIODS] for d in DAYS})
                    # If already filled by another section with same shared session, keep consistent label.
                    existing = faculty_workload[(year, fid)][day][i]
                    if existing is None:
                        faculty_workload[(year, fid)][day][i] = f"{sid};{section}"
                    else:
                        # Merge sections if same subject (shared session)
                        ex_sid, ex_secs = existing.split(";", 1)
                        if ex_sid == sid:
                            secs = set([s for s in ex_secs.split(",") if s])
                            secs.add(section)
                            faculty_workload[(year, fid)][day][i] = f"{sid};{','.join(sorted(secs))}"

        # Shared report sessions for shared classes come ONLY from placements
        # originating from Shared Class File (tagged as "shared_class_file").
        if solution_found:
            for entry in placement_log:
                t: Task = entry["task"]
                if entry["source"] != "shared_class_file":
                    continue
                if len(t.sections) < 2:
                    continue
                periods = list(range(entry["start"], entry["start"] + t.continuous_len))
                shared_report_sessions.append(
                    {
                        "year": year,
                        "subject_id": t.subject_id,
                        "faculty_id": t.faculty_id,
                        "sections": list(t.sections),
                        "day": entry["day"],
                        "periods": periods,
                        "source": "shared_class_file",
                    }
                )

        # Add lab shared sessions (from grouping) into report
        for k in shared_sessions:
            shared_report_sessions.append(
                {
                    "year": k.year,
                    "subject_id": k.subject_id,
                    "faculty_id": k.faculty_id,
                    "sections": sorted(list(lab_group_map.get(LabSessionKey(year=k.year, subject_id=k.subject_id, faculty_id=k.faculty_id, day=k.day, periods=k.periods), set()))),
                    "day": k.day,
                    "periods": list(k.periods),
                    "source": k.source,
                }
            )

    # Write outputs
    section_wb = _make_output_workbook_section_timetables(section_timetables, faculty_id_to_name, subject_id_to_name)
    faculty_wb = _make_output_workbook_faculty_workload(faculty_workload, faculty_id_to_name, subject_id_to_name)
    shared_wb = _make_output_workbook_shared_report(shared_report_sessions, faculty_id_to_name, subject_id_to_name)
    constraint_wb = _make_output_workbook_constraint_violations(constraint_violations, subject_id_to_name, faculty_id_to_name)

    section_path = out / "section_timetables.xlsx"
    faculty_path = out / "faculty_workload.xlsx"
    shared_path = out / "shared_classes_report.xlsx"
    constraint_path = out / "constraint_violation_report.xlsx"

    section_wb.save(section_path)
    faculty_wb.save(faculty_path)
    shared_wb.save(shared_path)
    constraint_wb.save(constraint_path)

    return {
        "sectionTimetables": section_path,
        "facultyWorkload": faculty_path,
        "sharedClassesReport": shared_path,
        "constraintViolationReport": constraint_path,
    }


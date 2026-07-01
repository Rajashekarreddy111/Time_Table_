from io import BytesIO, StringIO
from pathlib import Path

import cv2
import numpy as np
import openpyxl
import pandas as pd
import pdfplumber
import pytesseract
from fastapi import HTTPException, UploadFile
from openpyxl.utils import get_column_letter


def validation_error(message: str, details: list | None = None) -> HTTPException:
    return HTTPException(
        status_code=400,
        detail={
            "error": "ValidationError",
            "message": message,
            "details": details or [],
        },
    )


def _normalize_dataframe(df: pd.DataFrame) -> pd.DataFrame:
    cleaned = df.copy()
    cleaned.columns = [str(col).strip() for col in cleaned.columns]
    
    # Fill down YEAR column if it exists (handles merged cells in Excel)
    year_cols = [c for c in cleaned.columns if str(c).upper() == "YEAR"]
    if year_cols:
        cleaned[year_cols[0]] = cleaned[year_cols[0]].ffill()

    cleaned = cleaned.where(pd.notnull(cleaned), None)
    return cleaned


def _clean_excel_cell(value) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _normalize_grouped_header_token(value) -> str:
    text = (_clean_excel_cell(value) or "").upper()
    return " ".join(text.replace("_", " ").replace("-", " ").split())


def _is_grouped_main_header_pair(first_row: list[str | None], second_row: list[str | None]) -> bool:
    if len(first_row) < 5 or len(second_row) < 5:
        return False

    fixed_headers = [_normalize_grouped_header_token(cell) for cell in first_row[:2]]
    if fixed_headers != ["YEAR", "SUBJECT"]:
        return False

    section_seen = False
    for column_index in range(2, len(second_row), 3):
        if column_index + 2 >= len(second_row):
            break
        section = first_row[column_index]
        triplet = second_row[column_index:column_index + 3]
        if not section:
            continue
        normalized_triplet = [_normalize_grouped_header_token(cell) for cell in triplet]
        if normalized_triplet != ["NO OF HOURS", "FACULTY ID", "CONTINUOUS HOURS"]:
            return False
        section_seen = True
    return section_seen


def _looks_like_grouped_main_timetable_header(rows: list[tuple]) -> bool:
    if len(rows) < 2:
        return False
    for row_index in range(len(rows) - 1):
        first_row = [_clean_excel_cell(value) for value in rows[row_index]]
        second_row = [_clean_excel_cell(value) for value in rows[row_index + 1]]
        if _is_grouped_main_header_pair(first_row, second_row):
            return True
    return False


def _parse_grouped_main_timetable_excel(file_bytes: bytes, sheet_name: str | None = None) -> pd.DataFrame:
    workbook = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    if sheet_name and sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active
    rows = list(worksheet.iter_rows(values_only=True))

    if not _looks_like_grouped_main_timetable_header(rows):
        raise ValueError("Workbook does not match grouped main timetable template")

    def extract_block_columns(header_row: list[str | None], subheader_row: list[str | None]) -> tuple[list[str], int]:
        columns: list[str] = ["YEAR", "SUBJECT_ID"]
        last_used_column = 2
        for column_index in range(2, max(len(header_row), len(subheader_row)), 3):
            if column_index + 2 >= len(subheader_row):
                break
            section = header_row[column_index] if column_index < len(header_row) else None
            triplet = subheader_row[column_index:column_index + 3]
            if not section:
                continue
            normalized_triplet = [_normalize_grouped_header_token(cell) for cell in triplet]
            if normalized_triplet != ["NO OF HOURS", "FACULTY ID", "CONTINUOUS HOURS"]:
                continue
            section_name = str(section).strip().upper()
            columns.extend(
                [
                    f"{section_name}_HOURS",
                    f"{section_name}_FACULTY_ID",
                    f"{section_name}_CONTINUOUS_HOURS",
                ]
            )
            last_used_column = column_index + 3
        return columns, last_used_column

    records: list[dict] = []
    row_index = 0

    while row_index < len(rows) - 1:
        first_row = [_clean_excel_cell(value) for value in rows[row_index]]
        second_row = [_clean_excel_cell(value) for value in rows[row_index + 1]]

        if not _is_grouped_main_header_pair(first_row, second_row):
            row_index += 1
            continue

        columns, block_width = extract_block_columns(first_row, second_row)
        if len(columns) <= 2:
            row_index += 1
            continue

        row_index += 2
        while row_index < len(rows):
            current_row = list(rows[row_index])
            cleaned_current = [_clean_excel_cell(value) for value in current_row]

            if all(value is None for value in cleaned_current):
                row_index += 1
                continue

            next_second_row = [_clean_excel_cell(value) for value in rows[row_index + 1]] if row_index + 1 < len(rows) else []
            if _is_grouped_main_header_pair(cleaned_current, next_second_row):
                break

            values = current_row[:block_width]
            if len(values) < block_width:
                values.extend([None] * (block_width - len(values)))

            record: dict[str, object | None] = {
                "YEAR": values[0] if len(values) > 0 and values[0] is not None else (records[-1]["YEAR"] if records else None),
                "SUBJECT_ID": values[1] if len(values) > 1 else None,
            }

            output_column_index = 2
            for column_index in range(2, block_width, 3):
                if output_column_index + 2 >= len(columns):
                    break
                record[columns[output_column_index]] = values[column_index] if column_index < len(values) else None
                record[columns[output_column_index + 1]] = values[column_index + 1] if column_index + 1 < len(values) else None
                record[columns[output_column_index + 2]] = values[column_index + 2] if column_index + 2 < len(values) else None
                output_column_index += 3

            if any(value is not None and str(value).strip() != "" for value in record.values()):
                records.append(record)
            row_index += 1

    return pd.DataFrame(records)


def parse_tabular_upload(file_name: str, file_bytes: bytes) -> pd.DataFrame:
    suffix = Path(file_name).suffix.lower()
    if suffix == ".csv":
        frame = pd.read_csv(StringIO(file_bytes.decode("utf-8-sig")))
        return _normalize_dataframe(frame)
    if suffix in {".xlsx", ".xls"}:
        if suffix == ".xlsx":
            try:
                frame = _parse_grouped_main_timetable_excel(file_bytes)
                return _normalize_dataframe(frame)
            except ValueError:
                pass
        frame = pd.read_excel(BytesIO(file_bytes))
        return _normalize_dataframe(frame)
    if suffix == ".pdf":
        return _normalize_dataframe(parse_pdf_to_dataframe(file_bytes))
    if suffix in {".png", ".jpg", ".jpeg"}:
        return _normalize_dataframe(parse_image_to_dataframe(file_bytes))
    raise validation_error("Unsupported file format", [f"Received extension: {suffix}"])


def parse_pdf_to_dataframe(file_bytes: bytes) -> pd.DataFrame:
    records: list[dict] = []
    with pdfplumber.open(BytesIO(file_bytes)) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables() or []
            for table in tables:
                if not table or len(table) < 2:
                    continue
                headers = [str(h).strip().lower() for h in table[0]]
                for row in table[1:]:
                    records.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers))})
    if records:
        return pd.DataFrame(records)
    raise validation_error("No structured table found in PDF", [])


def parse_image_to_dataframe(file_bytes: bytes) -> pd.DataFrame:
    import re

    np_arr = np.frombuffer(file_bytes, dtype="uint8")
    image = cv2.imdecode(np_arr, cv2.IMREAD_COLOR)
    if image is None:
        raise validation_error("Unable to decode image", [])

    text = pytesseract.image_to_string(image)
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if not lines:
        raise validation_error("No readable text found in image", [])

    # Try different delimiters: comma, tab, or 2+ spaces
    def split_line(l):
        # Prefer tab/comma, then fallback to multi-space
        if "\t" in l:
            return [p.strip() for p in l.split("\t") if p.strip()]
        if "," in l:
            return [p.strip() for p in l.split(",") if p.strip()]
        # Fallback: Treat 2+ spaces as a delimiter (common in OCR tables)
        return [p.strip() for p in re.split(r"\s{2,}", l) if p.strip()]

    header = [h.strip().lower() for h in split_line(lines[0])]
    if not header:
        raise validation_error("Header row is empty or unparseable", [])

    rows = []
    for line in lines[1:]:
        parts = [p.strip() for p in split_line(line)]
        if len(parts) == 0:
            continue
        # Allow some flexibility in column count for OCR noise
        if len(parts) < len(header):
            # Pad with empty strings if slightly short
            parts = parts + [""] * (len(header) - len(parts))
        elif len(parts) > len(header):
            # Truncate if too long
            parts = parts[:len(header)]

        rows.append({header[i]: parts[i] for i in range(len(header))})

    if rows:
        return pd.DataFrame(rows)
    raise validation_error(
        "OCR output is not in a recognized tabular format",
        [{"tip": "Try formatting input as a CSV-like text with clear columns or borders."}],
    )


def require_columns(df: pd.DataFrame, required_columns: list[str]) -> None:
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise validation_error(
            "Required columns are missing",
            [{"missingColumns": missing, "receivedColumns": list(df.columns)}],
        )


def dataframe_rows(df: pd.DataFrame) -> list[dict]:
    # Convert to dict with lowercased keys for case-insensitive access, but preserve original column names
    records = []
    column_map = {str(col).strip().lower(): str(col).strip() for col in df.columns}
    for _, row in df.iterrows():
        record = {}
        for lower_col, orig_col in column_map.items():
            record[lower_col] = row[orig_col]
            # Also store original column name for cases where we need it
            record[f"__orig_{lower_col}"] = orig_col
        records.append(record)
    return records


def read_upload_bytes(file: UploadFile) -> bytes:
    content = file.file.read()
    if not content:
        raise validation_error("Uploaded file is empty", [])
    return content


def create_excel_template(records: list[dict], include_example_rows: bool = True) -> bytes:
    frame = pd.DataFrame(records)
    if not include_example_rows:
        frame = frame.head(0)
    stream = BytesIO()
    frame.to_excel(stream, index=False, engine="openpyxl")
    stream.seek(0)
    return stream.read()


def create_grouped_main_timetable_template(
    records: list[dict],
    include_example_rows: bool = True,
) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Main Timetable"

    normalized_records = [{str(key).strip().upper(): value for key, value in row.items()} for row in records]
    section_names: list[str] = []
    for row in normalized_records:
        for key in row.keys():
            if key.endswith("_HOURS") and "CONTINUOUS" not in key:
                section_name = key.removesuffix("_HOURS")
                if section_name not in section_names:
                    section_names.append(section_name)

    worksheet.cell(row=1, column=1, value="YEAR")
    worksheet.cell(row=1, column=2, value="SUBJECT")
    worksheet.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    worksheet.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

    current_column = 3
    for section_name in section_names:
        worksheet.cell(row=1, column=current_column, value=section_name)
        worksheet.merge_cells(start_row=1, start_column=current_column, end_row=1, end_column=current_column + 2)
        worksheet.cell(row=2, column=current_column, value="NO OF HOURS")
        worksheet.cell(row=2, column=current_column + 1, value="FACULTY-ID")
        worksheet.cell(row=2, column=current_column + 2, value="CONTINUOUS HOURS")
        current_column += 3

    if include_example_rows:
        for row_index, row in enumerate(normalized_records, start=3):
            worksheet.cell(row=row_index, column=1, value=row.get("YEAR"))
            worksheet.cell(row=row_index, column=2, value=row.get("SUBJECT_ID") or row.get("SUBJECT"))

            current_column = 3
            for section_name in section_names:
                worksheet.cell(row=row_index, column=current_column, value=row.get(f"{section_name}_HOURS"))
                worksheet.cell(row=row_index, column=current_column + 1, value=row.get(f"{section_name}_FACULTY_ID"))
                worksheet.cell(row=row_index, column=current_column + 2, value=row.get(f"{section_name}_CONTINUOUS_HOURS"))
                current_column += 3

    for row in worksheet.iter_rows():
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    for column_index in range(1, worksheet.max_column + 1):
        max_length = 0
        for row_index in range(1, worksheet.max_row + 1):
            value = worksheet.cell(row=row_index, column=column_index).value
            max_length = max(max_length, len(str(value)) if value is not None else 0)
        worksheet.column_dimensions[get_column_letter(column_index)].width = max(max_length + 2, 14)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.read()


def create_master_workbook_template(
    template_type: str,
    subjects_records: list[dict],
    faculty_records: list[dict],
    constraints_records: list[dict],
    labs_records: list[dict],
    shared_classes_records: list[dict],
    sessions_records: list[dict],
    classrooms_records: list[dict],
    faculty_availability_records: list[dict],
    continuous_rules_records: list[dict],
    fixed_classroom_blocks_records: list[dict],
) -> bytes:
    workbook = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    include_rows = (template_type == "example")

    # Helper to write standard sheets
    def write_standard_sheet(sheet_name: str, headers: list[str], records: list[dict]):
        ws = workbook.create_sheet(title=sheet_name)
        # Write headers
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)
        
        # Write records if include_rows
        if include_rows:
            for row_idx, r in enumerate(records, 2):
                for col_idx, h in enumerate(headers, 1):
                    ws.cell(row=row_idx, column=col_idx, value=r.get(h))

        # Formatting
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        for column_index in range(1, ws.max_column + 1):
            max_length = 0
            for row_index in range(1, ws.max_row + 1):
                value = ws.cell(row=row_index, column=column_index).value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            ws.column_dimensions[get_column_letter(column_index)].width = max(max_length + 2, 14)

    # Helper to write Constraints sheet (special grouped header layout)
    def write_constraints_sheet(sheet_name: str, records: list[dict]):
        ws = workbook.create_sheet(title=sheet_name)
        
        normalized_records = [{str(key).strip().upper(): value for key, value in row.items()} for row in records]
        section_names: list[str] = []
        for r in normalized_records:
            for key in r.keys():
                if key.endswith("_HOURS") and "CONTINUOUS" not in key:
                    section_name = key.removesuffix("_HOURS")
                    if section_name not in section_names:
                        section_names.append(section_name)

        if not section_names:
            section_names = ["C1", "C2", "C3"]

        ws.cell(row=1, column=1, value="YEAR")
        ws.cell(row=1, column=2, value="SUBJECT")
        ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
        ws.merge_cells(start_row=1, start_column=2, end_row=2, end_column=2)

        current_column = 3
        for section_name in section_names:
            ws.cell(row=1, column=current_column, value=section_name)
            ws.merge_cells(start_row=1, start_column=current_column, end_row=1, end_column=current_column + 2)
            ws.cell(row=2, column=current_column, value="NO OF HOURS")
            ws.cell(row=2, column=current_column + 1, value="FACULTY-ID")
            ws.cell(row=2, column=current_column + 2, value="CONTINUOUS HOURS")
            current_column += 3

        if include_rows:
            for row_idx, r in enumerate(normalized_records, start=3):
                ws.cell(row=row_idx, column=1, value=r.get("YEAR"))
                ws.cell(row=row_idx, column=2, value=r.get("SUBJECT_ID") or r.get("SUBJECT"))

                current_column = 3
                for section_name in section_names:
                    ws.cell(row=row_idx, column=current_column, value=r.get(f"{section_name}_HOURS"))
                    ws.cell(row=row_idx, column=current_column + 1, value=r.get(f"{section_name}_FACULTY_ID"))
                    ws.cell(row=row_idx, column=current_column + 2, value=r.get(f"{section_name}_CONTINUOUS_HOURS"))
                    current_column += 3

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

        for column_index in range(1, ws.max_column + 1):
            max_length = 0
            for row_index in range(1, ws.max_row + 1):
                value = ws.cell(row=row_index, column=column_index).value
                max_length = max(max_length, len(str(value)) if value is not None else 0)
            ws.column_dimensions[get_column_letter(column_index)].width = max(max_length + 2, 14)

    # 1. Subjects
    write_standard_sheet(
        "Subjects",
        ["SUBJECT_ID", "SUBJECT_NAME"],
        subjects_records,
    )
    # 2. Faculty Mapping
    write_standard_sheet(
        "Faculty Mapping",
        ["faculty name", "id assigned"],
        faculty_records,
    )
    # 3. Constraints
    write_constraints_sheet(
        "Constraints",
        constraints_records,
    )
    # 4. Labs
    write_standard_sheet(
        "Labs",
        ["YEAR", "SECTION", "SUBJECT_ID", "DAY", "HOURS", "VENUE"],
        labs_records,
    )
    # 5. Shared Classes
    write_standard_sheet(
        "Shared Classes",
        ["year", "sections", "subject"],
        shared_classes_records,
    )
    # 6. Sessions
    write_standard_sheet(
        "Sessions",
        ["period", "time"],
        sessions_records,
    )
    # 7. Classrooms
    write_standard_sheet(
        "Classrooms",
        ["class_number", "room_type", "capacity", "section", "strength"],
        classrooms_records,
    )
    # 8. Faculty Availability
    write_standard_sheet(
        "Faculty Availability",
        ["Faculty ID", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"],
        faculty_availability_records,
    )
    # 9. Continuous Rules
    write_standard_sheet(
        "Continuous Rules",
        ["SUBJECT_ID", "COMPULSORY_CONTINUOUS_HOURS"],
        continuous_rules_records,
    )
    # 10. Fixed Classroom Blocks
    write_standard_sheet(
        "Fixed Classroom Blocks",
        ["year", "section", "day", "periods", "classroom"],
        fixed_classroom_blocks_records,
    )

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream.read()
